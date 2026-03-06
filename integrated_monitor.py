"""
完全集成的链接监测工具
使用Python Playwright库直接控制浏览器
"""

import csv
import os
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from PIL import Image as PILImage
from playwright.sync_api import sync_playwright
import time

def read_links_from_csv(csv_file):
    """从CSV文件读取链接"""
    links = []
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
        for row in rows[1:]:
            if row and len(row) > 0 and row[0].strip():
                links.append(row[0].strip())
    return links

def resize_screenshot(image_path, max_width=300, max_height=200):
    """调整截图大小"""
    try:
        img = PILImage.open(image_path)
        width_ratio = max_width / img.width
        height_ratio = max_height / img.height
        ratio = min(width_ratio, height_ratio)
        new_width = int(img.width * ratio)
        new_height = int(img.height * ratio)
        return new_width, new_height
    except:
        return max_width, max_height

def create_excel_report(results, output_file):
    """创建Excel报告"""
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "链接监测结果"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 20
    
    # 标题行
    headers = ['序号', '链接', '状态', '截屏预览', '检测时间']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 25
    
    # 数据行
    current_row = 2
    for idx, result in enumerate(results, 1):
        ws.cell(row=current_row, column=1, value=idx).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=current_row, column=1).border = thin_border
        
        ws.cell(row=current_row, column=2, value=result['链接']).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=current_row, column=2).border = thin_border
        
        status_cell = ws.cell(row=current_row, column=3, value=result['状态'])
        if result['状态'] == '成功':
            status_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            status_cell.font = Font(color="FFFFFF", bold=True)
        else:
            status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            status_cell.font = Font(color="FFFFFF", bold=True)
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        status_cell.border = thin_border
        
        screenshot_path = result.get('截屏文件', '')
        cell_d = ws.cell(row=current_row, column=4)
        cell_d.border = thin_border
        
        if screenshot_path and os.path.exists(screenshot_path):
            try:
                width, height = resize_screenshot(screenshot_path, max_width=300, max_height=180)
                ws.row_dimensions[current_row].height = height * 0.75
                img = XLImage(screenshot_path)
                img.width = width
                img.height = height
                img.anchor = f'D{current_row}'
                ws.add_image(img)
            except Exception as e:
                cell_d.value = f'图片加载失败'
                ws.row_dimensions[current_row].height = 30
        else:
            cell_d.value = result.get('错误信息', '无截图')
            cell_d.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            ws.row_dimensions[current_row].height = 30
        
        cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.cell(row=current_row, column=5, value=result['检测时间']).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=current_row, column=5).border = thin_border
        
        current_row += 1
    
    # 统计
    stats_row = current_row + 1
    success_count = sum(1 for r in results if r['状态'] == '成功')
    fail_count = sum(1 for r in results if r['状态'] == '失败')
    
    ws.cell(row=stats_row, column=1, value='统计').font = Font(bold=True, size=12)
    ws.cell(row=stats_row, column=2, value=f'总计: {len(results)} | 成功: {success_count} | 失败: {fail_count}')
    ws.cell(row=stats_row, column=2).font = Font(bold=True, size=11)
    
    wb.save(output_file)
    return success_count, fail_count

def capture_screenshots_with_playwright(links, screenshot_dir):
    """使用Playwright捕获截图"""
    results = []
    
    print("\n正在启动浏览器...")
    print("提示: 浏览器会自动打开，如需登录请在首次访问时完成登录")
    print("-" * 70)
    
    with sync_playwright() as p:
        # 启动浏览器
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(viewport={'width': 1920, 'height': 1080})
        page = context.new_page()
        
        # 先访问主页，给用户登录机会
        print("\n[准备] 正在打开小红书主页，如需登录请现在登录...")
        page.goto('https://www.xiaohongshu.com/')
        print("等待10秒，请完成登录（如已登录可忽略）...")
        time.sleep(10)
        
        # 访问每个链接
        for idx, link in enumerate(links, 1):
            print(f"\n[{idx}/{len(links)}] 正在访问链接...")
            print(f"  URL: {link}")
            
            result = {
                '链接': link,
                '状态': '失败',
                '截屏文件': '',
                '错误信息': '',
                '检测时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            try:
                # 访问链接
                page.goto(link, timeout=30000)
                print(f"  等待页面加载...")
                time.sleep(3)
                
                # 截图
                screenshot_name = f'screenshot_{idx}.png'
                screenshot_path = os.path.join(screenshot_dir, screenshot_name)
                page.screenshot(path=screenshot_path)
                
                result['状态'] = '成功'
                result['截屏文件'] = screenshot_path
                print(f"  [OK] 截图已保存: {screenshot_name}")
                
            except Exception as e:
                result['错误信息'] = f'访问失败: {str(e)}'
                print(f"  [FAIL] 错误: {str(e)}")
            
            results.append(result)
        
        # 关闭浏览器
        browser.close()
        print("\n浏览器已关闭")
    
    return results

def main():
    print("\n" + "="*70)
    print("            链接监测工具 - 完全自动化版本")
    print("="*70)
    
    csv_file = 'yilideeplink.csv'
    output_dir = 'output'
    screenshot_dir = os.path.join(output_dir, 'screenshots')
    
    # 创建目录
    os.makedirs(screenshot_dir, exist_ok=True)
    
    # 读取链接
    print(f"\n[步骤 1/3] 读取CSV文件: {csv_file}")
    links = read_links_from_csv(csv_file)
    print(f"           找到 {len(links)} 个链接")
    
    # 显示链接
    print("\n链接列表:")
    for idx, link in enumerate(links, 1):
        print(f"  {idx}. {link[:80]}...")
    
    # 使用Playwright截图
    print(f"\n[步骤 2/3] 使用浏览器访问链接并截图")
    results = capture_screenshots_with_playwright(links, screenshot_dir)
    
    # 生成报告
    print(f"\n[步骤 3/3] 生成Excel报告")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f'链接监测报告_{timestamp}.xlsx')
    
    success_count, fail_count = create_excel_report(results, output_file)
    
    # 总结
    print("\n" + "="*70)
    print("                    监测完成")
    print("="*70)
    print(f"\n总计: {len(results)} 个链接")
    print(f"成功: {success_count} 个")
    print(f"失败: {fail_count} 个")
    print(f"\n报告: {output_file}")
    
    if fail_count > 0:
        print("\n失败链接:")
        for idx, result in enumerate(results, 1):
            if result['状态'] == '失败':
                print(f"  {idx}. {result['错误信息']}")
    
    print("\n" + "="*70)
    print("提示: Excel报告已生成，包含所有截图")
    print("="*70 + "\n")

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n用户中断程序")
    except Exception as e:
        print(f"\n错误: {str(e)}")
        print("请确保已安装: pip install playwright")
        print("并执行: playwright install chromium")








