"""
WPP MD 小红书链接监测工具
终极版本 - 蓝紫渐变风格
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import csv
import os
import threading
import time
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image as PILImage
from playwright.sync_api import sync_playwright

# 设置Playwright浏览器路径（支持打包后的EXE）
# 优先使用用户目录下的浏览器
if not os.environ.get('PLAYWRIGHT_BROWSERS_PATH'):
    playwright_path = os.path.join(os.path.expanduser('~'), '.playwright-browsers')
    if os.path.exists(playwright_path):
        os.environ['PLAYWRIGHT_BROWSERS_PATH'] = playwright_path

class UltimateLinkMonitorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("WPP MD 小红书链接监测")
        self.root.geometry("900x750")
        
        # 蓝紫渐变配色
        self.colors = {
            'bg': '#5B6FB5',
            'primary': '#5B7BF5',
            'secondary': '#6B5FB5',
            'white': '#FFFFFF',
            'text_light': '#E8EEFF',
            'text_dark': '#2D3748',
            'success': '#48BB78',
            'danger': '#F56565',
        }
        
        # 变量
        self.csv_file = tk.StringVar(value="yilideeplink.csv")
        self.output_dir = tk.StringVar(value="./output")
        self.schedule_enabled = tk.BooleanVar(value=False)
        self.schedule_time = tk.StringVar(value="09:00")
        self.login_wait_time = tk.StringVar(value="30")  # 扫码等待时间（秒）
        self.is_running = False
        self.stop_flag = False
        
        self.root.configure(bg=self.colors['bg'])
        self.create_ui()
        
    def create_ui(self):
        # 主容器 - 使用Canvas实现滚动
        main_canvas = tk.Canvas(self.root, bg=self.colors['bg'], highlightthickness=0)
        scrollbar = tk.Scrollbar(self.root, orient="vertical", command=main_canvas.yview)
        scrollable_frame = tk.Frame(main_canvas, bg=self.colors['bg'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        
        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=880)
        main_canvas.configure(yscrollcommand=scrollbar.set)
        
        main_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 内容区
        content = tk.Frame(scrollable_frame, bg=self.colors['bg'])
        content.pack(fill=tk.BOTH, expand=True, padx=25, pady=25)
        
        # ===== 1. 标题区 =====
        title_frame = tk.Frame(content, bg=self.colors['bg'])
        title_frame.pack(pady=(0, 20))
        
        # Logo
        logo_canvas = tk.Canvas(title_frame, width=70, height=70, bg=self.colors['bg'], highlightthickness=0)
        logo_canvas.pack()
        logo_canvas.create_oval(10, 10, 60, 60, fill=self.colors['primary'], outline='')
        logo_canvas.create_oval(22, 22, 48, 48, fill=self.colors['white'], outline='')
        logo_canvas.create_oval(30, 30, 40, 40, fill=self.colors['primary'], outline='')
        
        tk.Label(
            title_frame,
            text="WPP MD 小红书链接监测",
            font=('Microsoft YaHei UI', 20, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['white']
        ).pack(pady=(10, 2))
        
        tk.Label(
            title_frame,
            text="Professional Link Monitoring Tool",
            font=('Arial', 9),
            bg=self.colors['bg'],
            fg=self.colors['text_light']
        ).pack()
        
        # ===== 2. 文件配置卡片 =====
        file_card = self.create_card(content, "📁 文件配置")
        
        # CSV文件
        csv_row = tk.Frame(file_card, bg='white')
        csv_row.pack(fill=tk.X, pady=5)
        
        tk.Label(csv_row, text="输入CSV文件:", font=('Microsoft YaHei UI', 9), bg='white', fg=self.colors['text_dark']).pack(side=tk.LEFT)
        
        csv_entry_frame = tk.Frame(csv_row, bg='#F7FAFC')
        csv_entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 5))
        
        tk.Entry(
            csv_entry_frame,
            textvariable=self.csv_file,
            font=('Consolas', 9),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT
        ).pack(fill=tk.X, padx=8, pady=6)
        
        tk.Button(
            csv_row,
            text="浏览...",
            command=self.browse_csv,
            font=('Microsoft YaHei UI', 9),
            bg=self.colors['primary'],
            fg='white',
            relief=tk.FLAT,
            padx=15,
            pady=4,
            cursor='hand2'
        ).pack(side=tk.RIGHT)
        
        # 输出目录
        output_row = tk.Frame(file_card, bg='white')
        output_row.pack(fill=tk.X, pady=5)
        
        tk.Label(output_row, text="输出目录:", font=('Microsoft YaHei UI', 9), bg='white', fg=self.colors['text_dark']).pack(side=tk.LEFT, padx=(0, 15))
        
        output_entry_frame = tk.Frame(output_row, bg='#F7FAFC')
        output_entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 5))
        
        tk.Entry(
            output_entry_frame,
            textvariable=self.output_dir,
            font=('Consolas', 9),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT
        ).pack(fill=tk.X, padx=8, pady=6)
        
        tk.Button(
            output_row,
            text="浏览...",
            command=self.browse_output,
            font=('Microsoft YaHei UI', 9),
            bg=self.colors['primary'],
            fg='white',
            relief=tk.FLAT,
            padx=15,
            pady=4,
            cursor='hand2'
        ).pack(side=tk.RIGHT)
        
        # ===== 3. 定时任务卡片 =====
        schedule_card = self.create_card(content, "⏰ 定时任务")
        
        schedule_row = tk.Frame(schedule_card, bg='white')
        schedule_row.pack(fill=tk.X, pady=5)
        
        self.schedule_check = tk.Checkbutton(
            schedule_row,
            text="启用定时运行",
            variable=self.schedule_enabled,
            command=self.toggle_schedule,
            font=('Microsoft YaHei UI', 9),
            bg='white',
            fg=self.colors['text_dark'],
            activebackground='white',
            selectcolor='white',
            cursor='hand2'
        )
        self.schedule_check.pack(side=tk.LEFT)
        
        tk.Label(schedule_row, text="运行时间:", font=('Microsoft YaHei UI', 9), bg='white', fg=self.colors['text_dark']).pack(side=tk.LEFT, padx=(20, 8))
        
        time_frame = tk.Frame(schedule_row, bg='#F7FAFC')
        time_frame.pack(side=tk.LEFT)
        
        self.time_entry = tk.Entry(
            time_frame,
            textvariable=self.schedule_time,
            font=('Consolas', 9),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            width=8,
            state='disabled'
        )
        self.time_entry.pack(padx=8, pady=6)
        
        tk.Label(schedule_row, text="(格式: HH:MM)", font=('Arial', 8), bg='white', fg='#A0AEC0').pack(side=tk.LEFT, padx=(5, 0))
        
        # ===== 4. 登录设置卡片 =====
        login_card = self.create_card(content, "🔐 登录设置")
        
        login_row = tk.Frame(login_card, bg='white')
        login_row.pack(fill=tk.X, pady=5)
        
        tk.Label(
            login_row,
            text="扫码等待时间:",
            font=('Microsoft YaHei UI', 9),
            bg='white',
            fg=self.colors['text_dark']
        ).pack(side=tk.LEFT)
        
        wait_time_frame = tk.Frame(login_row, bg='#F7FAFC')
        wait_time_frame.pack(side=tk.LEFT, padx=(10, 5))
        
        tk.Entry(
            wait_time_frame,
            textvariable=self.login_wait_time,
            font=('Consolas', 9),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            width=6
        ).pack(padx=8, pady=6)
        
        tk.Label(
            login_row,
            text="秒 (建议30-60秒)",
            font=('Arial', 8),
            bg='white',
            fg='#A0AEC0'
        ).pack(side=tk.LEFT, padx=(5, 0))
        
        # ===== 5. 控制按钮 =====
        button_frame = tk.Frame(content, bg=self.colors['bg'])
        button_frame.pack(pady=15)
        
        self.start_button = tk.Button(
            button_frame,
            text="▶  立即开始监测",
            command=self.start_monitoring,
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            relief=tk.FLAT,
            padx=30,
            pady=12,
            cursor='hand2'
        )
        self.start_button.pack(side=tk.LEFT, padx=5)
        
        self.stop_button = tk.Button(
            button_frame,
            text="⬛  停止",
            command=self.stop_monitoring,
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg=self.colors['danger'],
            fg='white',
            relief=tk.FLAT,
            padx=25,
            pady=12,
            state='disabled',
            cursor='hand2'
        )
        self.stop_button.pack(side=tk.LEFT, padx=5)
        
        # ===== 6. 进度卡片 =====
        progress_card = self.create_card(content, "📊 执行进度")
        
        style = ttk.Style()
        style.theme_use('clam')
        style.configure(
            "Custom.Horizontal.TProgressbar",
            troughcolor='#E2E8F0',
            bordercolor='#E2E8F0',
            background=self.colors['primary'],
            lightcolor=self.colors['primary'],
            darkcolor=self.colors['primary']
        )
        
        self.progress = ttk.Progressbar(
            progress_card,
            mode='determinate',
            style="Custom.Horizontal.TProgressbar",
            length=400
        )
        self.progress.pack(fill=tk.X, pady=(5, 8))
        
        self.progress_label = tk.Label(
            progress_card,
            text="就绪",
            font=('Microsoft YaHei UI', 9),
            bg='white',
            fg=self.colors['text_dark']
        )
        self.progress_label.pack()
        
        # ===== 7. 日志卡片 =====
        log_card = self.create_card(content, "📝 运行日志")
        
        log_toolbar = tk.Frame(log_card, bg='white')
        log_toolbar.pack(fill=tk.X, pady=(0, 5))
        
        tk.Button(
            log_toolbar,
            text="🗑 清空日志",
            command=self.clear_log,
            font=('Microsoft YaHei UI', 8),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            padx=10,
            pady=3,
            cursor='hand2'
        ).pack(side=tk.RIGHT)
        
        self.log_text = scrolledtext.ScrolledText(
            log_card,
            height=12,
            font=('Consolas', 8),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            wrap=tk.WORD,
            state='disabled'
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
        
        # ===== 8. 状态栏 =====
        status_frame = tk.Frame(content, bg='#4B5FA5', height=35)
        status_frame.pack(fill=tk.X, pady=(10, 0))
        status_frame.pack_propagate(False)
        
        self.status_bar = tk.Label(
            status_frame,
            text="● 就绪",
            font=('Microsoft YaHei UI', 9),
            bg='#4B5FA5',
            fg=self.colors['text_light'],
            anchor=tk.W
        )
        self.status_bar.pack(side=tk.LEFT, padx=15, fill=tk.Y)
        
    def create_card(self, parent, title):
        """创建白色卡片"""
        card_outer = tk.Frame(parent, bg='white', relief=tk.FLAT)
        card_outer.pack(fill=tk.X, pady=(0, 12))
        
        # 标题
        title_frame = tk.Frame(card_outer, bg='white')
        title_frame.pack(fill=tk.X, padx=15, pady=(10, 5))
        
        tk.Label(
            title_frame,
            text=title,
            font=('Microsoft YaHei UI', 11, 'bold'),
            bg='white',
            fg=self.colors['text_dark']
        ).pack(side=tk.LEFT)
        
        # 分隔线
        tk.Frame(card_outer, bg='#E2E8F0', height=1).pack(fill=tk.X, padx=15)
        
        # 内容区
        content = tk.Frame(card_outer, bg='white')
        content.pack(fill=tk.BOTH, expand=True, padx=15, pady=(8, 10))
        
        return content
    
    def browse_csv(self):
        filename = filedialog.askopenfilename(
            title="选择CSV文件",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.csv_file.set(filename)
            self.log(f"✓ 已选择CSV: {os.path.basename(filename)}")
    
    def browse_output(self):
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.output_dir.set(directory)
            self.log(f"✓ 已选择输出目录: {directory}")
    
    def toggle_schedule(self):
        if self.schedule_enabled.get():
            self.time_entry.config(state='normal', bg='white')
            self.log("⏰ 定时任务已启用")
        else:
            self.time_entry.config(state='disabled', bg='#F7FAFC')
            self.log("⏰ 定时任务已禁用")
    
    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {message}\n"
        
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, log_message)
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')
        
        self.status_bar.config(text=f"● {message}")
    
    def clear_log(self):
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        self.log("🗑 日志已清空")
    
    def update_progress(self, current, total, message=""):
        progress = (current / total) * 100 if total > 0 else 0
        self.progress['value'] = progress
        self.progress_label.config(text=f"{message} ({current}/{total}) • {progress:.0f}%")
        self.root.update_idletasks()
    
    def read_links_from_csv(self, csv_file):
        links = []
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
            for row in rows[1:]:
                if row and len(row) > 0 and row[0].strip():
                    links.append(row[0].strip())
        return links
    
    def resize_screenshot(self, image_path, max_width=300, max_height=200):
        try:
            img = PILImage.open(image_path)
            width_ratio = max_width / img.width
            height_ratio = max_height / img.height
            ratio = min(width_ratio, height_ratio)
            return int(img.width * ratio), int(img.height * ratio)
        except:
            return max_width, max_height
    
    def create_excel_report(self, results, output_dir):
        os.makedirs(output_dir, exist_ok=True)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "链接监测结果"
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 20
        
        headers = ['序号', '链接', '状态', '截屏预览', '检测时间']
        header_fill = PatternFill(start_color="5B7BF5", end_color="5B7BF5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 25
        
        current_row = 2
        for idx, result in enumerate(results, 1):
            ws.cell(row=current_row, column=1, value=idx).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=1).border = thin_border
            
            ws.cell(row=current_row, column=2, value=result['链接']).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.cell(row=current_row, column=2).border = thin_border
            
            status_cell = ws.cell(row=current_row, column=3, value=result['状态'])
            if result['状态'] == '成功':
                status_cell.fill = PatternFill(start_color="48BB78", end_color="48BB78", fill_type="solid")
                status_cell.font = Font(color="FFFFFF", bold=True)
            else:
                status_cell.fill = PatternFill(start_color="F56565", end_color="F56565", fill_type="solid")
                status_cell.font = Font(color="FFFFFF", bold=True)
            status_cell.alignment = Alignment(horizontal='center', vertical='center')
            status_cell.border = thin_border
            
            screenshot_path = result.get('截屏文件', '')
            cell_d = ws.cell(row=current_row, column=4)
            cell_d.border = thin_border
            
            if screenshot_path and os.path.exists(screenshot_path):
                try:
                    width, height = self.resize_screenshot(screenshot_path, max_width=300, max_height=180)
                    ws.row_dimensions[current_row].height = height * 0.75
                    img = XLImage(screenshot_path)
                    img.width = width
                    img.height = height
                    img.anchor = f'D{current_row}'
                    ws.add_image(img)
                except:
                    cell_d.value = '图片加载失败'
                    ws.row_dimensions[current_row].height = 30
            else:
                cell_d.value = result.get('错误信息', '无截图')
                cell_d.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                ws.row_dimensions[current_row].height = 30
            
            cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            ws.cell(row=current_row, column=5, value=result['检测时间']).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=5).border = thin_border
            
            current_row += 1
        
        stats_row = current_row + 1
        success_count = sum(1 for r in results if r['状态'] == '成功')
        fail_count = sum(1 for r in results if r['状态'] == '失败')
        
        ws.cell(row=stats_row, column=1, value='统计').font = Font(bold=True, size=12)
        ws.cell(row=stats_row, column=2, value=f'总计: {len(results)} | 成功: {success_count} | 失败: {fail_count}')
        ws.cell(row=stats_row, column=2).font = Font(bold=True, size=11)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_dir, f'链接监测报告_{timestamp}.xlsx')
        wb.save(output_file)
        
        return output_file, success_count, fail_count
    
    def capture_with_playwright(self, links, screenshot_dir):
        results = []
        
        self.log("正在启动浏览器...")
        
        with sync_playwright() as p:
            # 尝试启动浏览器，如果失败则给出详细提示
            try:
                browser = p.chromium.launch(headless=False)
            except Exception as e:
                self.log(f"错误: 无法启动浏览器 - {str(e)}")
                self.log("正在尝试自动安装浏览器...")
                
                # 尝试自动安装
                import subprocess
                try:
                    result = subprocess.run(
                        ['playwright', 'install', 'chromium'],
                        capture_output=True,
                        text=True,
                        timeout=300
                    )
                    if result.returncode == 0:
                        self.log("浏览器安装成功，正在重试...")
                        browser = p.chromium.launch(headless=False)
                    else:
                        raise Exception("自动安装失败，请手动运行: playwright install chromium")
                except Exception as install_error:
                    raise Exception(f"无法安装浏览器: {str(install_error)}\n请手动在命令行运行: playwright install chromium")
            
            # 创建持久化的浏览器上下文，保持登录状态
            context = browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            )
            page = context.new_page()
            
            self.log("正在打开小红书主页...")
            try:
                page.goto('https://www.xiaohongshu.com/', wait_until='domcontentloaded', timeout=30000)
            except Exception:
                page.goto('https://www.xiaohongshu.com/', timeout=30000)
            time.sleep(3)
            
            # 获取用户设置的等待时间
            try:
                wait_seconds = int(self.login_wait_time.get())
                if wait_seconds < 5:
                    wait_seconds = 30
                if wait_seconds > 300:
                    wait_seconds = 300
            except:
                wait_seconds = 30
            
            self.log(f"等待{wait_seconds}秒，请完成扫码登录（如已登录可忽略）...")
            
            # 倒计时显示，每5秒检测一次登录状态
            login_success = False
            for remaining in range(wait_seconds, 0, -1):
                if self.stop_flag:
                    break
                
                if remaining % 5 == 0 or remaining <= 5:
                    self.log(f"   剩余 {remaining} 秒...")
                
                # 每5秒检测一次登录状态
                if remaining % 5 == 0:
                    try:
                        is_logged_in = page.evaluate("""() => {
                            // 检查多个登录标志
                            const avatar = document.querySelector('.avatar, .user-avatar, [class*="Avatar"]');
                            const userInfo = document.querySelector('.user-info, .login-info, [class*="UserInfo"]');
                            const loginBtn = document.querySelector('.login-btn, [class*="login-button"]');
                            
                            // 检查关键Cookie
                            const hasCookie = document.cookie.includes('web_session') || 
                                            document.cookie.includes('xsecappid') || 
                                            document.cookie.includes('a1');
                            
                            // 同时满足：有用户信息 AND 有Cookie AND 没有登录按钮
                            if ((avatar || userInfo) && hasCookie && !loginBtn) {
                                return true;
                            }
                            
                            return false;
                        }""")
                        
                        if is_logged_in:
                            login_success = True
                            self.log(f"✓ 检测到登录成功！")
                            time.sleep(2)  # 额外等待确保会话建立
                            break
                    except Exception:
                        pass
                
                time.sleep(1)
            
            if not self.stop_flag:
                if login_success:
                    self.log("✓ 登录成功，开始访问链接...")
                else:
                    self.log("✗ 未检测到登录成功！")
                    self.log("✗ 请确保扫码登录完成，否则每个链接都会弹出登录窗口")
                    self.log("等待10秒补救，请立即扫码...")
                    for i in range(10, 0, -1):
                        self.log(f"   倒计时 {i} 秒")
                        time.sleep(1)
                        if self.stop_flag:
                            break
            
            for idx, link in enumerate(links, 1):
                if self.stop_flag:
                    self.log("任务已被用户停止")
                    break
                
                self.update_progress(idx, len(links), f"正在访问链接 {idx}")
                self.log(f"[{idx}/{len(links)}] 访问: {link[:60]}...")
                
                result = {
                    '链接': link,
                    '状态': '失败',
                    '截屏文件': '',
                    '错误信息': '',
                    '检测时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                try:
                    # 增加重试机制
                    max_retries = 2
                    for attempt in range(max_retries):
                        try:
                            page.goto(link, timeout=45000, wait_until='domcontentloaded')
                            break
                        except Exception as e:
                            if attempt == max_retries - 1:
                                raise e
                            self.log(f"  重试中... ({attempt + 1}/{max_retries})")
                            time.sleep(2)
                    
                    # 等待页面完全加载
                    time.sleep(3)
                    
                    # 检查是否遇到登录弹窗
                    has_login_popup = page.evaluate("""() => {
                        const loginModal = document.querySelector('.login-modal, [class*="LoginModal"], [class*="login-dialog"]');
                        const qrCode = document.querySelector('[class*="qrcode"], [class*="QRCode"]');
                        return !!(loginModal || qrCode);
                    }""")
                    
                    if has_login_popup:
                        self.log(f"  ⚠ 检测到登录弹窗，请手动登录...")
                        # 等待用户手动登录
                        time.sleep(15)
                    
                    screenshot_name = f'screenshot_{idx}.png'
                    screenshot_path = os.path.join(screenshot_dir, screenshot_name)
                    page.screenshot(path=screenshot_path, full_page=False)
                    
                    result['状态'] = '成功'
                    result['截屏文件'] = screenshot_path
                    self.log(f"  ✓ 截图已保存: {screenshot_name}")
                    
                except Exception as e:
                    result['错误信息'] = f'访问失败: {str(e)}'
                    self.log(f"  ✗ 错误: {str(e)}")
                
                results.append(result)
                
                # 链接之间增加延迟，避免触发反爬
                if idx < len(links):
                    time.sleep(2)
            
            browser.close()
            self.log("浏览器已关闭")
        
        return results
    
    def monitor_task(self):
        try:
            self.is_running = True
            self.stop_flag = False
            self.start_button.config(state='disabled')
            self.stop_button.config(state='normal')
            
            csv_file = self.csv_file.get()
            if not os.path.exists(csv_file):
                messagebox.showerror("错误", f"CSV文件不存在: {csv_file}")
                return
            
            output_dir = self.output_dir.get()
            screenshot_dir = os.path.join(output_dir, 'screenshots')
            os.makedirs(screenshot_dir, exist_ok=True)
            
            self.log("="*50)
            self.log("🚀 开始链接监测任务")
            self.log(f"📁 输入: {csv_file}")
            self.log(f"📁 输出: {output_dir}")
            self.log("="*50)
            
            links = self.read_links_from_csv(csv_file)
            self.log(f"📋 找到 {len(links)} 个链接")
            
            if not links:
                messagebox.showwarning("警告", "CSV文件中没有找到链接")
                return
            
            results = self.capture_with_playwright(links, screenshot_dir)
            
            if not self.stop_flag:
                self.log("📊 正在生成Excel报告...")
                output_file, success_count, fail_count = self.create_excel_report(results, output_dir)
                
                self.log("="*50)
                self.log("✅ 监测任务完成！")
                self.log(f"📊 总计: {len(results)} | 成功: {success_count} | 失败: {fail_count}")
                self.log(f"📄 报告: {output_file}")
                self.log("="*50)
                
                messagebox.showinfo(
                    "完成",
                    f"监测完成！\n\n"
                    f"总计: {len(results)}\n"
                    f"成功: {success_count}\n"
                    f"失败: {fail_count}\n\n"
                    f"报告已保存至:\n{output_file}"
                )
        
        except Exception as e:
            self.log(f"❌ 错误: {str(e)}")
            messagebox.showerror("错误", f"监测过程中发生错误:\n{str(e)}")
        
        finally:
            self.is_running = False
            self.start_button.config(state='normal')
            self.stop_button.config(state='disabled')
            self.update_progress(0, 1, "就绪")
    
    def start_monitoring(self):
        if not self.is_running:
            thread = threading.Thread(target=self.monitor_task, daemon=True)
            thread.start()
    
    def stop_monitoring(self):
        if self.is_running:
            self.stop_flag = True
            self.log("⚠ 正在停止任务...")

def main():
    root = tk.Tk()
    
    # 居中窗口
    root.update_idletasks()
    width = 900
    height = 750
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    app = UltimateLinkMonitorGUI(root)
    root.mainloop()

if __name__ == '__main__':
    main()
