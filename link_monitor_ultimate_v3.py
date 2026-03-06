"""
WPP MD 小红书链接监测工具 - v3.0 终极整合版
整合原程序所有功能 + 反爬优化 + 登录增强 + 修复版功能
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import csv
import os
import threading
import time
import random
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image as PILImage
from playwright.sync_api import sync_playwright
import sys
import json
import re
from urllib.parse import urljoin

# 设置Playwright浏览器路径
def resolve_playwright_browsers_path():
    """解析打包的Playwright浏览器路径"""
    candidates = []
    if getattr(sys, "_MEIPASS", None):
        candidates.append(os.path.join(sys._MEIPASS, "browsers"))
    candidates.append(os.path.join(os.getcwd(), "browsers"))
    candidates.append(os.path.join(os.path.expanduser('~'), '.playwright-browsers'))
    for path in candidates:
        if path and os.path.exists(path):
            return path
    return None

if not os.environ.get('PLAYWRIGHT_BROWSERS_PATH'):
    pw_path = resolve_playwright_browsers_path()
    if pw_path:
        os.environ['PLAYWRIGHT_BROWSERS_PATH'] = pw_path


class UltimateMonitorV3GUI:
    def __init__(self, root):
        self.root = root
        self.root.title("WPP MD 小红书链接监测工具 v3.0 Ultimate")
        self.root.geometry("1100x850")

        # v3.0 终极配色方案
        self.colors = {
            'primary': '#5B6FB5',           # 主色调
            'secondary': '#6B5FB5',         # 次要色
            'accent': '#7B4FB5',            # 强调色
            'success': '#34C759',           # 成功绿
            'warning': '#FF9500',           # 警告橙
            'danger': '#FF3B30',            # 危险红
            'bg': '#F8F9FA',                # 背景色
            'card_bg': '#FFFFFF',          # 卡片背景
            'text_primary': '#2D3748',       # 主文字
            'text_secondary': '#718096',     # 次要文字
            'border': '#E2E8F0',            # 边框色
        }

        # v3.0 配置参数
        self.config = {
            # 优化的时间配置（反爬优化）
            'delays': {
                # 默认偏"均衡"，可按需要再调小（更快）或调大（更稳）
                'link_interval': (1.5, 3.5),       # 链接间随机延迟
                'page_load': (1.2, 3.0),           # 页面加载后等待
                'screenshot': (0.4, 1.2),          # 截图前等待
                'scroll': (0.3, 0.8),              # 滚动操作间隔
            },
            # 多个User-Agent轮换
            'user_agents': [
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36",
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36",
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36",
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0"
            ],
            # 性能优化配置
            'performance': {
                'concurrent_downloads': 3,         # 并发下载数量
                'download_timeout': 30,            # 下载超时时间(秒)
                'cache_enabled': True,             # 启用缓存
                'cache_ttl': 3600,                 # 缓存生存时间(秒)
                'max_memory_cache': 100,           # 最大内存缓存条目
            }
        }

        # 设置整体背景
        self.root.configure(bg=self.colors['bg'])

        # 核心变量
        self.csv_file = tk.StringVar(value="yilideeplink.csv")
        self.output_dir = tk.StringVar(value="./output_v3")
        self.login_wait_time = tk.StringVar(value="60")
        self.schedule_enabled = tk.BooleanVar(value=False)
        self.schedule_time = tk.StringVar(value="09:00")
        self.capture_mode = tk.StringVar(value="full")
        self.all_links_mode = tk.BooleanVar(value=False)
        self.enable_anti_detection = tk.BooleanVar(value=True)  # v4.0 反爬默认启用
        self.login_retries = tk.StringVar(value="3")  # v4.0 默认登录重试次数

        self.is_running = False
        self.stop_flag = False

        # v3.0 增强的登录和存储
        self.storage_state_path = "xhs_v3_storage.json"
        self.user_data_dir = "xhs_v3_profile"
        self.http_user_agent = random.choice(self.config['user_agents'])
        # 保留 Pro v2 粉丝数兜底能力（API/主页）
        self.enable_api_fans = True
        self.enable_homepage_fans = True
        # 粉丝数坐标探针：Chrome 悬浮卡片里“粉丝”行的典型位置（1920x1080 视窗）
        # 格式：(x, y, width, height)
        self.fans_probe_box = (742.59375, 307.8046875, 320, 16.796875)

        # 统计信息
        self.stats = {
            'total_processed': 0,
            'success_count': 0,
            'failed_count': 0,
            'anti_crawler_count': 0,
            'start_time': None
        }

        # v3.0 性能优化组件
        self.download_cache = {}        # 下载缓存
        self.fans_cache = {}           # 粉丝数缓存
        self.download_queue = []       # 下载队列
        self.active_downloads = set()  # 活跃下载集合

        self.create_ultimate_widgets()

    def cache_key(self, url, context=""):
        """生成缓存键"""
        import hashlib
        key_data = f"{url}_{context}".encode('utf-8')
        return hashlib.md5(key_data).hexdigest()[:16]

    def get_from_cache(self, cache_dict, key, ttl=None):
        """从缓存获取数据"""
        if not self.config['performance']['cache_enabled']:
            return None

        ttl = ttl or self.config['performance']['cache_ttl']

        try:
            if key in cache_dict:
                cached_item = cache_dict[key]
                # 检查是否过期
                if time.time() - cached_item['timestamp'] < ttl:
                    return cached_item['data']
                else:
                    # 删除过期缓存
                    del cache_dict[key]
        except Exception as e:
            self.log(f"缓存读取失败: {str(e)}", 'warning')

        return None

    def set_cache(self, cache_dict, key, data):
        """设置缓存数据"""
        if not self.config['performance']['cache_enabled']:
            return

        try:
            cache_dict[key] = {
                'data': data,
                'timestamp': time.time()
            }

            # 清理过大的缓存
            max_size = self.config['performance']['max_memory_cache']
            if len(cache_dict) > max_size:
                # 删除最旧的缓存项
                oldest_key = min(cache_dict.keys(),
                                key=lambda k: cache_dict[k]['timestamp'])
                del cache_dict[oldest_key]
                self.log(f"缓存已满，清理最旧项", 'debug')

        except Exception as e:
            self.log(f"缓存写入失败: {str(e)}", 'warning')

    def concurrent_download_images(self, page, image_urls, output_dir, name_prefix, max_workers=3):
        """并发下载图片"""
        import threading
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import queue

        if not image_urls:
            return []

        self.log(f"开始并发下载 {len(image_urls)} 张图片，工作线程: {max_workers}", 'info')

        results = []
        seen_urls = set()

        def download_single_image(url_data):
            """下载单张图片"""
            url, index = url_data
            if url in seen_urls:
                return None

            try:
                # 检查缓存
                cache_key = self.cache_key(url, "image")
                cached_result = self.get_from_cache(self.download_cache, cache_key)
                if cached_result:
                    self.log(f"    图片 {index} 使用缓存", 'info')
                    return cached_result

                # 下载图片
                cover_name = f"{name_prefix}_cover_{index:02d}.png"
                cover_path = os.path.join(output_dir, cover_name)

                if self.download_image_via_page(page, url, cover_path):
                    if os.path.exists(cover_path) and self.is_large_image_file(cover_path):
                        result = cover_path
                        # 缓存成功结果
                        self.set_cache(self.download_cache, cache_key, result)
                        seen_urls.add(url)
                        return result
                    else:
                        try:
                            if os.path.exists(cover_path):
                                os.remove(cover_path)
                        except Exception:
                            pass

            except Exception as e:
                self.log(f"    图片 {index} 下载失败: {str(e)}", 'error')

            return None

        # 使用线程池并发下载
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 准备下载任务
            download_tasks = [(url, i + 1) for i, url in enumerate(image_urls[:10])]  # 限制最多10张

            # 提交所有任务
            future_to_index = {
                executor.submit(download_single_image, task): task[1]
                for task in download_tasks
            }

            # 收集结果
            for future in as_completed(future_to_index):
                try:
                    result = future.result(timeout=self.config['performance']['download_timeout'])
                    if result:
                        results.append(result)
                        self.log(f"    图片下载完成: {os.path.basename(result)}", 'success')
                except Exception as e:
                    index = future_to_index[future]
                    self.log(f"    图片 {index} 下载超时或失败: {str(e)}", 'error')

        self.log(f"并发下载完成，成功 {len(results)}/{len(image_urls)} 张", 'success')
        return results

    def batch_download_with_progress(self, page, items, callback=None):
        """批量下载带进度回调"""
        total = len(items)
        if total == 0:
            return []

        results = []
        batch_size = self.config['performance']['concurrent_downloads']

        for i in range(0, total, batch_size):
            batch = items[i:i + batch_size]
            batch_num = i // batch_size + 1
            total_batches = (total + batch_size - 1) // batch_size

            self.log(f"处理批次 {batch_num}/{total_batches} ({len(batch)} 项)", 'info')

            # 处理当前批次
            batch_results = self.concurrent_download_images(
                page, batch, "", f"batch_{batch_num}"
            )

            results.extend([r for r in batch_results if r])

            # 进度回调
            if callback:
                progress = min(i + len(batch), total)
                callback(progress, total, f"批次 {batch_num}/{total_batches}")

            # 防止过于频繁的请求
            if batch_num < total_batches:
                delay = self.get_random_delay('link_interval')
                time.sleep(delay)

        return results

    def create_ultimate_widgets(self):
        # ===== 顶部标题栏 =====
        header_frame = tk.Frame(self.root, bg=self.colors['primary'], height=90)
        header_frame.pack(fill=tk.X, side=tk.TOP)
        header_frame.pack_propagate(False)

        # Logo区域
        logo_canvas = tk.Canvas(header_frame, width=60, height=60, bg=self.colors['primary'], highlightthickness=0)
        logo_canvas.place(x=30, y=15)
        logo_canvas.create_oval(10, 10, 50, 50, fill=self.colors['card_bg'], outline='')
        logo_canvas.create_oval(20, 20, 40, 40, fill=self.colors['primary'], outline='')

        # 标题区域
        title_container = tk.Frame(header_frame, bg=self.colors['primary'])
        title_container.pack(side=tk.LEFT, padx=(100, 0), expand=True, fill=tk.BOTH)

        title_label = tk.Label(
            title_container,
            text="WPP MD 小红书链接监测工具 v3.0",
            font=('Microsoft YaHei UI', 22, 'bold'),
            bg=self.colors['primary'],
            fg=self.colors['card_bg']
        )
        title_label.pack(side=tk.TOP, pady=(20, 2))

        subtitle_label = tk.Label(
            title_container,
            text="Ultimate Anti-Detection Version | 智能反爬监测 | 完整功能整合",
            font=('Microsoft YaHei UI', 10),
            bg=self.colors['primary'],
            fg=self.colors['accent']
        )
        subtitle_label.pack(side=tk.TOP)

        # ===== 主容器 =====
        main_container = tk.Frame(self.root, bg=self.colors['bg'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=25, pady=20)

        # ===== 文件配置卡片 =====
        file_card = self.create_ultimate_card(main_container, "📁 文件配置")
        file_card.pack(fill=tk.X, pady=(0, 15))

        # CSV文件配置
        csv_frame = tk.Frame(file_card, bg=self.colors['card_bg'])
        csv_frame.pack(fill=tk.X, pady=10, padx=15)

        csv_label = tk.Label(csv_frame, text="输入CSV文件", font=('Microsoft YaHei UI', 11, 'bold'),
                           bg=self.colors['card_bg'], fg=self.colors['text_primary'])
        csv_label.pack(anchor=tk.W)

        csv_input_frame = tk.Frame(csv_frame, bg=self.colors['bg'])
        csv_input_frame.pack(fill=tk.X, pady=(5, 0))

        csv_entry = tk.Entry(csv_input_frame, textvariable=self.csv_file, font=('Consolas', 10),
                           bg=self.colors['bg'], fg=self.colors['text_primary'], relief=tk.FLAT)
        csv_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        csv_browse_btn = tk.Button(csv_input_frame, text="浏览", command=self.browse_csv,
                                  font=('Microsoft YaHei UI', 9, 'bold'),
                                  bg=self.colors['primary'], fg=self.colors['card_bg'],
                                  relief=tk.FLAT, padx=20, pady=6, cursor='hand2')
        csv_browse_btn.pack(side=tk.RIGHT)

        # 输出目录配置
        output_frame = tk.Frame(file_card, bg=self.colors['card_bg'])
        output_frame.pack(fill=tk.X, pady=10, padx=15)

        output_label = tk.Label(output_frame, text="输出目录", font=('Microsoft YaHei UI', 11, 'bold'),
                              bg=self.colors['card_bg'], fg=self.colors['text_primary'])
        output_label.pack(anchor=tk.W)

        output_input_frame = tk.Frame(output_frame, bg=self.colors['bg'])
        output_input_frame.pack(fill=tk.X, pady=(5, 0))

        output_entry = tk.Entry(output_input_frame, textvariable=self.output_dir, font=('Consolas', 10),
                               bg=self.colors['bg'], fg=self.colors['text_primary'], relief=tk.FLAT)
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        output_browse_btn = tk.Button(output_input_frame, text="浏览", command=self.browse_output,
                                     font=('Microsoft YaHei UI', 9, 'bold'),
                                     bg=self.colors['primary'], fg=self.colors['card_bg'],
                                     relief=tk.FLAT, padx=20, pady=6, cursor='hand2')
        output_browse_btn.pack(side=tk.RIGHT)

        # ===== 高级配置卡片 =====
        advanced_card = self.create_ultimate_card(main_container, "⚙️ 高级配置")
        advanced_card.pack(fill=tk.X, pady=(0, 15))

        config_frame = tk.Frame(advanced_card, bg=self.colors['card_bg'])
        config_frame.pack(fill=tk.X, pady=15, padx=15)

        # 左列配置
        left_column = tk.Frame(config_frame, bg=self.colors['card_bg'])
        left_column.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # 登录设置
        login_frame = tk.Frame(left_column, bg=self.colors['card_bg'])
        login_frame.pack(fill=tk.X, pady=5)

        login_label = tk.Label(login_frame, text="登录等待时间(秒):", font=('Microsoft YaHei UI', 10),
                             bg=self.colors['card_bg'], fg=self.colors['text_primary'])
        login_label.pack(side=tk.LEFT)

        login_spinbox = tk.Spinbox(login_frame, from_=10, to=300, textvariable=self.login_wait_time,
                                   width=8, font=('Consolas', 10))
        login_spinbox.pack(side=tk.LEFT, padx=(10, 0))

        login_note = tk.Label(login_frame, text="(建议60-90秒)", font=('Microsoft YaHei UI', 8),
                            bg=self.colors['card_bg'], fg=self.colors['text_secondary'])
        login_note.pack(side=tk.LEFT, padx=(5, 0))

        # 登录重试次数
        retry_frame = tk.Frame(left_column, bg=self.colors['card_bg'])
        retry_frame.pack(fill=tk.X, pady=5)

        retry_label = tk.Label(retry_frame, text="登录重试次数:", font=('Microsoft YaHei UI', 10),
                             bg=self.colors['card_bg'], fg=self.colors['text_primary'])
        retry_label.pack(side=tk.LEFT)

        retry_spinbox = tk.Spinbox(retry_frame, from_=1, to=10, textvariable=self.login_retries,
                                   width=8, font=('Consolas', 10))
        retry_spinbox.pack(side=tk.LEFT, padx=(10, 0))

        # 右列配置
        right_column = tk.Frame(config_frame, bg=self.colors['card_bg'])
        right_column.pack(side=tk.RIGHT, fill=tk.X, expand=True)

        # 模式选择
        mode_frame = tk.Frame(right_column, bg=self.colors['card_bg'])
        mode_frame.pack(fill=tk.X, pady=5)

        mode_label = tk.Label(mode_frame, text="抓取模式:", font=('Microsoft YaHei UI', 10),
                             bg=self.colors['card_bg'], fg=self.colors['text_primary'])
        mode_label.pack(anchor=tk.W)

        mode_options = tk.Frame(mode_frame, bg=self.colors['card_bg'])
        mode_options.pack(fill=tk.X, pady=(3, 0))

        tk.Radiobutton(mode_options, text="快速模式(仅截图)", variable=self.capture_mode, value="simple",
                      font=('Microsoft YaHei UI', 9), bg=self.colors['card_bg'], fg=self.colors['text_primary'],
                      activebackground=self.colors['card_bg'], selectcolor=self.colors['card_bg']).pack(side=tk.LEFT, padx=(0, 15))

        tk.Radiobutton(mode_options, text="完整模式(深度抓取)", variable=self.capture_mode, value="full",
                      font=('Microsoft YaHei UI', 9, 'bold'), bg=self.colors['card_bg'], fg=self.colors['primary'],
                      activebackground=self.colors['card_bg'], selectcolor=self.colors['card_bg']).pack(side=tk.LEFT)

        # v3.0 反爬检测开关
        anti_detection_frame = tk.Frame(right_column, bg=self.colors['card_bg'])
        anti_detection_frame.pack(fill=tk.X, pady=8)

        anti_check = tk.Checkbutton(anti_detection_frame, text="启用智能反爬检测", variable=self.enable_anti_detection,
                                  font=('Microsoft YaHei UI', 10, 'bold'), bg=self.colors['card_bg'], fg=self.colors['success'],
                                  activebackground=self.colors['card_bg'], selectcolor=self.colors['card_bg'])
        anti_check.pack(side=tk.LEFT)

        anti_note = tk.Label(anti_detection_frame, text="(v3.0新功能)", font=('Microsoft YaHei UI', 8),
                            bg=self.colors['card_bg'], fg=self.colors['text_secondary'])
        anti_note.pack(side=tk.LEFT, padx=(5, 0))

        # ===== 控制按钮区域 =====
        control_frame = tk.Frame(main_container, bg=self.colors['bg'])
        control_frame.pack(fill=tk.X, pady=15)

        self.start_button = tk.Button(control_frame, text="▶ 开始监测", command=self.start_monitoring,
                                     font=('Microsoft YaHei UI', 12, 'bold'),
                                     bg=self.colors['success'], fg=self.colors['card_bg'],
                                     relief=tk.FLAT, padx=30, pady=12, cursor='hand2')
        self.start_button.pack(side=tk.LEFT, padx=5)

        self.stop_button = tk.Button(control_frame, text="■ 停止", command=self.stop_monitoring,
                                    font=('Microsoft YaHei UI', 12, 'bold'),
                                    bg=self.colors['danger'], fg=self.colors['card_bg'],
                                    relief=tk.FLAT, padx=25, pady=12, state='disabled', cursor='hand2')
        self.stop_button.pack(side=tk.LEFT, padx=5)

        # ===== 进度区域 =====
        progress_card = self.create_ultimate_card(main_container, "📊 执行进度")
        progress_card.pack(fill=tk.X, pady=(0, 15))

        progress_inner = tk.Frame(progress_card, bg=self.colors['card_bg'])
        progress_inner.pack(fill=tk.X, padx=15, pady=15)

        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Ultimate.Horizontal.TProgressbar",
                       troughcolor=self.colors['border'], bordercolor=self.colors['border'],
                       background=self.colors['success'], lightcolor=self.colors['success'],
                       darkcolor=self.colors['success'])

        self.progress = ttk.Progressbar(progress_inner, mode='determinate',
                                        style="Ultimate.Horizontal.TProgressbar", length=600)
        self.progress.pack(fill=tk.X, pady=(0, 8))

        self.progress_label = tk.Label(progress_inner, text="就绪 - v3.0 Ultimate已就绪",
                                      font=('Microsoft YaHei UI', 10), bg=self.colors['card_bg'],
                                      fg=self.colors['text_secondary'])
        self.progress_label.pack()

        # ===== 统计信息区域 =====
        stats_card = self.create_ultimate_card(main_container, "📈 实时统计")
        stats_card.pack(fill=tk.X, pady=(0, 15))

        stats_inner = tk.Frame(stats_card, bg=self.colors['card_bg'])
        stats_inner.pack(fill=tk.X, padx=15, pady=10)

        # 创建统计标签
        self.stats_labels = {}
        stats_config = [
            ('total', '总链接数:', '0'),
            ('success', '成功数:', '0'),
            ('failed', '失败数:', '0'),
            ('anti_crawler', '反爬触发:', '0'),
        ]

        for i, (key, label, default) in enumerate(stats_config):
            stat_frame = tk.Frame(stats_inner, bg=self.colors['card_bg'])
            stat_frame.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)

            label_widget = tk.Label(stat_frame, text=label, font=('Microsoft YaHei UI', 9, 'bold'),
                                  bg=self.colors['card_bg'], fg=self.colors['text_secondary'])
            label_widget.pack()

            value_widget = tk.Label(stat_frame, text=default, font=('Microsoft YaHei UI', 12, 'bold'),
                                    bg=self.colors['card_bg'], fg=self.colors['text_primary'])
            value_widget.pack()

            self.stats_labels[key] = value_widget

        # ===== 日志区域 =====
        log_card = self.create_ultimate_card(main_container, "📝 运行日志")
        log_card.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        log_inner = tk.Frame(log_card, bg=self.colors['card_bg'])
        log_inner.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)

        # 日志工具栏
        log_toolbar = tk.Frame(log_inner, bg=self.colors['card_bg'])
        log_toolbar.pack(fill=tk.X, pady=(0, 5))

        tk.Button(log_toolbar, text="清空日志", command=self.clear_log,
                 font=('Microsoft YaHei UI', 8), bg=self.colors['bg'], fg=self.colors['text_primary'],
                 relief=tk.FLAT, padx=10, pady=3, cursor='hand2').pack(side=tk.RIGHT)

        # 日志文本框
        self.log_text = scrolledtext.ScrolledText(log_inner, height=12, font=('Consolas', 8),
                                                 bg=self.colors['bg'], fg=self.colors['text_primary'],
                                                 relief=tk.FLAT, wrap=tk.WORD, state='disabled')
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # ===== 状态栏 =====
        status_frame = tk.Frame(main_container, bg=self.colors['primary'], height=35)
        status_frame.pack(fill=tk.X, pady=(0, 0))
        status_frame.pack_propagate(False)

        self.status_bar = tk.Label(status_frame, text="● v3.0 Ultimate就绪 | 智能反爬检测已启用",
                                  font=('Microsoft YaHei UI', 9), bg=self.colors['primary'],
                                  fg=self.colors['card_bg'], anchor=tk.W)
        self.status_bar.pack(side=tk.LEFT, padx=15, fill=tk.Y, expand=True)

    def create_ultimate_card(self, parent, title):
        """创建v3.0风格的卡片"""
        card_outer = tk.Frame(parent, bg=self.colors['border'], relief=tk.FLAT)
        card_outer.pack(fill=tk.X)

        card_inner = tk.Frame(card_outer, bg=self.colors['card_bg'])
        card_inner.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        # 标题栏
        title_frame = tk.Frame(card_inner, bg=self.colors['primary'], height=35)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)

        title_label = tk.Label(title_frame, text=title, font=('Microsoft YaHei UI', 11, 'bold'),
                              bg=self.colors['primary'], fg=self.colors['card_bg'])
        title_label.pack(side=tk.LEFT, padx=15, pady=8)

        # 版本标识
        version_label = tk.Label(title_frame, text="v3.0", font=('Microsoft YaHei UI', 8, 'bold'),
                                bg=self.colors['primary'], fg=self.colors['accent'])
        version_label.pack(side=tk.RIGHT, padx=15, pady=8)

        # 内容区域
        content = tk.Frame(card_inner, bg=self.colors['card_bg'])
        content.pack(fill=tk.BOTH, expand=True, padx=15, pady=(15, 10))

        return content

    # ===== 核心功能方法 =====

    def get_random_delay(self, delay_type):
        """获取随机延迟时间"""
        min_delay, max_delay = self.config['delays'][delay_type]
        return random.uniform(min_delay, max_delay)

    def update_stats(self):
        """更新统计信息"""
        key_mapping = {
            'total': 'total_processed',
            'success': 'success_count',
            'failed': 'failed_count',
            'anti_crawler': 'anti_crawler_count',
        }

        for display_key, label in self.stats_labels.items():
            stats_key = key_mapping.get(display_key, display_key)
            count = self.stats.get(stats_key, 0)
            label.config(text=str(count), fg=self.colors['text_primary'])

            # 根据统计数据设置颜色
            if display_key == 'success' and count > 0:
                label.config(fg=self.colors['success'])
            elif display_key == 'failed' and count > 0:
                label.config(fg=self.colors['danger'])
            elif display_key == 'anti_crawler' and count > 0:
                label.config(fg=self.colors['warning'])

    def log(self, message, level='info'):
        """v4.0增强日志记录（包含错误追踪和性能监控）"""
        import traceback
        from datetime import datetime

        timestamp = datetime.now().strftime("%H:%M:%S")

        # 根据级别设置颜色前缀和详细程度
        level_config = {
            'info': {'prefix': '📝', 'show_traceback': False, 'min_level': 0},
            'success': {'prefix': '✅', 'show_traceback': False, 'min_level': 0},
            'warning': {'prefix': '⚠️', 'show_traceback': False, 'min_level': 1},
            'error': {'prefix': '❌', 'show_traceback': True, 'min_level': 2},
            'anti_crawler': {'prefix': '🛡️', 'show_traceback': False, 'min_level': 1},
            'debug': {'prefix': '🔍', 'show_traceback': False, 'min_level': 3}
        }

        config = level_config.get(level, level_config['info'])
        prefix = config['prefix']

        # 构建基础日志消息
        log_message = f"[{timestamp}] {prefix} {message}"

        # 添加调用栈信息（仅对error级别）
        if config['show_traceback']:
            try:
                # 获取调用栈的最后几层（排除本方法和内部调用）
                stack = traceback.extract_stack()[-3:-1]
                if stack:
                    caller_info = f" ({stack[-1].filename.split('/')[-1]}:{stack[-1].lineno})"
                    log_message += caller_info
            except Exception:
                pass

        log_message += "\n"

        # 写入日志文本框
        try:
            self.log_text.config(state='normal')
            self.log_text.insert(tk.END, log_message)
            self.log_text.see(tk.END)
            self.log_text.config(state='disabled')
        except Exception as e:
            # 如果日志系统本身出错，至少输出到控制台
            print(f"日志系统错误: {e}")
            print(log_message)

        # 更新状态栏（简化版本）
        try:
            status_message = f"● {message[:80]}..."
            if len(message) > 80:
                status_message += ".."
            self.status_bar.config(text=status_message)
        except Exception:
            pass

        # 性能和错误统计（可选）
        if level == 'error':
            try:
                # 记录错误统计
                if not hasattr(self, 'error_stats'):
                    self.error_stats = {}
                error_type = message.split(':')[0] if ':' in message else 'unknown'
                self.error_stats[error_type] = self.error_stats.get(error_type, 0) + 1
            except Exception:
                pass

    def log_with_context(self, message, level='info', context=None):
        """带上下文信息的日志记录"""
        if context:
            context_str = " | ".join([f"{k}:{v}" for k, v in context.items()])
            enhanced_message = f"{message} [{context_str}]"
        else:
            enhanced_message = message
        self.log(enhanced_message, level)

    def safe_execute(self, func, *args, error_context=None, **kwargs):
        """安全执行函数，自动处理异常"""
        try:
            return func(*args, **kwargs)
        except Exception as e:
            error_msg = f"执行失败: {func.__name__} - {str(e)}"
            if error_context:
                self.log_with_context(error_msg, 'error', error_context)
            else:
                self.log(error_msg, 'error')
            return None

    def retry_with_backoff(self, func, max_retries=3, base_delay=1, backoff_factor=2,
                          context=None, retry_on=None):
        """带退避策略的重试机制"""
        import time
        import random

        if retry_on is None:
            retry_on = Exception  # 默认所有异常都重试

        last_exception = None

        for attempt in range(max_retries + 1):  # +1 包含初始尝试
            try:
                if attempt > 0:
                    # 计算延迟时间：基础延迟 * (退避因子 ^ 尝试次数) + 随机抖动
                    delay = base_delay * (backoff_factor ** (attempt - 1))
                    jitter = random.uniform(0.1, 0.5) * delay
                    total_delay = delay + jitter

                    if context:
                        self.log_with_context(
                            f"重试第 {attempt} 次，等待 {total_delay:.1f} 秒",
                            'warning', context
                        )
                    else:
                        self.log(f"重试第 {attempt} 次，等待 {total_delay:.1f} 秒", 'warning')

                    time.sleep(total_delay)

                result = func()

                if attempt > 0:
                    if context:
                        self.log_with_context(f"重试成功 (第{attempt}次)", 'success', context)
                    else:
                        self.log(f"重试成功 (第{attempt}次)", 'success')

                return result

            except retry_on as e:
                last_exception = e

                if attempt < max_retries:
                    if context:
                        self.log_with_context(
                            f"第 {attempt + 1} 次尝试失败: {str(e)}", 'warning', context
                        )
                    else:
                        self.log(f"第 {attempt + 1} 次尝试失败: {str(e)}", 'warning')
                else:
                    if context:
                        self.log_with_context(
                            f"所有 {max_retries + 1} 次尝试均失败", 'error', context
                        )
                    else:
                        self.log(f"所有 {max_retries + 1} 次尝试均失败", 'error')
            except Exception as e:
                # 非预期异常，不重试
                if context:
                    self.log_with_context(
                        f"非预期异常，不重试: {str(e)}", 'error', context
                    )
                else:
                    self.log(f"非预期异常，不重试: {str(e)}", 'error')
                raise e

        # 如果所有重试都失败，抛出最后一个异常
        raise last_exception

    def measure_performance(self, func_name=None):
        """性能测量装饰器"""
        def decorator(func):
            def wrapper(*args, **kwargs):
                start_time = time.time()
                func_display_name = func_name or func.__name__

                try:
                    result = func(*args, **kwargs)
                    return result
                finally:
                    end_time = time.time()
                    duration = end_time - start_time

                    if duration > 1.0:  # 只记录超过1秒的操作
                        self.log(f"⏱️ {func_display_name} 耗时 {duration:.1f} 秒", 'info')

                    # 记录性能统计
                    try:
                        if not hasattr(self, 'performance_stats'):
                            self.performance_stats = {}

                        if func_display_name not in self.performance_stats:
                            self.performance_stats[func_display_name] = {
                                'total_time': 0,
                                'call_count': 0,
                                'avg_time': 0,
                                'max_time': 0
                            }

                        stats = self.performance_stats[func_display_name]
                        stats['total_time'] += duration
                        stats['call_count'] += 1
                        stats['avg_time'] = stats['total_time'] / stats['call_count']
                        stats['max_time'] = max(stats['max_time'], duration)

                    except Exception:
                        pass

            return wrapper
        return decorator

    def browse_csv(self):
        filename = filedialog.askopenfilename(
            title="选择CSV文件",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.csv_file.set(filename)
            self.log(f"选择CSV文件: {os.path.basename(filename)}", 'info')

    def browse_output(self):
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.output_dir.set(directory)
            self.log(f"选择输出目录: {directory}", 'info')

    def clear_log(self):
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        self.log("日志已清空", 'info')

    def update_progress(self, current, total, message=""):
        progress = (current / total) * 100 if total > 0 else 0
        self.progress['value'] = progress
        self.progress_label.config(text=f"{message} ({current}/{total}) • {progress:.0f}%")
        self.root.update_idletasks()

    def read_links_from_csv(self, csv_file):
        """读取CSV文件，支持多种格式"""
        items = []
        try:
            with open(csv_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                if reader.fieldnames and ('链接' in reader.fieldnames or reader.fieldnames[0]):
                    for idx, row in enumerate(reader, 1):
                        link = (row.get('链接') or row.get(reader.fieldnames[0]) or '').strip()
                        if not link:
                            continue
                        product = (row.get('产品') or '').strip()
                        seq = (row.get('序号') or '').strip() or str(idx)
                        items.append({'产品': product, '序号': seq, '链接': link})
        except Exception:
            # 尝试简单格式
            try:
                with open(csv_file, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                    for idx, row in enumerate(rows[1:], 1):
                        if row and len(row) > 0 and row[0].strip():
                            items.append({'产品': '', '序号': str(idx), '链接': row[0].strip()})
            except Exception:
                pass
        return items

    def check_login_state_comprehensive(self, page):
        """v4.0全面登录状态检查"""
        try:
            # 多维度检测登录状态
            login_state = page.evaluate("""() => {
                const checks = {
                    // DOM元素检测
                    hasAvatar: !!document.querySelector('.avatar, .user-avatar, [class*="Avatar"]'),
                    hasUserInfo: !!document.querySelector('.user-info, .login-info, [class*="UserInfo"]'),
                    hasLoginBtn: !!document.querySelector('.login-btn, [class*="login-button"]'),

                    // Cookie检测
                    cookies: document.cookie,
                    hasWebSession: document.cookie.includes('web_session'),
                    hasXsecAppid: document.cookie.includes('xsecappid'),

                    // 页面内容检测
                    isLoginPage: document.body.innerText.includes('登录') ||
                                document.body.innerText.includes('扫码'),
                };

                // 综合判断登录状态
                checks.isLoggedIn = (
                    (checks.hasAvatar || checks.hasUserInfo) &&
                    (checks.hasWebSession || checks.hasXsecAppid) &&
                    !checks.hasLoginBtn &&
                    !checks.isLoginPage
                );

                return checks;
            }""")
            return login_state
        except Exception:
            return {'isLoggedIn': False}

    def detect_anti_crawler(self, page):
        """v4.0反爬虫检测"""
        try:
            indicators = page.evaluate("""() => {
                return {
                    hasCaptcha: !!document.querySelector('.captcha, [class*="captcha"], [class*="verify"]'),
                    hasBlockPage: !!document.querySelector('[class*="block"], [class*="forbidden"]'),
                    hasRateLimit: !!document.querySelector('[class*="rate-limit"], [class*="too-many"]'),
                    blockText: document.body.innerText.includes('访问过于频繁') ||
                              document.body.innerText.includes('请求过于频繁') ||
                              document.body.innerText.includes('请稍后再试') ||
                              document.body.innerText.includes('验证'),
                };
            }""")

            is_blocked = (
                indicators['hasCaptcha'] or
                indicators['hasBlockPage'] or
                indicators['hasRateLimit'] or
                indicators['blockText']
            )

            return is_blocked, indicators
        except Exception:
            return False, {}

    def perform_ultimate_login(self, page):
        """v4.0终极登录流程"""
        max_retries = int(self.login_retries.get())

        for attempt in range(max_retries):
            try:
                self.log(f"登录尝试 {attempt + 1}/{max_retries}", 'info')

                # 1. 打开小红书首页
                self.log("正在打开小红书首页...", 'info')
                page.goto('https://www.xiaohongshu.com/', wait_until='domcontentloaded', timeout=30000)
                time.sleep(3)

                # 2. 检查登录状态
                login_state = self.check_login_state_comprehensive(page)

                if login_state.get('isLoggedIn', False):
                    self.log("检测到已登录状态", 'success')

                    # 保存登录状态
                    try:
                        page.context.storage_state(path=self.storage_state_path)
                        self.log("登录状态已保存", 'success')
                    except Exception:
                        pass

                    return True

                # 3. 需要扫码登录
                self.log("需要扫码登录", 'info')
                self.log("请使用小红书App扫描浏览器中的二维码", 'warning')

                login_wait = int(self.login_wait_time.get())
                self.log(f"等待扫码登录，最长{login_wait}秒...", 'info')

                for remaining in range(login_wait, 0, -5):
                    if self.stop_flag:
                        break

                    if remaining % 15 == 0:
                        self.log(f"剩余 {remaining} 秒", 'info')

                    # 检查登录状态
                    if self.check_login_state_comprehensive(page).get('isLoggedIn', False):
                        self.log("检测到登录成功！", 'success')
                        time.sleep(2)

                        # 保存登录状态
                        try:
                            page.context.storage_state(path=self.storage_state_path)
                            self.log("登录状态已保存", 'success')
                        except Exception:
                            pass

                        return True

                    time.sleep(5)

                if attempt < max_retries - 1:
                    self.log(f"登录超时，准备重试...", 'warning')
                    time.sleep(5)
                else:
                    self.log("所有登录尝试均失败", 'error')
                    return False

            except Exception as e:
                self.log(f"登录尝试失败: {str(e)}", 'error')
                if attempt < max_retries - 1:
                    time.sleep(10)
                else:
                    return False

        return False

    def capture_with_ultimate_features(self, items, screenshot_dir):
        """v4.0终极版抓取功能"""
        results = []
        covers_dir = os.path.join(screenshot_dir, 'covers')
        os.makedirs(covers_dir, exist_ok=True)

        self.log("启动v3.0 Ultimate浏览器引擎...", 'info')

        with sync_playwright() as p:
            try:
                os.makedirs(self.user_data_dir, exist_ok=True)

                # 随机选择User-Agent
                self.http_user_agent = random.choice(self.config['user_agents'])
                self.log(f"使用User-Agent: {self.http_user_agent[:50]}...", 'info')

                launch_kwargs = dict(
                    user_data_dir=self.user_data_dir,
                    headless=False,
                    viewport={'width': 1920, 'height': 1080},
                    user_agent=self.http_user_agent,
                    ignore_https_errors=True,
                )

                # 优先使用系统 Chrome（更接近手动悬停行为）；失败则回退 Playwright Chromium
                try:
                    context = p.chromium.launch_persistent_context(channel="chrome", **launch_kwargs)
                    self.log("已使用 Chrome 启动浏览器", "success")
                except Exception as e:
                    self.log(f"Chrome 启动失败，回退 Chromium: {e}", "warning")
                    context = p.chromium.launch_persistent_context(**launch_kwargs)
            except Exception as e:
                self.log(f"启动浏览器失败: {e}", 'error')
                raise

            page = context.pages[0] if context.pages else context.new_page()
            page.set_default_timeout(60000)

            # 设置额外请求头
            extra_headers = {
                'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
                'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8',
                'accept-encoding': 'gzip, deflate, br',
                'sec-fetch-dest': 'document',
                'sec-fetch-mode': 'navigate',
                'sec-fetch-site': 'none'
            }
            page.set_extra_http_headers(extra_headers)

            # 执行登录
            login_success = self.perform_ultimate_login(page)
            if not login_success:
                self.log("登录失败，将以无登录模式继续", 'warning')

            # 处理每个链接
            for idx, item in enumerate(items, 1):
                if self.stop_flag:
                    self.log("用户中断处理", 'info')
                    break

                link = item.get('链接', '')
                product = item.get('产品', '')
                seq_val = item.get('序号', str(idx))
                name_prefix = self.sanitize_name(f"{product}_{seq_val}") if product else self.sanitize_name(f"item_{seq_val}")

                self.update_progress(idx, len(items), f"处理链接 {idx}")
                self.log(f"[{idx}/{len(items)}] 处理: {link[:60]}...", 'info')

                result = {
                    '序号': seq_val,
                    '产品': product,
                    '链接': link,
                    '笔记ID': self.extract_note_id(link),
                    '采集状态': '失败',
                    '错误信息': '',
                    'HTTP状态': '',
                    '标题': '',
                    '作者昵称': '',
                    '作者ID': '',
                    '粉丝数': '',
                    '博主主页': '',
                    '发布时间': '',
                    '点赞数': '',
                    '收藏数': '',
                    '评论数': '',
                    '分享数': '',
                    '封面链接': '',
                    '视频链接': '',
                    '正文': '',
                    '截屏文件': '',
                    '封面图列表': [],
                    '封面图数量': 0,
                    '检测时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    '登录状态': '成功' if login_success else '未登录',
                    '反爬状态': '正常'
                }

                try:
                    # v4.0 链接间随机延迟
                    if idx > 1:
                        delay = self.get_random_delay('link_interval')
                        self.log(f"链接间延迟 {delay:.1f} 秒...", 'info')
                        time.sleep(delay)

                    # 访问链接（增强版重试和错误处理）
                    max_retries = 3  # 增加重试次数
                    resp = None
                    page_load_success = False

                    for attempt in range(max_retries):
                        try:
                            self.log(f"  [链接] 尝试访问 ({attempt + 1}/{max_retries})...", 'info')
                            resp = page.goto(link, timeout=60000, wait_until='domcontentloaded')

                            # 验证页面是否真正加载成功
                            if resp and resp.status == 200:
                                # 等待页面完全加载
                                time.sleep(3)
                                # 检查页面是否有内容
                                page_content = page.evaluate("""() => {
                                    return document.body && document.body.innerText.length > 0;
                                }""")
                                if page_content:
                                    page_load_success = True
                                    self.log(f"  [链接] 页面加载成功 (状态: {resp.status})", 'success')
                                    break
                                else:
                                    self.log(f"  [链接] 页面内容为空，重试...", 'warning')
                            else:
                                self.log(f"  [链接] HTTP错误: {resp.status if resp else 'unknown'}", 'warning')

                        except Exception as e:
                            self.log(f"  [链接] 访问失败: {str(e)}", 'warning')
                            if attempt < max_retries - 1:
                                time.sleep(3)  # 等待后重试
                                continue

                    if not page_load_success:
                        self.log(f"  [链接] 所有尝试均失败", 'error')
                        result['错误信息'] = "页面加载失败"
                        # 不要抛出异常，继续处理下一个链接
                        self.stats['failed_count'] += 1
                        continue

                    result['HTTP状态'] = resp.status if resp else ''

                    # 页面加载后随机等待
                    page_load_delay = self.get_random_delay('page_load')
                    self.log(f"  [页面] 页面稳定等待 {page_load_delay:.1f} 秒...", 'info')
                    time.sleep(page_load_delay)

                    # v4.0 反爬检测
                    if self.enable_anti_detection.get():
                        is_blocked, indicators = self.detect_anti_crawler(page)
                        if is_blocked:
                            result['反爬状态'] = '触发反爬'
                            self.stats['anti_crawler_count'] += 1
                            self.update_stats()

                            # 尝试反爬处理
                            self.log("  [反爬] 检测到反爬机制，尝试处理...", 'warning')
                            time.sleep(random.uniform(8, 15))

                            # 刷新页面
                            try:
                                page.reload(wait_until='domcontentloaded')
                                time.sleep(page_load_delay)
                                self.log("  [反爬] 页面刷新完成", 'success')
                            except Exception as e:
                                self.log(f"  [反爬] 页面刷新失败: {str(e)}", 'error')

                    # 检查登录弹窗和页面状态
                    try:
                        page_status = page.evaluate("""() => {
                            return {
                                hasLoginPopup: !!(document.querySelector('.login-modal, [class*="LoginModal"]') ||
                                                 document.querySelector('[class*="qrcode"]')),
                                hasContent: document.body.innerText.length > 100,
                                title: document.title || '',
                                url: window.location.href
                            };
                        }""")

                        if page_status['hasLoginPopup']:
                            self.log("  [页面] 检测到登录弹窗，可能需要重新登录", 'warning')

                        if not page_status['hasContent']:
                            self.log("  [页面] 页面内容过少，可能需要重新加载", 'warning')

                        self.log(f"  [页面] 标题: {page_status.get('title', 'unknown')[:50]}...", 'info')

                    except Exception as e:
                        self.log(f"  [页面] 页面状态检查失败: {str(e)}", 'warning')

                    # 截图
                    screenshot_name = f'{name_prefix}_screenshot.png'
                    screenshot_path = os.path.join(screenshot_dir, screenshot_name)

                    try:
                        screenshot_delay = self.get_random_delay('screenshot')
                        self.log(f"截图前等待 {screenshot_delay:.1f} 秒...", 'info')
                        time.sleep(screenshot_delay)

                        page.screenshot(path=screenshot_path, full_page=False, timeout=60000)
                        result['截屏文件'] = screenshot_path
                        self.log(f"截图成功: {screenshot_name}", 'success')

                    except Exception as se:
                        self.log(f"截图失败: {str(se)}", 'error')
                        try:
                            page.locator("body").screenshot(path=screenshot_path, timeout=60000)
                            result['截屏文件'] = screenshot_path
                            self.log(f"使用body兜底截图成功", 'success')
                        except Exception as se2:
                            self.log(f"兜底截图也失败: {str(se2)}", 'error')

                    # 完整模式深度抓取
                    is_full_mode = self.capture_mode.get() == "full"
                    extraction_success = True
                    extraction_errors = []

                    if is_full_mode:
                        self.log("进入完整模式深度抓取...", 'info')

                        # 分别处理不同的数据提取，避免一个失败影响所有
                        try:
                            details, image_urls = self.extract_detailed_content_ultimate(page, link)
                            result.update(details)
                            self.log("基础信息提取成功", 'success')
                        except Exception as detail_e:
                            extraction_success = False
                            extraction_errors.append(f"基础信息提取失败: {str(detail_e)}")
                            self.log(f"基础信息提取失败: {str(detail_e)}", 'error')
                            # 设置默认的空值，避免字段缺失
                            details = {}
                            image_urls = []

                        # 封面图抓取 - 独立处理
                        try:
                            cover_paths = self.capture_cover_images_ultimate(page, covers_dir, name_prefix, image_urls)
                            result['封面图列表'] = cover_paths
                            result['封面图数量'] = len(cover_paths) if cover_paths else len(image_urls or [])

                            if not result.get('封面链接') and image_urls:
                                result['封面链接'] = image_urls[0]

                            self.log(f"封面图获取 {len(cover_paths)} 张", 'success')
                        except Exception as cover_e:
                            extraction_errors.append(f"封面图获取失败: {str(cover_e)}")
                            self.log(f"封面图获取失败: {str(cover_e)}", 'error')
                            # 设置默认值
                            result['封面图列表'] = []
                            result['封面图数量'] = 0

                        # 统计和报告提取结果
                        if details:
                            self.log(f"深度抓取完成 - 点赞:{details.get('点赞数','')} 收藏:{details.get('收藏数','')} 评论:{details.get('评论数','')} 粉丝:{details.get('粉丝数','')}", 'success')
                        else:
                            self.log("基础信息提取为空，但继续处理", 'warning')

                        # 判断整体提取是否成功 - 更宽松的成功标准
                        # 成功标准：满足以下任一条件即可
                        # 1. 获取到基础信息（标题、作者等）
                        # 2. 获取到粉丝数（即使其他数据失败）
                        # 3. 获取到封面图
                        # 4. 获取到互动数据（点赞、收藏、评论等）
                        has_basic_info = bool(details.get('标题') or details.get('作者昵称'))
                        has_fans = bool(details.get('粉丝数'))
                        has_cover = bool(result.get('封面图列表'))
                        has_interaction = bool(details.get('点赞数') or details.get('收藏数') or details.get('评论数'))
                        has_content = bool(details.get('正文'))
                        has_screenshot = bool(result.get('截屏文件'))

                    # MCP同步：记录各个指标的具体状态
                    self.log(f"  [MCP状态] 详细指标检查:", 'info')
                    self.log(f"    - 基础信息(标题/作者): {has_basic_info} (标题:{bool(details.get('标题'))}, 作者:{bool(details.get('作者昵称'))})", 'info')
                    self.log(f"    - 粉丝数: {has_fans} (值:{details.get('粉丝数', 'None')})", 'info')
                    self.log(f"    - 封面图: {has_cover} (数量:{result.get('封面图数量', 0)})", 'info')
                    self.log(f"    - 互动数据: {has_interaction} (点赞:{details.get('点赞数')}, 收藏:{details.get('收藏数')}, 评论:{details.get('评论数')})", 'info')
                    self.log(f"    - 正文内容: {has_content} (长度:{len(details.get('正文', ''))})", 'info')
                    self.log(f"    - 截图文件: {has_screenshot} (路径:{bool(result.get('截屏文件'))})", 'info')

                    # 计算成功指标数量
                    success_indicators = sum([
                        has_basic_info,
                        has_fans,
                        has_cover,
                        has_interaction,
                        has_content,
                        has_screenshot
                    ])

                    self.log(f"  [MCP评估] 成功指标数: {success_indicators}/6", 'success')

                    if extraction_errors and success_indicators == 0:
                        # 只有在完全没有任何成功数据时才标记为失败
                        result['错误信息'] = "; ".join(extraction_errors)
                        self.log(f"  [评估] ✗ 完全失败: {result['错误信息']}", 'error')
                    elif extraction_errors and success_indicators > 0:
                        # 部分成功，记录错误但标记为成功
                        result['错误信息'] = f"部分成功但有错误: {'; '.join(extraction_errors)}" if extraction_errors else ""
                        self.log(f"  [评估] ✓ 部分成功: {success_indicators}个指标成功", 'success')
                    else:
                        # 完全成功
                        result['错误信息'] = ""
                        self.log(f"  [评估] ✓ 完全成功: {success_indicators}个指标成功", 'success')

                    # 设置采集状态 - 使用更宽松的成功标准
                    if success_indicators >= 2 or (has_fans and success_indicators >= 1):
                        # 有2个以上成功指标，或有粉丝数+其他指标，标记为成功
                        result['采集状态'] = '成功'
                        self.stats['success_count'] += 1
                        self.log(f"  [状态] ✓ 采集成功: {success_indicators}个指标", 'success')
                    elif success_indicators >= 1:
                        # 至少1个成功指标，也算作成功（特别是有粉丝数的情况）
                        result['采集状态'] = '成功'
                        self.stats['success_count'] += 1
                        self.log(f"  [状态] ✓ 采集部分成功: {success_indicators}个指标", 'success')
                    else:
                        # 完全没有成功数据
                        result['采集状态'] = '失败'
                        self.stats['failed_count'] += 1
                        self.log(f"  [状态] ✗ 采集失败: 无有效数据", 'error')

                    # 更详细的错误信息处理
                    if extraction_errors:
                        if result['采集状态'] == '成功':
                            result['错误信息'] = f"成功但有警告: {'; '.join(extraction_errors)}"
                        else:
                            result['错误信息'] = f"失败原因: {'; '.join(extraction_errors)}"
  
                except Exception as e:
                    result['错误信息'] = str(e)
                    self.stats['failed_count'] += 1
                    self.log(f"处理失败: {str(e)}", 'error')

                results.append(result)
                self.stats['total_processed'] += 1
                self.update_stats()

            try:
                context.close()
            except Exception:
                pass

        return results

    def fetch_user_fans_via_api(self, page, author_id):
        """增强版API粉丝数获取（多端点、重试机制、智能解析）"""
        import time
        import json

        if not author_id:
            self.log("  [API] 作者ID为空，跳过API调用", 'warning')
            return ""

        # 多个API端点作为备选
        api_endpoints = [
            # 主端点 - 用户详细信息
            {
                "url": f"https://edith.xiaohongshu.com/api/sns/web/v1/user/otherinfo?user_id={author_id}",
                "name": "用户详细信息API",
                "primary": True
            },
            # 备选端点1 - 用户基本信息
            {
                "url": f"https://edith.xiaohongshu.com/api/sns/web/v1/user/basicinfo?user_id={author_id}",
                "name": "用户基本信息API",
                "primary": False
            },
            # 备选端点2 - 用户卡片信息
            {
                "url": f"https://edith.xiaohongshu.com/api/sns/web/v1/user/cardinfo?user_id={author_id}",
                "name": "用户卡片信息API",
                "primary": False
            },
            # 备选端点3 - 用户统计信息
            {
                "url": f"https://www.xiaohongshu.com/api/sns/web/v1/user/{author_id}/stats",
                "name": "用户统计信息API",
                "primary": False
            }
        ]

        # 智能重试配置
        max_total_retries = 3
        endpoint_timeout = 20  # 增加超时时间

        for attempt in range(max_total_retries):
            if attempt > 0:
                # 指数退避，抖动处理
                base_delay = 2 ** attempt
                jitter = random.uniform(0.5, 1.5)
                delay = min(base_delay * jitter, 15)  # 最大15秒
                self.log(f"  [API] 第{attempt + 1}次重试，等待 {delay:.1f} 秒", 'info')
                time.sleep(delay)

            for endpoint in api_endpoints:
                try:
                    api_url = endpoint["url"]
                    endpoint_name = endpoint["name"]
                    is_primary = endpoint["primary"]

                    self.log(f"  [API] 尝试 {endpoint_name}: {api_url[:60]}...", 'info')

                    # 构建更全面的请求头
                    headers = {
                        "Accept": "application/json, text/plain, */*",
                        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
                        "Accept-Encoding": "gzip, deflate, br",
                        "Cache-Control": "no-cache",
                        "Pragma": "no-cache",
                        "Origin": "https://www.xiaohongshu.com",
                        "Referer": page.url,
                        "User-Agent": self.http_user_agent,
                        "Sec-Fetch-Dest": "empty",
                        "Sec-Fetch-Mode": "cors",
                        "Sec-Fetch-Site": "same-site",
                        # 尝试添加可能的认证信息
                        "X-Requested-With": "XMLHttpRequest"
                    }

                    # 执行API请求
                    response = page.context.request.get(
                        api_url,
                        headers=headers,
                        timeout=endpoint_timeout * 1000  # 转换为毫秒
                    )

                    # 记录响应状态
                    status_code = response.status if hasattr(response, 'status') else 'unknown'
                    self.log(f"  [API] 响应状态: {status_code}", 'info')

                    if response.ok:
                        try:
                            # 尝试解析JSON响应
                            response_text = response.text() if hasattr(response, 'text') else str(response.body())
                            data = json.loads(response_text)

                            self.log(f"  [API] JSON解析成功", 'success')

                            # 智能数据解析 - 多层次查找粉丝数
                            fans_value = None

                            # 1. 直接查找粉丝数字段
                            fans_fields = [
                                # 新版本字段
                                "fans", "followers_count", "fans_count",
                                # 兼容旧版本字段
                                "followed_count", "fansNum", "fans_num",
                                # 可能的其他字段
                                "subscriber_count", "fans_total", "followers"
                            ]

                            # 递归查找函数
                            def find_fans_in_data(data, fields, depth=0):
                                if depth > 5:  # 防止无限递归
                                    return None

                                if isinstance(data, dict):
                                    for field in fields:
                                        if field in data:
                                            value = data[field]
                                            if value is not None and value != "":
                                                return value

                                    # 在嵌套结构中查找
                                    for key, value in data.items():
                                        if isinstance(value, (dict, list)):
                                            result = find_fans_in_data(value, fields, depth + 1)
                                            if result is not None:
                                                return result

                                elif isinstance(data, list):
                                    for item in data:
                                        result = find_fans_in_data(item, fields, depth + 1)
                                        if result is not None:
                                            return result

                                return None

                            # 2. 从响应的不同层级查找
                            # 常见的数据结构路径
                            search_paths = [
                                ["data"],  # 直接在data字段
                                ["data", "user_info"],  # 在data.user_info
                                ["data", "user"],  # 在data.user
                                ["data", "basic_info"],  # 在data.basic_info
                                ["data", "stats"],  # 在data.stats
                                ["user_info"],  # 在user_info
                                ["user"],  # 在user
                                ["result"],  # 在result
                                ["items", 0] if isinstance(data.get("items"), list) and data.get("items") else None,  # 在items[0]
                            ]

                            for path in search_paths:
                                if path is None:
                                    continue

                                try:
                                    current_data = data
                                    for key in path:
                                        current_data = current_data[key]

                                    fans_value = find_fans_in_data(current_data, fans_fields)
                                    if fans_value is not None:
                                        break

                                except (KeyError, TypeError, IndexError):
                                    continue

                            # 3. 如果路径查找失败，尝试全局搜索
                            if fans_value is None:
                                fans_value = find_fans_in_data(data, fans_fields)

                            # 4. 处理找到的粉丝数值
                            if fans_value is not None:
                                # 处理不同格式的粉丝数
                                if isinstance(fans_value, (int, float)):
                                    fans_str = str(int(fans_value))
                                elif isinstance(fans_value, str):
                                    fans_str = fans_value.strip()
                                    # 处理格式化的数字（如 "1.2万"）
                                    fans_str = self.parse_fans_number(fans_str)
                                else:
                                    fans_str = str(fans_value)

                                if fans_str and fans_str != "0":
                                    self.log(f"  [API] ✓ 粉丝数获取成功: {fans_str} (来源: {endpoint_name})", 'success')
                                    return fans_str

                            # 5. 检查API响应中的错误信息
                            if isinstance(data, dict):
                                error_msg = data.get('msg') or data.get('error') or data.get('message')
                                if error_msg:
                                    self.log(f"  [API] API返回错误: {error_msg}", 'warning')

                        except json.JSONDecodeError as e:
                            self.log(f"  [API] JSON解析失败: {str(e)}", 'error')
                            # 记录原始响应用于调试
                            try:
                                raw_response = response.text()[:200] if hasattr(response, 'text') else str(response.body())[:200]
                                self.log(f"  [API] 原始响应: {raw_response}...", 'info')
                            except Exception:
                                pass

                    else:
                        # 处理HTTP错误状态
                        status_text = getattr(response, 'status_text', '')
                        if status_code == 429:
                            self.log(f"  [API] 请求频率限制，等待后重试", 'warning')
                            time.sleep(5)  # 额外等待
                            continue
                        elif status_code == 403:
                            self.log(f"  [API] 访问被拒绝，可能需要认证", 'warning')
                        elif status_code == 404:
                            self.log(f"  [API] API端点不存在或用户不存在", 'warning')
                            if is_primary:
                                continue  # 主端点404时尝试其他端点
                        elif status_code >= 500:
                            self.log(f"  [API] 服务器错误: {status_code} {status_text}", 'error')
                        else:
                            self.log(f"  [API] HTTP错误: {status_code} {status_text}", 'warning')

                except Exception as e:
                    self.log(f"  [API] {endpoint_name} 请求失败: {str(e)}", 'error')
                    continue

        # 所有API尝试均失败
        self.log("  [API] 所有API端点均失败", 'error')
        return ""

    def parse_fans_number(self, fans_str):
        """解析格式化的粉丝数字符串（如"1.2万"、"3.4K"等）"""
        try:
            if not fans_str or not isinstance(fans_str, str):
                return ""

            fans_str = fans_str.strip().upper()

            # 处理中文单位
            if '万' in fans_str:
                number_part = fans_str.replace('万', '').strip()
                try:
                    number = float(number_part)
                    return str(int(number * 10000))
                except ValueError:
                    pass

            # 处理英文单位
            if 'W' in fans_str:
                number_part = fans_str.replace('W', '').strip()
                try:
                    number = float(number_part)
                    return str(int(number * 10000))
                except ValueError:
                    pass

            if 'K' in fans_str:
                number_part = fans_str.replace('K', '').strip()
                try:
                    number = float(number_part)
                    return str(int(number * 1000))
                except ValueError:
                    pass

            # 移除非数字字符，保留数字
            import re
            numbers_only = re.sub(r'[^\d]', '', fans_str)
            if numbers_only:
                return numbers_only

            # 尝试直接转换为整数
            try:
                number = int(fans_str)
                return str(number)
            except ValueError:
                pass

            return ""

        except Exception as e:
            self.log(f"    粉丝数格式解析失败: {str(e)}", 'warning')
            return ""

    def fetch_user_fans_via_homepage(self, context, author_home_url):
        """打开博主主页作为粉丝数兜底"""
        if not author_home_url:
            return ""
        try:
            self.log("  [主页] 打开博主主页获取粉丝数...", 'info')
            home_page = context.new_page()
            home_page.goto(author_home_url, wait_until='domcontentloaded', timeout=20000)
            time.sleep(2)
            fans_text = home_page.evaluate("""() => {
                const bodyText = document.body.innerText.substring(0, 8000);
                const patterns = [
                    /(粉丝|关注者)[^\\d]{0,3}([0-9]+\\.?[0-9]*[万wWkK]?)/i,
                    /([0-9]+\\.?[0-9]*[万wWkK]?)\\s*(粉丝|关注者)/i,
                    /粉丝数[^\\d]{0,3}([0-9]+\\.?[0-9]*[万wWkK]?)/i,
                ];
                for (const p of patterns) {
                    const m = bodyText.match(p);
                    if (m) return m[2] || m[1] || '';
                }
                return '';
            }""")
            try:
                home_page.close()
            except Exception:
                pass
            return fans_text or ""
        except Exception as e:
            self.log(f"  [主页] 粉丝数获取失败: {str(e)}", 'warning')
            return ""

    def fetch_user_fans_directly(self, page, url):
        """直接从页面提取粉丝数（基于测试结果，粉丝数直接显示在页面上）"""
        try:
            self.log("  [粉丝] 开始直接提取粉丝数信息", 'info')

            # 基于测试结果，我们知道粉丝数信息直接显示在页面上
            # 格式如：5.1万粉丝，位置通常在用户信息区域
            fans_data = page.evaluate("""
                () => {
                    const result = {
                        fans: '',
                        confidence: 0,
                        method: '',
                        debug: []
                    };

                    // 1. 查找包含"粉丝"文本的所有元素
                    const elements = document.querySelectorAll('*');
                    const fansElements = [];

                    elements.forEach(el => {
                        const text = (el.innerText || el.textContent || '').trim();
                        if (text.includes('粉丝') || text.includes('关注者')) {
                            const rect = el.getBoundingClientRect();
                            if (rect.width > 0 && rect.height > 0 && el.offsetParent !== null) {
                                fansElements.push({
                                    element: el,
                                    text: text,
                                    rect: rect,
                                    className: (el.className || '').toString(),
                                    tagName: el.tagName
                                });
                            }
                        }
                    });

                    result.debug.push(`找到 ${fansElements.length} 个包含粉丝信息的元素`);

                    // 2. 按优先级解析粉丝数
                    // 优先级1: 查找用户信息区域（如 "26关注 5.1万粉丝 32.8万获赞与收藏"）
                    for (const el of fansElements) {
                        const text = el.text;
                        if (text.includes('关注') && text.includes('粉丝') &&
                           (text.includes('获赞') || text.includes('收藏'))) {
                            const fansMatch = text.match(/([0-9]+\.?[0-9]*[万wWkK]?)\s*粉丝/);
                            if (fansMatch) {
                                result.fans = fansMatch[1];
                                result.confidence = 95;
                                result.method = 'user_info_block';
                                result.debug.push(`从用户信息块提取: ${text.substring(0, 50)}...`);
                                return result;
                            }
                        }
                    }

                    // 优先级2: 直接匹配"数字+粉丝"格式
                    for (const el of fansElements) {
                        const text = el.text;
                        const patterns = [
                            /([0-9]+\.?[0-9]*[万wWkK]?)\s*粉丝/,
                            /粉丝[:\s：]*([0-9]+\.?[0-9]*[万wWkK]?)/,
                            /([0-9]+\.?[0-9]*[万wWkK]?)\s*(粉丝|关注者)/
                        ];

                        for (const pattern of patterns) {
                            const match = text.match(pattern);
                            if (match) {
                                const fansNumber = match[1] || match[2];
                                if (fansNumber) {
                                    result.fans = fansNumber;
                                    result.confidence = 90;
                                    result.method = 'direct_text_match';
                                    result.debug.push(`找到匹配: ${text.substring(0, 50)}...`);
                                    result.debug.push(`提取粉丝数: ${fansNumber}`);
                                    return result;
                                }
                            }
                        }
                    }

                    // 优先级3: 单独的粉丝数元素
                    for (const el of fansElements) {
                        const text = el.text.trim();
                        if (text.includes('粉丝') && text.length < 50) {
                            const fansMatch = text.match(/([0-9]+\.?[0-9]*[万wWkK]?)/);
                            if (fansMatch) {
                                result.fans = fansMatch[1];
                                result.confidence = 70;
                                result.method = 'fans_element_only';
                                result.debug.push(`从单独粉丝元素提取: ${text}`);
                                return result;
                            }
                        }
                    }

                    // 优先级4: 基于位置查找（右侧用户信息区域）
                    const targetElements = fansElements.filter(el => {
                        const rect = el.rect;
                        // 基于测试位置，大约在页面右侧用户信息区域
                        return rect.left > 600 && rect.top > 100 && rect.top < 600;
                    });

                    for (const el of targetElements) {
                        const text = el.text;
                        const fansMatch = text.match(/([0-9]+\.?[0-9]*[万wWkK]?)\s*粉丝/);
                        if (fansMatch) {
                            result.fans = fansMatch[1];
                            result.confidence = 80;
                            result.method = 'position_based';
                            result.debug.push(`基于位置提取: ${text.substring(0, 50)}...`);
                            return result;
                        }
                    }

                    result.debug.push('所有解析方法均失败');
                    return result;
                }
            """)

            # 记录调试信息
            self.log(f"  [粉丝] 解析方法: {fans_data.get('method', 'none')}", 'info')
            self.log(f"  [粉丝] 置信度: {fans_data.get('confidence', 0)}%", 'info')

            # 显示调试信息
            debug_messages = fans_data.get('debug', [])
            for msg in debug_messages[:3]:  # 只显示前3条
                self.log(f"  [调试] {msg}", 'debug')

            # 提取粉丝数
            if fans_data and fans_data.get('fans'):
                fans_number = fans_data['fans']
                method = fans_data.get('method', '')

                self.log(f"  [粉丝] ✓ 提取成功: {fans_number} (方法: {method})", 'success')

                # 解析粉丝数（统一格式）
                if fans_number:
                    import re
                    pattern = r"([0-9]+\.?[0-9]*[万wWkK]?)"
                    match = re.search(pattern, fans_number, re.IGNORECASE)
                    if match:
                        fans_result = match.group(1)
                        fans_result = self.parse_fans_number(fans_result)
                        if fans_result and fans_result != "0":
                            self.log(f"  [粉丝] ✓ 解析成功: {fans_result} (原始: {fans_number})", 'success')
                            return fans_result

            self.log("  [粉丝] 未能提取到有效的粉丝数", 'warning')
            return ""

        except Exception as e:
            self.log(f"  [粉丝] 直接提取失败: {str(e)}", 'error')
            return ""

    def fetch_user_fans_via_hover(self, page, author_id="", author_name=""):
        """直接从页面获取粉丝数（基于测试结果，粉丝数直接显示在页面上）"""
        try:
            self.log("  [粉丝] 开始直接抓取页面粉丝数信息", 'info')

            # 基于测试结果，我们知道粉丝数信息直接显示在页面上
            # 格式如：5.1万粉丝，位置通常在用户信息区域
            fans_data = page.evaluate("""
                () => {
                    const result = {
                        fans: '',
                        confidence: 0,
                        method: '',
                        allFansInfo: [],
                        debug: []
                    };

                    // 1. 查找包含"粉丝"文本的所有元素
                    const elements = document.querySelectorAll('*');
                    const fansElements = [];

                    elements.forEach(el => {
                        const text = (el.innerText || el.textContent || '').trim();
                        if (text.includes('粉丝') || text.includes('关注者')) {
                            const rect = el.getBoundingClientRect();
                            if (rect.width > 0 && rect.height > 0 && el.offsetParent !== null) {
                                fansElements.push({
                                    element: el,
                                    text: text,
                                    rect: rect,
                                    className: (el.className || '').toString(),
                                    tagName: el.tagName,
                                    html: el.innerHTML.substring(0, 200)
                                });
                            }
                        }
                    });

                    result.debug.push(`找到 ${fansElements.length} 个包含粉丝信息的元素`);
                    result.allFansInfo = fansElements.map(el => ({
                        text: el.text,
                        className: el.className,
                        tagName: el.tagName,
                        visible: true,
                        rect: el.rect.toString()
                    }));

                    // 2. 按优先级解析粉丝数
                    // 优先级1: 包含"粉丝"的数字信息
                    for (const el of fansElements) {
                        const text = el.text;

                        // 匹配数字+粉丝的格式
                        const patterns = [
                            /([0-9]+\.?[0-9]*[万wWkK]?)\s*粉丝/,
                            /粉丝[:\s：]*([0-9]+\.?[0-9]*[万wWkK]?)/,
                            /关注者[:\s：]*([0-9]+\.?[0-9]*[万wWkK]?)/,
                            /([0-9]+\.?[0-9]*[万wWkK]?)\s*(粉丝|关注者)/
                        ];

                        for (const pattern of patterns) {
                            const match = text.match(pattern);
                            if (match) {
                                const fansNumber = match[1] || match[2];
                                if (fansNumber) {
                                    result.fans = fansNumber;
                                    result.confidence = 95;
                                    result.method = 'direct_text_match';
                                    result.debug.push(`找到匹配: ${pattern}`);
                                    result.debug.push(`提取粉丝数: ${fansNumber}`);
                                    return result;
                                }
                            }
                        }
                    }

                    // 优先级2: 查找用户信息区域（通常包含关注、粉丝、获赞）
                    for (const el of fansElements) {
                        const text = el.text;
                        if (text.includes('关注') && text.includes('粉丝') && text.includes('获赞')) {
                            // 这可能是完整的用户信息，如 "26关注 5.1万粉丝 32.8万获赞与收藏"
                            const fansMatch = text.match(/([0-9]+\.?[0-9]*[万wWkK]?)\s*粉丝/);
                            if (fansMatch) {
                                result.fans = fansMatch[1];
                                result.confidence = 90;
                                result.method = 'user_info_block';
                                result.debug.push(`从用户信息块提取: ${text}`);
                                return result;
                            }
                        }
                    }

                    // 优先级3: 单独的粉丝数元素
                    for (const el of fansElements) {
                        const text = el.text.trim();
                        if (text.includes('粉丝') && text.length < 20) {
                            // 短文本，可能是单独的粉丝数元素
                            const fansMatch = text.match(/([0-9]+\.?[0-9]*[万wWkK]?)/);
                            if (fansMatch) {
                                result.fans = fansMatch[1];
                                result.confidence = 70;
                                result.method = 'fans_element_only';
                                result.debug.push(`从单独粉丝元素提取: ${text}`);
                                return result;
                            }
                        }
                    }

                    // 优先级4: 查找特定位置的元素（基于测试的坐标位置）
                    const targetElements = elements.filter(el => {
                        const rect = el.getBoundingClientRect();
                        const text = (el.innerText || '').toLowerCase();

                        // 基于测试位置，大约在页面右侧用户信息区域
                        return rect.left > 600 && rect.top > 200 && rect.top < 400 &&
                               (text.includes('粉丝') || text.includes('关注'));
                    });

                    for (const el of targetElements) {
                        const text = (el.innerText || '').trim();
                        const fansMatch = text.match(/([0-9]+\.?[0-9]*[万wWkK]?)\s*粉丝/);
                        if (fansMatch) {
                            result.fans = fansMatch[1];
                            result.confidence = 80;
                            result.method = 'position_based';
                            result.debug.push(`基于位置提取: ${text} at (${el.getBoundingClientRect().left}, ${el.getBoundingClientRect().top})`);
                            return result;
                        }
                    }

                    result.debug.push('所有解析方法均失败');
                    return result;
                }
            """)

            # 记录调试信息
            self.log(f"  [粉丝] 解析方法: {fans_data.get('method', 'none')}", 'info')
            self.log(f"  [粉丝] 置信度: {fans_data.get('confidence', 0)}%", 'info')
            self.log(f"  [粉丝] 找到元素数: {len(fans_data.get('allFansInfo', []))}", 'info')

            # 显示调试信息
            debug_messages = fans_data.get('debug', [])
            for msg in debug_messages[:5]:  # 只显示前5条
                self.log(f"  [调试] {msg}", 'debug')

            # 提取粉丝数
            if fans_data and fans_data.get('fans'):
                fans_number = fans_data['fans']
                confidence = fans_data.get('confidence', 0)
                method = fans_data.get('method', '')

                self.log(f"  [粉丝] ✓ 提取成功: {fans_number} (置信度: {confidence}% | 方法: {method})", 'success')

                # 解析粉丝数（统一格式）
                if fans_number:
                    import re
                    patterns = [
                        r"([0-9]+\.?[0-9]*[万wWkK]?)"
                    ]

                    for pattern in patterns:
                        match = re.search(pattern, fans_number, re.IGNORECASE)
                        if match:
                            fans_result = match.group(1)
                            fans_result = self.parse_fans_number(fans_result)
                            if fans_result and fans_result != "0":
                                self.log(f"  [粉丝] ✓ 解析成功: {fans_result} (原始: {fans_number})", 'success')
                                return fans_result

            self.log("  [粉丝] 未能提取到有效的粉丝数", 'warning')
            return ""

        except Exception as e:
            self.log(f"  [粉丝] 直接提取失败: {str(e)}", 'error')
            return ""

    def fetch_user_fans_via_hover(self, page, author_id="", author_name=""):
        """
        悬停作者区域抓取粉丝数（Chrome 悬浮卡片）。
        先用你提供的坐标框做“探针”读取粉丝行，再用悬浮卡片容器兜底，避免全页扫描误命中。
        """
        try:
            # 让页面回到顶部，保证悬浮卡片位置相对稳定
            try:
                page.evaluate("() => window.scrollTo(0, 0)")
            except Exception:
                pass

            def pick_topmost(selector):
                best = None
                best_y = None
                for el in page.query_selector_all(selector) or []:
                    try:
                        box = el.bounding_box()
                        if not box or box.get("width", 0) < 10 or box.get("height", 0) < 10:
                            continue
                        y = box.get("y", 10**9)
                        if best is None or y < best_y:
                            best = el
                            best_y = y
                    except Exception:
                        continue
                return best

            # 1) 定位作者元素
            author_el = None
            if author_id:
                author_el = pick_topmost(f'a[href*="/user/profile/{author_id}"]')
            if not author_el:
                author_el = pick_topmost('a[href*="/user/profile/"]')

            if not author_el and author_name:
                try:
                    locator = page.locator(f'text="{author_name}"').first
                    handle = locator.element_handle() if locator else None
                    if handle:
                        author_el = handle
                except Exception:
                    pass

            if not author_el:
                return ""

            try:
                author_el.scroll_into_view_if_needed(timeout=4000)
            except Exception:
                pass

            # 2) 悬停触发卡片
            hovered = False
            try:
                author_el.hover(timeout=5000)
                hovered = True
            except Exception:
                try:
                    box = author_el.bounding_box()
                    if box:
                        page.mouse.move(box["x"] + box["width"] / 2, box["y"] + box["height"] / 2)
                        hovered = True
                except Exception:
                    hovered = False

            if not hovered:
                return ""

            # 3) 坐标探针（你给的坐标：粉丝行 bbox）
            probe = getattr(self, "fans_probe_box", None) or (742.59375, 307.8046875, 320, 16.796875)
            fans_text = ""
            try:
                for _ in range(12):  # ~3-4s
                    fans_text = page.evaluate(
                        r"""(box) => {
                            const arr = Array.isArray(box) ? box : [];
                            const x = Number(arr[0] || 0), y = Number(arr[1] || 0), w = Number(arr[2] || 0), h = Number(arr[3] || 0);
                            if (!w || !h) return '';

                            const cx = x + w / 2;
                            const cy = y + h / 2;
                            const vw = window.innerWidth || 0;
                            const vh = window.innerHeight || 0;
                            const clamp = (n, min, max) => Math.max(min, Math.min(max, n));

                            const want = (t) => {
                                if (!t) return false;
                                const s = String(t).trim();
                                if (!s || s.length > 240) return false;
                                if (!/\\d/.test(s)) return false;
                                return s.includes('粉丝') || s.includes('关注者') || /followers/i.test(s);
                            };

                            const tryExtract = (el) => {
                                let cur = el;
                                for (let depth = 0; depth < 6 && cur; depth++) {
                                    const t = (cur.innerText || cur.textContent || '').trim();
                                    if (want(t)) return t;
                                    cur = cur.parentElement;
                                }
                                return '';
                            };

                            // 3.1 精确命中
                            const points = [
                                [cx, cy],
                                [x + 8, cy],
                                [x + w - 8, cy],
                                [cx, y + 2],
                                [cx, y + h - 2],
                            ];
                            for (const [px0, py0] of points) {
                                const px = clamp(px0, 0, Math.max(0, vw - 1));
                                const py = clamp(py0, 0, Math.max(0, vh - 1));
                                const el = document.elementFromPoint(px, py);
                                if (!el) continue;
                                const t = tryExtract(el);
                                if (t) return t;
                            }

                            // 3.2 附近区域搜索（限定坐标附近，避免命中推荐流）
                            const els = Array.from(document.querySelectorAll('div, span, p, a'));
                            const candidates = [];
                            for (const el of els) {
                                if (!el || el.offsetParent === null) continue;
                                const t = (el.innerText || el.textContent || '').trim();
                                if (!want(t)) continue;
                                const r = el.getBoundingClientRect();
                                if (!r || r.width < 1 || r.height < 1) continue;
                                if (r.top < cy - 160 || r.top > cy + 160) continue;
                                if (r.left < cx - 600 || r.left > cx + 600) continue;
                                const dx = (r.left + r.width / 2) - cx;
                                const dy = (r.top + r.height / 2) - cy;
                                candidates.push({ t, d: dx * dx + dy * dy });
                            }
                            candidates.sort((a, b) => a.d - b.d);
                            return candidates.length ? candidates[0].t : '';
                        }""",
                        list(probe),
                    ) or ""
                    if fans_text:
                        break
                    page.wait_for_timeout(300)
            except Exception:
                fans_text = ""

            # 4) 悬浮卡片容器兜底
            if not fans_text:
                try:
                    fans_text = page.evaluate(
                        r"""() => {
                            const cardSelectors = [
                                '.user-card-container',
                                '.user-info-card',
                                '[class*="UserCard"]',
                                '[class*="popup"]',
                                '[class*="popover"]',
                                '[class*="tooltip"]',
                                '.author-card',
                                '[class*="user-card"]',
                            ];
                            for (const sel of cardSelectors) {
                                const cards = document.querySelectorAll(sel);
                                for (const card of cards) {
                                    if (!card || card.offsetParent === null) continue;
                                    const t = (card.innerText || '').trim();
                                    if (t && t.includes('粉丝') && /\\d/.test(t)) return t;
                                }
                            }
                            return '';
                        }""",
                    ) or ""
                except Exception:
                    fans_text = ""

            if not fans_text:
                return ""

            patterns = [
                r"(粉丝|关注者|followers)[:\\s：]*([0-9]+\\.?[0-9]*[万wWkK]?)",
                r"([0-9]+\\.?[0-9]*[万wWkK]?)\\s*(粉丝|关注者|followers)",
            ]
            for pattern in patterns:
                match = re.search(pattern, str(fans_text), re.IGNORECASE)
                if not match:
                    continue
                value = match.group(1) if re.match(r"[0-9]", match.group(1) or "") else match.group(2)
                value = (value or "").strip()
                if value:
                    return value.replace("W", "万").replace("w", "万")

            return ""
        except Exception:
            return ""

    def _simulate_realistic_hover(self, page, box):
        r'''
                    self.log(f"  [悬停] 尝试通过作者名称文本定位: {author_name}", 'info')

                    # 使用Playwright的文本定位策略
                    text_selectors = [
                        f'text="{author_name}"',
                        f'[class*="author"]:has-text("{author_name}")',
                        f'[class*="name"]:has-text("{author_name}")',
                        f'.user-name:has-text("{author_name}")',
                        f'.nickname:has-text("{author_name}")',
                        f'span:has-text("{author_name}")'
                    ]

                    for selector in text_selectors:
                        try:
                            self.log(f"    [定位] 文本选择器: {selector}", 'info')
                            element = page.wait_for_selector(selector, timeout=3000)

                            if element:
                                is_visible = element.is_visible()
                                box = element.bounding_box()

                                debug_info["elements_found"].append({
                                    "selector": selector,
                                    "visible": is_visible,
                                    "bounding_box": box
                                })

                                if is_visible and box and box.width > 10 and box.height > 10:
                                    author_element = element
                                    selection_method = f"作者名称文本定位 ({selector})"
                                    self.log(f"  [悬停] ✓ 找到名称元素: {selector}", 'success')
                                    break

                        except Exception as e:
                            self.log(f"    [定位] 文本选择器失败 {selector}: {str(e)}", 'warning')
                            continue

                    if author_element:
                        self.log(f"  [悬停] {selection_method}", 'success')

                except Exception as e:
                    self.log(f"  [悬停] 作者名称定位失败: {str(e)}", 'error')

            # 3. 智能元素搜索（Playwright高级定位）
            if not author_element:
                try:
                    self.log("  [悬停] 开始智能元素搜索...", 'info')

                    # 使用Playwright的JavaScript执行进行高级搜索
                    js_search_result = page.evaluate("""() => {
                        const elements = [];

                        // 搜索所有可能的用户链接
                        const userLinks = Array.from(document.querySelectorAll('a[href*="/user/profile/"]'));

                        for (const link of userLinks) {
                            const rect = link.getBoundingClientRect();
                            if (rect.width > 10 && rect.height > 10 &&
                                rect.top > -200 && rect.top < window.innerHeight + 200) {

                                const isVisible = link.offsetParent !== null;
                                const text = (link.innerText || link.textContent || '').trim();

                                elements.push({
                                    selector: 'a[href*="/user/profile/"]',
                                    visible: isVisible,
                                    text: text,
                                    rect: {
                                        top: rect.top,
                                        left: rect.left,
                                        width: rect.width,
                                        height: rect.height
                                    },
                                    html: link.outerHTML.substring(0, 200)
                                });
                            }
                        }

                        // 按位置排序，优先选择页面顶部的元素
                        elements.sort((a, b) => a.rect.top - b.rect.top);

                        return elements.slice(0, 5); // 返回前5个最可能的元素
                    }""")

                    if js_search_result:
                        self.log(f"  [悬停] JavaScript搜索找到 {len(js_search_result)} 个候选元素", 'info')

                        for idx, element_info in enumerate(js_search_result):
                            self.log(f"    [搜索] 元素{idx+1}: text='{element_info.get('text', '')}' rect={element_info.get('rect')}", 'info')

                            # 尝试通过元素信息重新定位
                            try:
                                # 使用文本内容定位
                                text = element_info.get('text', '').strip()
                                if text and len(text) < 50:
                                    text_element = page.wait_for_selector(f'text="{text}"', timeout=2000)
                                    if text_element and text_element.is_visible():
                                        author_element = text_element
                                        selection_method = f"JavaScript智能定位 (文本: {text})"
                                        self.log(f"  [悬停] ✓ 通过文本重新定位成功: {text}", 'success')
                                        break
                            except Exception:
                                continue

                            if idx == 0:  # 如果是第一个元素，尝试直接定位
                                try:
                                    first_element = page.query_selector('a[href*="/user/profile/"]')
                                    if first_element and first_element.is_visible():
                                        author_element = first_element
                                        selection_method = "JavaScript智能定位 (首个用户链接)"
                                        self.log(f"  [悬停] ✓ 定位到首个用户链接", 'success')
                                        break
                                except Exception:
                                    continue

                    if author_element:
                        self.log(f"  [悬停] {selection_method}", 'success')
                    else:
                        self.log("  [悬停] JavaScript智能定位未找到合适元素", 'warning')

                except Exception as e:
                    self.log(f"  [悬停] 智能搜索失败: {str(e)}", 'error')

            # 4. 如果仍未找到元素，尝试智能查找
            if not author_element:
                try:
                    self.log("  [悬停] 尝试智能查找用户元素...", 'info')
                    # 使用JavaScript进行更智能的查找
                    author_element = page.evaluate("""() => {
                        // 查找可能的用户链接
                        const userLinks = Array.from(document.querySelectorAll('a[href*="/user/profile/"]'))
                            .filter(a => {
                                const rect = a.getBoundingClientRect();
                                return rect.width > 10 && rect.height > 10 &&
                                       rect.top > -100 && rect.top < 1200;
                            })
                            .sort((a, b) => a.getBoundingClientRect().top - b.getBoundingClientRect().top);

                        if (userLinks.length > 0) {
                            const link = userLinks[0];
                            // 查找链接内部或相邻的文本元素
                            const textElement = link.querySelector('[class*="name"], [class*="nick"], .name') || link;
                            if (textElement && textElement.textContent.trim()) {
                                return textElement.tagName + '.' + textElement.className;
                            }
                        }

                        // 查找包含用户相关class的元素
                        const userElements = Array.from(document.querySelectorAll('[class*="user"], [class*="author"], [class*="creator"]'))
                            .filter(el => {
                                const rect = el.getBoundingClientRect();
                                return rect.width > 10 && rect.height > 10 &&
                                       rect.top > -100 && rect.top < 1200 &&
                                       el.textContent.trim().length > 0 && el.textContent.trim().length < 50;
                            })
                            .sort((a, b) => a.getBoundingClientRect().top - b.getBoundingClientRect().top);

                        return userElements.length > 0 ? userElements[0].tagName + '.' + userElements[0].className : '';
                    }""")

                    if author_element:
                        try:
                            # 尝试根据返回的选择器找到元素
                            element = page.query_selector(author_element)
                            if element and element.is_visible():
                                author_element = element
                                selection_method = f"智能查找 ({author_element})"
                                self.log(f"  [悬停] {selection_method}", 'info')
                        except Exception:
                            author_element = None

                except Exception as e:
                    self.log(f"  [悬停] 智能查找失败: {str(e)}", 'warning')

            # 4. 增强的悬停执行（Playwright精确定位）
            if not author_element:
                self.log("  [悬停] 未找到有效的作者元素，返回失败", 'error')
                return ""

            try:
                self.log("  [悬停] 开始增强悬停操作...", 'info')

                # 4.1 确保元素在视窗中并滚动到元素
                try:
                    self.log("  [悬停] 滚动到元素位置...", 'info')
                    author_element.scroll_into_view_if_needed(timeout=5000)
                    time.sleep(1)  # 等待滚动完成
                    self.log("  [悬停] ✓ 滚动完成", 'success')
                except Exception as e:
                    self.log(f"  [悬停] 滚动失败: {str(e)}", 'warning')

                # 4.2 获取元素精确位置
                box = author_element.bounding_box()
                if not box:
                    self.log("  [悬停] 无法获取元素位置", 'error')
                    return ""

                self.log(f"  [悬停] 元素位置: x={box['x']}, y={box['y']}, width={box['width']}, height={box['height']}", 'info')

                # 4.3 多种悬停策略（Playwright增强版）
                hover_strategies = [
                    {
                        "name": "内置hover方法",
                        "method": lambda: author_element.hover(timeout=5000),
                        "delay": 0.5
                    },
                    {
                        "name": "鼠标移动到中心点",
                        "method": lambda: page.mouse.move(
                            box["x"] + box["width"] / 2,
                            box["y"] + box["height"] / 2
                        ),
                        "delay": 0.8
                    },
                    {
                        "name": "鼠标移动到偏移位置（避免过于精确）",
                        "method": lambda: page.mouse.move(
                            box["x"] + box["width"] / 2 + random.randint(-3, 3),
                            box["y"] + box["height"] / 2 + random.randint(-3, 3)
                        ),
                        "delay": 1.0
                    },
                    {
                        "name": "分步移动（模拟真实用户行为）",
                        "method": lambda: self._simulate_realistic_hover(page, box),
                        "delay": 1.5
                    }
                ]

                hover_success = False
                for strategy in hover_strategies:
                    try:
                        self.log(f"  [悬停] 尝试策略: {strategy['name']}", 'info')
                        strategy["method"]()
                        time.sleep(strategy["delay"])
                        hover_success = True
                        self.log(f"  [悬停] ✓ {strategy['name']} 成功", 'success')
                        break
                    except Exception as e:
                        self.log(f"  [悬停] {strategy['name']} 失败: {str(e)}", 'warning')
                        continue

                if not hover_success:
                    self.log("  [悬停] 所有悬停策略均失败", 'error')
                    return ""

                # 4.4 智能等待悬浮卡片出现
                self.log("  [悬停] 等待悬浮卡片出现...", 'info')
                card_appeared = self._wait_for_hover_card_with_retry(page)
                if not card_appeared:
                    self.log("  [悬停] 未检测到悬浮卡片", 'warning')
                else:
                    self.log("  [悬停] ✓ 检测到悬浮卡片", 'success')

                # 4.5 增强的粉丝数提取（Playwright精确定位）
                self.log("  [悬停] 开始提取粉丝数...", 'info')
                fans_data = self._extract_fans_with_enhanced_detection(page)

                # MCP同步：记录详细的粉丝数提取调试信息
                self.log_fans_extraction_debug(fans_data, page)

                if fans_data and fans_data.get('fans'):
                    fans_number = fans_data['fans']
                    confidence = fans_data.get('confidence', 0)
                    method = fans_data.get('method', '')

                    self.log(f"  [悬停] ✓ 提取到粉丝数: {fans_number} (置信度: {confidence}% | 方法: {method})", 'success')

                    # 解析粉丝数
                    if fans_number:
                        import re
                        patterns = [
                            r"(粉丝|关注者|followers)[:\s：]*([0-9]+\.?[0-9]*[万wWkK]?)",
                            r"([0-9]+\.?[0-9]*[万wWkK]?)\s*(粉丝|关注者|followers)",
                            r"粉丝[：:\s]*([0-9]+\.?[0-9]*[万wWkK]?)",
                        ]

                        for pattern in patterns:
                            match = re.search(pattern, fans_number, re.IGNORECASE)
                            if match:
                                if re.match(r"[0-9]", match.group(1)):
                                    fans_result = match.group(1)
                                else:
                                    fans_result = match.group(2)

                                fans_result = self.parse_fans_number(fans_result)
                                if fans_result and fans_result != "0":
                                    self.log(f"  [悬停] ✓ 粉丝数解析成功: {fans_result} (原始: {fans_number})", 'success')
                                    return fans_result

                self.log("  [悬停] 未能提取到有效的粉丝数", 'warning')
                return ""

            except Exception as e:
                self.log(f"  [悬停] 悬停操作异常: {str(e)}", 'error')
                return ""

        except Exception as e:
            self.log(f"  [悬停] 整体流程失败: {str(e)}", 'error')
            return ""

        '''

    def _simulate_realistic_hover(self, page, box):
        """模拟真实用户的悬停行为（带详细可视化调试）"""
        try:
            # 创建调试截屏目录
            debug_dir = os.path.join(self.output_dir.get(), "debug_screenshots")
            os.makedirs(debug_dir, exist_ok=True)

            # 分步移动到目标位置
            start_x, start_y = page.viewport_size["width"] // 2, page.viewport_size["height"] // 2
            target_x = box["x"] + box["width"] / 2 + random.randint(-2, 2)
            target_y = box["y"] + box["height"] / 2 + random.randint(-2, 2)

            self.log(f"    [定位] 鼠标移动路径: 从({start_x}, {start_y}) → 目标({target_x:.1f}, {target_y:.1f})", 'info')
            self.log(f"    [定位] 目标元素大小: {box['width']:.1f} x {box['height']:.1f}", 'info')

            # 添加视觉标记
            try:
                # 1. 高亮目标区域
                page.evaluate(f"""
                    () => {{
                        const highlight = document.createElement('div');
                        highlight.style.position = 'fixed';
                        highlight.style.left = '{box["x"] - 2}px';
                        highlight.style.top = '{box["y"] - 2}px';
                        highlight.style.width = '{box["width"] + 4}px';
                        highlight.style.height = '{box["height"] + 4}px';
                        highlight.style.border = '3px solid red';
                        highlight.style.backgroundColor = 'rgba(255, 0, 0, 0.2)';
                        highlight.style.zIndex = '10000';
                        highlight.style.pointerEvents = 'none';
                        highlight.id = 'hover_highlight';
                        document.body.appendChild(highlight);

                        // 2. 添加目标中心点标记
                        const center = document.createElement('div');
                        center.style.position = 'fixed';
                        center.style.left = '{target_x - 8}px';
                        center.style.top = '{target_y - 8}px';
                        center.style.width = '16px';
                        center.style.height = '16px';
                        center.style.borderRadius = '50%';
                        center.style.backgroundColor = 'red';
                        center.style.border = '2px solid white';
                        center.style.zIndex = '10001';
                        center.style.pointerEvents = 'none';
                        center.innerHTML = '🎯';
                        center.style.fontSize = '12px';
                        center.style.textAlign = 'center';
                        center.style.lineHeight = '16px';
                        center.id = 'hover_target';
                        document.body.appendChild(center);
                    }}
                """)
                self.log("    [可视化] ✓ 添加红色高亮框和目标标记 🎯", 'success')

                # 截屏：标记状态
                marker_screenshot = os.path.join(debug_dir, "mcp_visual_markers.png")
                page.screenshot(path=marker_screenshot, full_page=False)
                self.log(f"    [MCP截屏] 视觉标记截图保存: mcp_visual_markers.png", 'info')

            except Exception as e:
                self.log(f"    [可视化] 添加视觉标记失败: {str(e)}", 'warning')

            # 3. 分3-4步移动到目标，每步都添加轨迹点
            steps = random.randint(3, 4)
            self.log(f"    [鼠标] 开始分{steps}步移动到目标", 'info')

            for i in range(1, steps + 1):
                progress = i / steps
                current_x = start_x + (target_x - start_x) * progress
                current_y = start_y + (target_y - start_y) * progress

                # 添加轨迹点
                try:
                    page.evaluate(f"""
                        () => {{
                            const trail = document.createElement('div');
                            trail.style.position = 'fixed';
                            trail.style.left = '{current_x - 4}px';
                            trail.style.top = '{current_y - 4}px';
                            trail.style.width = '8px';
                            trail.style.height = '8px';
                            trail.style.borderRadius = '50%';
                            trail.style.backgroundColor = 'rgba(0, 255, 0, 0.8)';
                            trail.style.border = '1px solid white';
                            trail.style.zIndex = '9999';
                            trail.style.pointerEvents = 'none';
                            trail.innerHTML = '{i}';
                            trail.style.fontSize = '6px';
                            trail.style.textAlign = 'center';
                            trail.style.lineHeight = '8px';
                            trail.style.color = 'white';
                            trail.id = 'trail_{i}';
                            document.body.appendChild(trail);
                        }}
                    """)
                except Exception:
                    pass

                # 移动鼠标
                page.mouse.move(current_x, current_y)
                self.log(f"    [鼠标] 步骤{i}: 移动到({current_x:.1f}, {current_y:.1f})", 'debug')
                time.sleep(random.uniform(0.2, 0.4))

                # 每步都截图
                try:
                    step_screenshot = os.path.join(debug_dir, f"mcp_mouse_step_{i}.png")
                    page.screenshot(path=step_screenshot, full_page=False)
                    self.log(f"    [MCP截屏] 鼠标步骤{i}截图保存", 'debug')
                except Exception:
                    pass

            # 最终悬停
            self.log("    [悬停] 执行最终悬停操作...", 'info')
            page.hover(f"xpath=//div[@style*='position: fixed'][@style*='left: {box['x'] - 2}px']")

            # 检查悬停后的视觉效果
            time.sleep(2)
            try:
                final_screenshot = os.path.join(debug_dir, "mcp_hover_complete.png")
                page.screenshot(path=final_screenshot, full_page=False)
                self.log(f"    [MCP截屏] 悬停完成截图保存: mcp_hover_complete.png", 'success')

                # 检查是否有新的弹出元素
                popup_elements = page.evaluate("""
                    () => {
                        const allElements = document.querySelectorAll('*');
                        const newElements = Array.from(allElements).filter(el => {
                            const style = window.getComputedStyle(el);
                            const rect = el.getBoundingClientRect();
                            const text = el.innerText || '';

                            return (style.position === 'fixed' || style.position === 'absolute') &&
                                   rect.width > 0 && rect.height > 0 &&
                                   (text.includes('粉丝') || text.includes('关注') ||
                                    el.className.includes('card') || el.className.includes('popup') ||
                                    el.className.includes('tooltip')) &&
                                   el.style.zIndex > 1000;
                        });

                        return newElements.map(el => ({
                            tagName: el.tagName,
                            className: el.className,
                            text: el.innerText.substring(0, 50),
                            rect: el.getBoundingClientRect().toString(),
                            zIndex: el.style.zIndex
                        }));
                    }
                """)

                self.log(f"    [检测] 发现 {len(popup_elements)} 个可能的悬浮元素", 'info')
                for i, elem in enumerate(popup_elements[:3]):  # 只显示前3个
                    self.log(f"      元素{i+1}: {elem['tagName']}.{elem['className']} - {elem['text']}", 'info')

                # 清理所有视觉标记
                page.evaluate("""
                    () => {
                        const elements = ['hover_highlight', 'hover_target', 'trail_1', 'trail_2', 'trail_3', 'trail_4'];
                        elements.forEach(id => {
                            const el = document.getElementById(id);
                            if (el) el.remove();
                        });
                    }
                """)
                self.log("    [清理] 已移除所有视觉标记", 'info')

            except Exception as e:
                self.log(f"    [检测] 悬停后检查失败: {str(e)}", 'warning')

        except Exception as e:
            self.log(f"    [鼠标] 可视化悬停失败: {str(e)}", 'error')
            # 如果可视化失败，回退到简单移动
            try:
                page.mouse.move(target_x, target_y)
                page.hover(f"xpath=//div[@style*='position: fixed'][@style*='left: {box['x'] - 2}px']")
            except Exception:
                pass

    def add_crosshair_cursor(self, page):
        """添加十字光标来显示鼠标的确切位置"""
        try:
            debug_dir = os.path.join(self.output_dir.get(), "debug_screenshots")
            os.makedirs(debug_dir, exist_ok=True)

            # 创建自定义光标
            page.evaluate("""
                () => {
                    // 创建十字光标
                    const cursor = document.createElement('div');
                    cursor.id = 'mouse_crosshair';
                    cursor.style.position = 'fixed';
                    cursor.style.width = '20px';
                    cursor.style.height = '20px';
                    cursor.style.border = '2px solid red';
                    cursor.style.borderRadius = '50%';
                    cursor.style.backgroundColor = 'rgba(255, 0, 0, 0.3)';
                    cursor.style.zIndex = '99999';
                    cursor.style.pointerEvents = 'none';
                    cursor.style.display = 'none';
                    cursor.style.transform = 'translate(-50%, -50%)';

                    // 添加十字线
                    const horizontal = document.createElement('div');
                    horizontal.style.position = 'absolute';
                    horizontal.style.width = '30px';
                    horizontal.style.height = '1px';
                    horizontal.style.backgroundColor = 'red';
                    horizontal.style.top = '50%';
                    horizontal.style.left = '50%';
                    horizontal.style.transform = 'translate(-50%, -50%)';

                    const vertical = document.createElement('div');
                    vertical.style.position = 'absolute';
                    vertical.style.width = '1px';
                    vertical.style.height = '30px';
                    vertical.style.backgroundColor = 'red';
                    vertical.style.top = '50%';
                    vertical.style.left = '50%';
                    vertical.style.transform = 'translate(-50%, -50%)';

                    cursor.appendChild(horizontal);
                    cursor.appendChild(vertical);
                    document.body.appendChild(cursor);

                    // 添加坐标显示
                    const coords = document.createElement('div');
                    coords.id = 'mouse_coords';
                    coords.style.position = 'fixed';
                    coords.style.top = '10px';
                    coords.style.right = '10px';
                    coords.style.backgroundColor = 'rgba(0, 0, 0, 0.8)';
                    coords.style.color = 'white';
                    coords.style.padding = '5px 10px';
                    coords.style.borderRadius = '3px';
                    coords.style.fontSize = '12px';
                    coords.style.fontFamily = 'monospace';
                    coords.style.zIndex = '99998';
                    coords.innerHTML = 'Mouse: (0, 0)';
                    document.body.appendChild(coords);

                    // 监听鼠标移动
                    document.addEventListener('mousemove', (e) => {
                        const cursor = document.getElementById('mouse_crosshair');
                        const coords = document.getElementById('mouse_coords');
                        if (cursor && coords) {
                            cursor.style.left = e.clientX + 'px';
                            cursor.style.top = e.clientY + 'px';
                            cursor.style.display = 'block';
                            coords.innerHTML = `Mouse: (${e.clientX}, ${e.clientY})`;
                        }
                    });
                }
            """)
            self.log("    [光标] ✓ 已添加鼠标十字光标和坐标显示", 'success')

            # 截屏：光标状态
            cursor_screenshot = os.path.join(debug_dir, "mcp_cursor_ready.png")
            page.screenshot(path=cursor_screenshot, full_page=False)
            self.log(f"    [MCP截屏] 光标准备状态截图: mcp_cursor_ready.png", 'info')

        except Exception as e:
            self.log(f"    [光标] 添加十字光标失败: {str(e)}", 'warning')

    def show_elements_with_coordinates(self, page):
        """显示所有可悬停元素的坐标"""
        try:
            debug_dir = os.path.join(self.output_dir.get(), "debug_screenshots")
            os.makedirs(debug_dir, exist_ok=True)

            elements_info = page.evaluate("""
                () => {
                    const elements = document.querySelectorAll('*');
                    const hoverableElements = [];

                    elements.forEach((el, index) => {
                        const rect = el.getBoundingClientRect();
                        const text = (el.innerText || el.textContent || '').substring(0, 30);
                        const className = el.className || '';
                        const hasClickHandler = el.onclick || el.onmouseover || el.onmouseenter;
                        const isInteractive = el.tagName === 'A' || el.tagName === 'BUTTON' ||
                                             el.role === 'button' || el.style.cursor === 'pointer' ||
                                             hasClickHandler;

                        if (rect.width > 10 && rect.height > 10 &&
                            (text.length > 0 || isInteractive || className.includes('user'))) {
                            hoverableElements.push({
                                index: index,
                                tagName: el.tagName,
                                className: className.substring(0, 50),
                                text: text,
                                x: Math.round(rect.left + rect.width / 2),
                                y: Math.round(rect.top + rect.height / 2),
                                width: Math.round(rect.width),
                                height: Math.round(rect.height),
                                isInteractive: isInteractive,
                                zIndex: window.getComputedStyle(el).zIndex
                            });
                        }
                    });

                    // 只返回前50个最相关的元素
                    return hoverableElements
                        .sort((a, b) => b.isInteractive - a.isInteractive)
                        .slice(0, 50);
                }
            """)

            self.log(f"    [元素] 发现 {len(elements_info)} 个可悬停元素", 'info')

            # 标记前10个元素
            marked_count = 0
            for i, elem in enumerate(elements_info[:10]):
                try:
                    page.evaluate(f"""
                        () => {{
                            const marker = document.createElement('div');
                            marker.style.position = 'fixed';
                            marker.style.left = '{elem['x'] - 15}px';
                            marker.style.top = '{elem['y'] - 15}px';
                            marker.style.width = '30px';
                            marker.style.height = '30px';
                            marker.style.border = '2px solid blue';
                            marker.style.borderRadius = '50%';
                            marker.style.backgroundColor = 'rgba(0, 0, 255, 0.2)';
                            marker.style.zIndex = '10000';
                            marker.style.pointerEvents = 'none';
                            marker.style.fontSize = '10px';
                            marker.style.textAlign = 'center';
                            marker.style.lineHeight = '30px';
                            marker.style.color = 'blue';
                            marker.style.fontWeight = 'bold';
                            marker.innerHTML = '{i + 1}';
                            marker.id = 'elem_marker_{i}';
                            document.body.appendChild(marker);
                        }}
                    """)
                    marked_count += 1
                except Exception:
                    pass

            self.log(f"    [标记] 已标记 {marked_count} 个候选元素", 'success')

            # 显示元素信息
            for i, elem in enumerate(elements_info[:10]):
                interactive_text = "✓" if elem['isInteractive'] else "✗"
                self.log(f"      元素{i+1}: {elem['tagName']} at ({elem['x']}, {elem['y']}) 交互:{interactive_text} - {elem['text']}", 'info')

            # 截屏：元素标记状态
            elements_screenshot = os.path.join(debug_dir, "mcp_elements_marked.png")
            page.screenshot(path=elements_screenshot, full_page=False)
            self.log(f"    [MCP截屏] 元素标记截图保存: mcp_elements_marked.png", 'info')

            return elements_info

        except Exception as e:
            self.log(f"    [元素] 显示元素坐标失败: {str(e)}", 'warning')
            return []

    def _wait_for_hover_card_with_retry(self, page, max_wait_ms=8000):
        """等待悬浮卡片出现（带MCP同步截屏调试）"""
        try:
            # 检查间隔
            check_interval = 300
            total_waited = 0

            # 创建调试截屏目录
            debug_dir = os.path.join(self.output_dir.get(), "debug_screenshots")
            os.makedirs(debug_dir, exist_ok=True)

            while total_waited < max_wait_ms:
                time.sleep(check_interval / 1000)
                total_waited += check_interval

                # MCP同步：定期截屏调试
                if total_waited % 2000 == 0:
                    try:
                        screenshot_name = f"mcp_hover_debug_{total_waited//1000}s.png"
                        screenshot_path = os.path.join(debug_dir, screenshot_name)

                        # 截取当前页面状态
                        page.screenshot(path=screenshot_path, full_page=False)
                        self.log(f"    [MCP截屏] 调试截图保存: {screenshot_name}", 'info')

                        # 获取页面元素状态
                        element_status = page.evaluate("""() => {
                            const allElements = document.querySelectorAll('*');
                            const visibleElements = Array.from(allElements).filter(el => {
                                const rect = el.getBoundingClientRect();
                                return rect.width > 0 && rect.height > 0 && el.offsetParent !== null;
                            });

                            const userElements = visibleElements.filter(el => {
                                const text = (el.innerText || el.textContent || '').toLowerCase();
                                const className = (el.className || '').toLowerCase();
                                return text.includes('粉丝') || text.includes('关注') ||
                                       className.includes('user') || className.includes('card');
                            });

                            return {
                                totalVisible: visibleElements.length,
                                userRelatedElements: userElements.length,
                                userElementsPreview: userElements.slice(0, 5).map(el => ({
                                    tag: el.tagName,
                                    className: el.className,
                                    text: (el.innerText || '').substring(0, 50)
                                }))
                            };
                        }""")

                        self.log(f"    [MCP状态] 页面元素: 可见{element_status.get('totalVisible')}个, 用户相关{element_status.get('userRelatedElements')}个", 'info')

                        for preview in element_status.get('userElementsPreview', []):
                            self.log(f"    [MCP元素] {preview.get('tag')}.{preview.get('className')} - {preview.get('text')}", 'debug')

                    except Exception as e:
                        self.log(f"    [MCP截屏] 截屏失败: {str(e)}", 'warning')

                # 检查悬浮卡片
                card_detected = page.evaluate("""() => {
                    const selectors = [
                        '.user-card-container',
                        '.user-info-card',
                        '[class*="UserCard"]',
                        '[class*="user-card"]',
                        '[class*="popup"]',
                        '[class*="popover"]',
                        '[class*="tooltip"]',
                        '.author-card',
                        '.user-tooltip'
                    ];

                    for (const sel of selectors) {
                        const elements = document.querySelectorAll(sel);
                        for (const el of elements) {
                            if (el.offsetParent !== null &&
                                (el.innerText.includes('粉丝') || el.innerText.includes('关注'))) {
                                return {
                                    found: true,
                                    selector: sel,
                                    text: el.innerText.substring(0, 100),
                                    visible: true,
                                    rect: el.getBoundingClientRect().toString(),
                                    innerHTML: el.innerHTML.substring(0, 200)
                                };
                            }
                        }
                    }
                    return {found: false};
                }""")

                if card_detected.get('found'):
                    self.log(f"    [MCP检测] 检测到卡片: {card_detected.get('selector')}", 'success')
                    self.log(f"    [MCP检测] 卡片文本: {card_detected.get('text', '')[:50]}...", 'success')
                    self.log(f"    [MCP检测] 卡片位置: {card_detected.get('rect')}", 'info')

                    # 保存成功检测到的卡片截图
                    try:
                        success_screenshot = os.path.join(debug_dir, "mcp_hover_success.png")
                        page.screenshot(path=success_screenshot, full_page=False)
                        self.log(f"    [MCP截屏] 成功截图保存: mcp_hover_success.png", 'success')
                    except Exception as e:
                        self.log(f"    [MCP截屏] 成功截图失败: {str(e)}", 'warning')

                    return True
                else:
                    self.log(f"    [MCP检测] 第{total_waited//1000}秒: 未检测到悬浮卡片", 'warning')

                # 偶尔移动鼠标一点，防止超时
                if total_waited % 2000 == 0:
                    try:
                        current_pos = page.mouse.position
                        page.mouse.move(current_pos["x"] + 1, current_pos["y"] + 1)
                        time.sleep(0.1)
                        page.mouse.move(current_pos["x"], current_pos["y"])
                    except Exception:
                        pass

            # 保存最终失败状态的截图
            try:
                final_screenshot = os.path.join(debug_dir, "mcp_hover_failed.png")
                page.screenshot(path=final_screenshot, full_page=False)
                self.log(f"    [MCP截屏] 失败状态截图保存: mcp_hover_failed.png", 'warning')
            except Exception as e:
                self.log(f"    [MCP截屏] 失败状态截图失败: {str(e)}", 'warning')

            return False

        except Exception as e:
            self.log(f"    [MCP检测] 等待悬浮卡片异常: {str(e)}", 'error')
            return False

    def _extract_fans_with_enhanced_detection(self, page):
        """增强的粉丝数提取（带MCP同步调试）"""
        try:
            # 创建调试截屏目录
            debug_dir = os.path.join(self.output_dir.get(), "debug_screenshots")
            os.makedirs(debug_dir, exist_ok=True)

            return page.evaluate("""() => {
                const result = {
                    fans: '',
                    confidence: 0,
                    method: '',
                    allText: '',
                    debug: []
                };

                // MCP调试：记录所有可能的粉丝相关元素
                const allElements = document.querySelectorAll('*');
                const fansRelatedElements = Array.from(allElements).filter(el => {
                    const text = (el.innerText || el.textContent || '').toLowerCase();
                    const className = (el.className || '').toLowerCase();
                    return text.includes('粉丝') || text.includes('关注') ||
                           className.includes('fans') || className.includes('follower');
                });

                result.debug.push(`找到 ${fansRelatedElements.length} 个粉丝相关元素`);

                // 记录每个相关元素的详细信息
                fansRelatedElements.forEach((el, index) => {
                    const rect = el.getBoundingClientRect();
                    const isVisible = rect.width > 0 && rect.height > 0 && el.offsetParent !== null;
                    const text = (el.innerText || el.textContent || '').substring(0, 100);
                    const className = el.className || '';

                    result.debug.push(`元素${index}: ${el.tagName}.${className} - 可见:${isVisible} - ${text}`);
                });

                // 1. 从悬浮卡片中提取
                const cardSelectors = [
                    '.user-card-container',
                    '.user-info-card',
                    '[class*="UserCard"]',
                    '[class*="user-card"]',
                    '.author-card'
                ];

                for (const sel of cardSelectors) {
                    const cards = document.querySelectorAll(sel);
                    result.debug.push(`检查卡片选择器 ${sel}: 找到 ${cards.length} 个卡片`);

                    for (const card of cards) {
                        if (card.offsetParent === null) {
                            result.debug.push(`卡片不可见，跳过`);
                            continue;
                        }

                        const text = (card.innerText || '').trim();
                        result.allText += text + ' | ';
                        result.debug.push(`卡片内容: ${text.substring(0, 100)}...`);

                        if (text.includes('粉丝') && /\\d/.test(text)) {
                            result.fans = text;
                            result.confidence = 90;
                            result.method = 'card';
                            result.debug.push(`卡片匹配成功: ${sel}`);
                            return result;
                        }
                    }
                }

                // 2. 从弹出层中提取
                const popupSelectors = [
                    '[class*="popup"]',
                    '[class*="popover"]',
                    '[class*="tooltip"]'
                ];

                for (const sel of popupSelectors) {
                    const elements = document.querySelectorAll(sel);
                    for (const el of elements) {
                        if (el.offsetParent === null) continue;

                        const text = (el.innerText || '').trim();
                        result.allText += text + ' | ';
                        result.debug.push(`检查弹出层 ${sel}: ${text.substring(0, 50)}...`);

                        if (text.includes('粉丝') && /\\d/.test(text)) {
                            result.fans = text;
                            result.confidence = 70;
                            result.method = 'popup';
                            result.debug.push(`弹出层匹配成功: ${sel}`);
                            return result;
                        }
                    }
                }

                // 3. 从页面顶部用户信息区域提取
                const topSelectors = [
                    '.user-info',
                    '.author-info',
                    '[class*="UserInfo"]',
                    '.side-bar'
                ];

                for (const sel of topSelectors) {
                    const elements = document.querySelectorAll(sel);
                    for (const el of elements) {
                        if (el.offsetParent === null) continue;

                        const rect = el.getBoundingClientRect();
                        if (rect.top > 200) continue; // 只考虑页面顶部

                        const text = (el.innerText || '').trim();
                        result.allText += text + ' | ';
                        result.debug.push(`检查顶部区域 ${sel}: ${text.substring(0, 50)}...`);

                        if (text.includes('粉丝') && /\\d/.test(text)) {
                            result.fans = text;
                            result.confidence = 50;
                            result.method = 'top';
                            result.debug.push(`顶部区域匹配成功: ${sel}`);
                            return result;
                        }
                    }
                }

                result.debug.push('所有提取方法均失败');
                return result;
            }""")

        except Exception as e:
            self.log(f"    [提取] 增强检测失败: {str(e)}", 'error')
            return None

    def log_fans_extraction_debug(self, fans_data, page):
        """记录粉丝数提取的详细调试信息"""
        if not fans_data:
            return

        try:
            # 创建调试截屏目录
            debug_dir = os.path.join(self.output_dir.get(), "debug_screenshots")
            os.makedirs(debug_dir, exist_ok=True)

            # 记录调试信息
            debug_info = fans_data.get('debug', [])
            all_text = fans_data.get('allText', '')
            method = fans_data.get('method', '')
            confidence = fans_data.get('confidence', 0)

            self.log(f"    [MCP提取] 方法: {method}, 置信度: {confidence}%", 'info')
            self.log(f"    [MCP提取] 找到元素数: {len(debug_info)}", 'info')

            # 记录每个调试信息
            for i, debug_msg in enumerate(debug_info):
                if i < 10:  # 只记录前10条，避免日志过长
                    self.log(f"    [MCP调试] {debug_msg}", 'debug')
                elif i == 10:
                    self.log(f"    [MCP调试] ... 还有 {len(debug_info) - 10} 条调试信息", 'info')
                    break

            # 保存粉丝数提取调试截图
            if all_text:
                try:
                    extraction_screenshot = os.path.join(debug_dir, "mcp_fans_extraction.png")
                    page.screenshot(path=extraction_screenshot, full_page=False)
                    self.log(f"    [MCP截屏] 粉丝提取截图保存: mcp_fans_extraction.png", 'info')
                except Exception as e:
                    self.log(f"    [MCP截屏] 粉丝提取截图失败: {str(e)}", 'warning')

            # 保存提取到的文本内容到文件
            try:
                debug_text_file = os.path.join(debug_dir, "mcp_fans_text.txt")
                with open(debug_text_file, 'w', encoding='utf-8') as f:
                    f.write(f"提取时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\\n")
                    f.write(f"提取方法: {method}\\n")
                    f.write(f"置信度: {confidence}%\\n")
                    f.write(f"粉丝数据: {fans_data.get('fans', '')}\\n")
                    f.write(f"所有文本: {all_text}\\n")
                    f.write("\\n调试信息:\\n")
                    for debug_msg in debug_info:
                        f.write(f"{debug_msg}\\n")
                self.log(f"    [MCP调试] 调试文本保存: mcp_fans_text.txt", 'info')
            except Exception as e:
                self.log(f"    [MCP调试] 调试文本保存失败: {str(e)}", 'warning')

        except Exception as e:
            self.log(f"    [MCP调试] 调试信息记录失败: {str(e)}", 'error')

    def extract_detailed_content_ultimate(self, page, url):
        """v4.0终极版内容提取（兼容Pro v2字段）"""
        details = {
            "笔记ID": self.extract_note_id(url),
            "标题": "",
            "作者昵称": "",
            "作者ID": "",
            "粉丝数": "",
            "博主主页": "",
            "发布时间": "",
            "点赞数": "",
            "收藏数": "",
            "评论数": "",
            "分享数": "",
            "封面链接": "",
            "视频链接": "",
            "正文": "",
            "封面图数量": 0
        }
        image_urls = []
        note_id = details["笔记ID"]

        try:
            self.log("开始提取详细内容...", 'info')
            time.sleep(1.2)

            page_data = page.evaluate(
                r"""(noteId) => {
                    const pick = (...values) => {
                        for (const val of values) {
                            if (val !== undefined && val !== null && val !== '') return val;
                        }
                        return '';
                    };

                    const result = {
                        title: '',
                        authorName: '',
                        authorId: '',
                        authorHome: '',
                        fans: '',
                        publishTime: '',
                        likes: '',
                        collects: '',
                        comments: '',
                        shares: '',
                        coverUrl: '',
                        videoUrl: '',
                        content: '',
                        images: []
                    };

                    const globalState = window.__INITIAL_STATE__ || window.__REDUX_STATE__ || {};
                    const noteStore = globalState.note || globalState.noteDetail || (globalState.NoteView && globalState.NoteView.note) || {};
                    const detailMap = noteStore.noteDetailMap || noteStore.notes || noteStore.noteInfoMap || {};
                    let detail = null;
                    if (noteId && detailMap[noteId]) {
                        detail = detailMap[noteId];
                    } else {
                        const values = Object.values(detailMap);
                        if (values.length > 0) detail = values[0];
                    }

                    if (detail) {
                        const note = detail.note || detail;
                        const noteUserId = note.user_id || note.userId;
                        const user = detail.user || note.user || (globalState.user && globalState.user.userDetailMap && globalState.user.userDetailMap[noteUserId]) || {};
                        const stats = detail.noteStat || detail.noteStats || detail.interactInfo || note.interactInfo || detail.statistics || {};

                        result.title = (note.title || note.note_title || '').trim();
                        result.content = (note.desc || note.content || note.note_desc || '').trim();
                        result.publishTime = pick(note.time, note.publish_time, note.create_time, note.first_publish_time);

                        result.authorName = pick(user.nickname, user.nick_name, user.name);
                        const authorId = pick(user.user_id, user.userid, user.id, noteUserId);
                        if (authorId) {
                            result.authorId = authorId.toString();
                            result.authorHome = `https://www.xiaohongshu.com/user/profile/${authorId}`;
                        }

                        const fansValue = pick(user.fans, user.fans_num, user.fansNum, user.fansCount, user.followers);
                        if (fansValue !== '') result.fans = fansValue.toString();

                        result.likes = pick(stats.likes, stats.likeCount, stats.like_cnt, stats.likedCount, stats.like_num, stats.noteLikeCount, stats.liked_count);
                        result.collects = pick(
                            stats.collectCount, stats.collect_cnt, stats.collects, stats.collect_num, stats.collect_count, stats.collectCount,
                            stats.favoriteCount, stats.favorite_cnt, stats.favoritedCount, stats.favorited_count,
                            stats.saveCount, stats.savedCount, stats.collectedCount, stats.collected_count
                        );
                        result.comments = pick(stats.commentCount, stats.comment_cnt, stats.comments, stats.comment_num, stats.comment_count);
                        result.shares = pick(stats.shareCount, stats.share_cnt, stats.share, stats.share_num, stats.share_count);

                        const imageList = note.imageList || note.images || detail.imageList || [];
                        result.images = imageList
                            .map(img => (img && (img.url || img.url_default || img.img_url || img.original)) || '')
                            .filter(Boolean);
                        if (result.images.length > 0) result.coverUrl = result.images[0];

                        const video = note.video || note.video_info || note.videoInfo || note.video_info_v2 || note.videoInfoV2 || {};
                        const videoUrl = pick(video.url, video.play_url, video.playUrl, video.originUrl, video.origin_url);
                        if (videoUrl) result.videoUrl = videoUrl.toString();
                    }

                    return result;
                }""",
                note_id,
            ) or {}

            # 基础字段
            details["标题"] = (page_data.get("title") or "")[:200]
            details["正文"] = (page_data.get("content") or "")[:500]
            details["作者昵称"] = page_data.get("authorName") or ""
            details["作者ID"] = page_data.get("authorId") or ""
            details["博主主页"] = page_data.get("authorHome") or (f"https://www.xiaohongshu.com/user/profile/{details['作者ID']}" if details["作者ID"] else "")
            # 注意：不能用 `or ""`，否则 0 会被当成空值丢失
            details["粉丝数"] = str(page_data.get("fans", ""))
            details["发布时间"] = page_data.get("publishTime") or ""
            # 注意：不能用 `or ""`，否则 0 会被当成空值丢失
            details["点赞数"] = str(page_data.get("likes", ""))
            details["收藏数"] = str(page_data.get("collects", ""))
            details["评论数"] = str(page_data.get("comments", ""))
            details["分享数"] = str(page_data.get("shares", ""))
            details["封面链接"] = page_data.get("coverUrl") or ""
            details["视频链接"] = page_data.get("videoUrl") or ""

            image_urls = (page_data.get("images") or []) if isinstance(page_data, dict) else []
            # 标准化图片URL，保持顺序
            normalized_images = []
            for u in image_urls:
                normalized = self.normalize_media_url(page.url, u)
                if normalized:
                    normalized_images.append(normalized)
            image_urls = normalized_images

            if not details["封面链接"] and image_urls:
                details["封面链接"] = image_urls[0]
            details["封面图数量"] = len(image_urls)

            if image_urls:
                self.log(f"元数据提取成功 - 获取到{len(image_urls)}张图片", 'success')

            # DOM兜底：当元数据缺失/不完整时，从页面结构中补齐关键字段与图片URL
            try:
                dom_data = page.evaluate(r"""() => {
                    const pickText = (el) => (el && (el.innerText || el.textContent) || '').trim();
                    const pickAttr = (el, ...names) => {
                        if (!el) return '';
                        for (const name of names) {
                            const v = el.getAttribute(name);
                            if (v) return v;
                        }
                        return '';
                    };
                    const getImgSrc = (img) => pickAttr(img, 'src', 'data-src', 'data-original', 'data-origin', 'data-lazy-src');

                    const data = {
                        title: '',
                        authorName: '',
                        authorId: '',
                        authorHome: '',
                        fans: '',
                        publishTime: '',
                        likes: '',
                        collects: '',
                        comments: '',
                        shares: '',
                        coverUrl: '',
                        videoUrl: '',
                        content: '',
                        images: []
                    };

                    const titleEl = document.querySelector('#detail-title, [id*=\"title\"], .title, [class*=\"title\"]');
                    data.title = pickText(titleEl);

                    const authorEl = document.querySelector('.author-name, .username, .user-nickname, [class*=\"author\"][class*=\"name\"], [class*=\"nickname\"]');
                    data.authorName = pickText(authorEl);

                    // 作者链接：只取页面顶部可见的 profile 链接，避免命中推荐流/其他模块
                    const profileLinks = Array.from(document.querySelectorAll('a[href*=\"/user/profile/\"]'))
                        .filter(a => a && a.href && a.offsetParent !== null)
                        .map(a => {
                            const rect = a.getBoundingClientRect();
                            const text = ((a.innerText || a.textContent) || '').trim();
                            return { a, rect, text };
                        })
                        .filter(x => x.rect && x.rect.width > 10 && x.rect.height > 10 && x.rect.top > -200 && x.rect.top < 1400)
                        .sort((x, y) => x.rect.top - y.rect.top);

                    let authorLink = null;
                    if (profileLinks.length) {
                        const withText = profileLinks.filter(x => x.text && x.text.length <= 30 && !x.text.includes('关注'));
                        authorLink = (withText.length ? withText[0] : profileLinks[0]).a;
                    }

                    if (authorLink) {
                        data.authorHome = authorLink.href || pickAttr(authorLink, 'href') || '';
                        const match = (data.authorHome || '').match(/profile\\/([a-zA-Z0-9]+)/);
                        if (match) data.authorId = match[1];
                        if (!data.authorName) data.authorName = ((authorLink.innerText || authorLink.textContent) || '').trim();
                    }

                    const fansEl = document.querySelector('.fans-count, [class*=\"fans\"], .follower-count');
                    data.fans = pickText(fansEl);

                    const timeEl = document.querySelector('.publish-time, .date, [class*=\"time\"], [class*=\"date\"]');
                    data.publishTime = pickText(timeEl);

                    // 互动数据：优先从可能的按钮/计数区域读取
                    const interactEls = document.querySelectorAll('[class*=\"interact\"] span, [class*=\"count\"], button span, a span');
                    interactEls.forEach((el) => {
                        const text = pickText(el);
                        if (!text || !/[0-9]/.test(text)) return;
                        const className = (el.className || '').toString();
                        const parentClass = (el.closest('[class]') && el.closest('[class]').className || '').toString();
                        const hint = (className + ' ' + parentClass).toLowerCase();
                        if (!data.likes && (hint.includes('like') || hint.includes('zan'))) data.likes = text;
                        if (!data.collects && (hint.includes('collect') || hint.includes('favorite') || hint.includes('fav') || hint.includes('save'))) data.collects = text;
                        if (!data.comments && (hint.includes('comment') || hint.includes('reply'))) data.comments = text;
                        if (!data.shares && (hint.includes('share') || hint.includes('forward'))) data.shares = text;
                    });

                    // 文本兜底：从页面全文匹配“收藏/点赞/评论/分享”
                    const bodyText = (document.body && document.body.innerText) ? document.body.innerText : '';
                    const findCount = (label) => {
                        const patterns = [
                            new RegExp(label + '\\\\s*[:：]?\\\\s*([0-9]+\\\\.?[0-9]*[万wWkK]?)', 'i'),
                            new RegExp('([0-9]+\\\\.?[0-9]*[万wWkK]?)\\\\s*' + label, 'i'),
                        ];
                        for (const p of patterns) {
                            const m = bodyText.match(p);
                            if (m) return m[1];
                        }
                        return '';
                    };
                    if (!data.likes) data.likes = findCount('点赞');
                    if (!data.collects) data.collects = findCount('收藏');
                    if (!data.comments) data.comments = findCount('评论');
                    if (!data.shares) data.shares = findCount('分享');

                    const contentEl = document.querySelector('#detail-desc, .note-content, .content, [class*=\"desc\"], [class*=\"content\"]');
                    const contentText = pickText(contentEl);
                    if (contentText) data.content = contentText.slice(0, 500);

                    // 封面/图片列表
                    const coverImg = document.querySelector('img[src*=\"sns-img\"], img[src*=\"xhscdn\"], img[data-src*=\"sns-img\"], img[data-src*=\"xhscdn\"]');
                    if (coverImg) data.coverUrl = getImgSrc(coverImg) || '';

                    const videoEl = document.querySelector('video');
                    const sourceEl = videoEl ? videoEl.querySelector('source') : null;
                    data.videoUrl = (sourceEl && sourceEl.src) || (videoEl && videoEl.src) || '';

                    // 不在这里扫全站 img，避免把推荐流/其他模块图片当封面

                    return data;
                }""") or {}
            except Exception:
                dom_data = {}

            if dom_data:
                def set_if_empty(target_key, value, max_len=None):
                    if value is None:
                        return
                    text = str(value).strip()
                    if not text:
                        return
                    if max_len:
                        text = text[:max_len]
                    if not details.get(target_key):
                        details[target_key] = text

                set_if_empty("标题", dom_data.get("title"), 200)
                set_if_empty("正文", dom_data.get("content"), 500)
                set_if_empty("作者昵称", dom_data.get("authorName"))
                set_if_empty("作者ID", dom_data.get("authorId"))
                set_if_empty("博主主页", dom_data.get("authorHome"))
                # 注意：笔记详情页 DOM 中的“粉丝”容易误命中其他文本，这里不直接采用
                set_if_empty("发布时间", dom_data.get("publishTime"))
                set_if_empty("点赞数", dom_data.get("likes"))
                set_if_empty("收藏数", dom_data.get("collects"))
                set_if_empty("评论数", dom_data.get("comments"))
                set_if_empty("分享数", dom_data.get("shares"))
                set_if_empty("视频链接", dom_data.get("videoUrl"))

                if not details.get("博主主页") and details.get("作者ID"):
                    details["博主主页"] = f"https://www.xiaohongshu.com/user/profile/{details['作者ID']}"

                if not details.get("作者ID") and details.get("博主主页"):
                    match = re.search(r"/user/profile/([a-zA-Z0-9]+)", str(details.get("博主主页")))
                    if match:
                        details["作者ID"] = match.group(1)

                # 图片URL：只在元数据缺失时，从页面中提取“当前笔记”的图片（严格避免推荐流污染）
                if not image_urls:
                    try:
                        self.ensure_media_loaded(page)
                        dom_images = self.extract_note_image_urls_from_dom(page, limit=10)
                        if dom_images:
                            image_urls = dom_images
                            self.log(f"  [DOM] 获取到{len(image_urls)}张笔记图片", 'success')
                    except Exception:
                        pass

                if not details.get("封面链接") and image_urls:
                    details["封面链接"] = image_urls[0]
                details["封面图数量"] = len(image_urls)

            # 粉丝数兜底：重排序获取策略 + 增强错误处理
            fans_extraction_methods = [
                {
                    "name": "API获取",
                    "enabled": self.enable_api_fans,
                    "condition": lambda: details.get("作者ID"),
                    "method": lambda: self.fetch_user_fans_via_api(page, details.get("作者ID")),
                    "priority": 1,
                    "retry_on_failure": True
                },
                {
                    "name": "悬停卡片",
                    "enabled": True,
                    "condition": lambda: True,
                    "method": lambda: self.fetch_user_fans_via_hover(
                        page,
                        author_id=details.get("作者ID", ""),
                        author_name=details.get("作者昵称", ""),
                    ),
                    "priority": 2,
                    "retry_on_failure": False
                },
                {
                    "name": "直接提取",
                    "enabled": True,
                    "condition": lambda: True,
                    "method": lambda: self.fetch_user_fans_directly(page, url),
                    "priority": 3,
                    "retry_on_failure": False
                },
                {
                    "name": "主页获取",
                    "enabled": self.enable_homepage_fans,
                    "condition": lambda: details.get("博主主页"),
                    "method": lambda: self.fetch_user_fans_via_homepage(page.context, details.get("博主主页")),
                    "priority": 4,
                    "retry_on_failure": False
                },
                {
                    "name": "页面文本搜索",
                    "enabled": True,
                    "condition": lambda: True,  # 最后的兜底方法
                    "method": lambda: self.fans_from_page_text(page),
                    "priority": 5,
                    "retry_on_failure": False
                }
            ]

            # 按优先级排序并执行粉丝数提取
            successful_method = None
            for method_info in sorted(fans_extraction_methods, key=lambda x: x["priority"]):
                if not method_info["enabled"]:
                    continue

                if not method_info["condition"]():
                    continue

                try:
                    method_name = method_info["name"]
                    self.log(f"  [粉丝数] 尝试{method_name}...", 'info')

                    fans_result = method_info["method"]()

                    if fans_result and fans_result != "0" and len(fans_result.strip()) > 0:
                        # 验证粉丝数合理性
                        if self.validate_fans_number(fans_result):
                            details["粉丝数"] = fans_result
                            successful_method = method_name
                            self.log(f"  [粉丝数] ✓ {method_name}成功: {fans_result}", 'success')
                            break
                        else:
                            self.log(f"  [粉丝数] {method_name}结果不合理: {fans_result}", 'warning')

                except Exception as e:
                    error_msg = f"  [粉丝数] {method_info['name']}失败: {str(e)}"
                    if method_info["retry_on_failure"]:
                        self.log(error_msg, 'warning')
                        # 对于可重试的方法，记录更详细的错误信息
                        if "API" in method_info["name"]:
                            self.log(f"  [粉丝数] API错误详情: {str(e)}", 'error')
                    else:
                        self.log(error_msg, 'warning')
                    continue

            if not details["粉丝数"]:
                self.log("  [粉丝数] 所有方法均失败", 'error')
            else:
                self.log(f"  [粉丝数] 最终结果: {details['粉丝数']} (来源: {successful_method})", 'success')

        except Exception as e:
            self.log(f"详细内容提取失败: {str(e)}", 'error')

        return details, image_urls

    def ensure_media_loaded(self, page):
        """滚动页面触发懒加载图片"""
        try:
            page.evaluate("() => window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(800)
            page.evaluate("() => window.scrollTo(0, document.body.scrollHeight / 2)")
            page.wait_for_timeout(400)
            page.evaluate("() => window.scrollTo(0, 0)")
            page.wait_for_timeout(200)
        except Exception:
            pass

    def extract_note_image_urls_from_dom(self, page, limit=10):
        """仅提取当前笔记的图片URL（避免把推荐流/其他模块图片当封面）"""
        try:
            raw_urls = page.evaluate(
                r"""(maxCount) => {
                    const uniq = (arr) => {
                        const seen = new Set();
                        const out = [];
                        for (const u of arr) {
                            if (!u || seen.has(u)) continue;
                            seen.add(u);
                            out.push(u);
                        }
                        return out;
                    };

                    const pickAttr = (el, ...names) => {
                        if (!el) return '';
                        for (const name of names) {
                            const v = el.getAttribute(name);
                            if (v) return v;
                        }
                        return '';
                    };
                    const getImgSrc = (img) => {
                        if (!img) return '';
                        const direct = img.currentSrc || img.src || '';
                        return direct || pickAttr(img, 'src', 'data-src', 'data-original', 'data-origin', 'data-lazy-src');
                    };
                    const isVisible = (el) => !!(el && el.offsetParent !== null);

                    const pickMainCarousel = () => {
                        const candidates = Array.from(
                            document.querySelectorAll(
                                '.swiper, [class*="swiper"], [class*="Swiper"], .carousel, [class*="carousel"], [class*="Carousel"]'
                            )
                        );
                        const scored = candidates
                            .map((el) => {
                                const rect = el.getBoundingClientRect();
                                const imgs = el.querySelectorAll('img').length;
                                const slides = el.querySelectorAll('.swiper-slide').length;
                                const bullets = el.querySelectorAll('.swiper-pagination-bullet').length;
                                const area = Math.max(0, rect.width) * Math.max(0, rect.height);
                                const nearTop = rect.top > -200 && rect.top < 1400;
                                const visible = isVisible(el) && rect.width > 260 && rect.height > 180 && imgs > 0 && nearTop;
                                const score = (visible ? 1 : 0) * 1e9 + bullets * 1e7 + slides * 1e6 + imgs * 1e5 + area;
                                return { el, score, bullets, slides };
                            })
                            .sort((a, b) => b.score - a.score);
                        return scored.length ? scored[0] : null;
                    };

                    const carousel = pickMainCarousel();
                    let expected = 0;
                    let urls = [];

                    if (carousel && carousel.el) {
                        expected = carousel.bullets || 0;
                        if (!expected && carousel.slides && carousel.slides < 30) expected = carousel.slides;

                        const imgs = Array.from(carousel.el.querySelectorAll('img'));
                        urls = uniq(imgs.map(getImgSrc).filter(Boolean));

                        // 兼容部分页面用 background-image 渲染内容图：只在主轮播容器内取，避免推荐流污染
                        const bgEls = Array.from(carousel.el.querySelectorAll('[style*="background-image"]'));
                        for (const el of bgEls) {
                            try {
                                const bg = getComputedStyle(el).backgroundImage || '';
                                const m = bg.match(/url\\([\"\\']?(.*?)[\"\\']?\\)/);
                                if (m && m[1]) urls.push(m[1]);
                            } catch (e) {}
                        }
                        urls = uniq(urls.filter(Boolean));
                        if (expected && urls.length > expected) urls = urls.slice(0, expected);
                    }

                    if (!urls.length) {
                        // 兜底：只取页面顶部可见区域的内容图（避免推荐流）
                        const imgs = Array.from(
                            document.querySelectorAll(
                                'img[src*="sns-img"], img[src*="xhscdn"], img[src*="xiaohongshu.com"], img[data-src*="sns-img"], img[data-src*="xhscdn"], img[data-src*="xiaohongshu.com"]'
                            )
                        )
                            .map((img) => ({ img, src: getImgSrc(img), rect: img.getBoundingClientRect() }))
                            .filter((x) => x.src && isVisible(x.img) && x.rect.top > -200 && x.rect.top < 1800 && x.rect.width > 160 && x.rect.height > 160)
                            .sort((a, b) => a.rect.top - b.rect.top);

                        urls = uniq(imgs.map((x) => x.src));
                    }

                    return urls.slice(0, maxCount);
                }""",
                limit,
            ) or []

            if not isinstance(raw_urls, list):
                return []

            out = []
            seen = set()
            for u in raw_urls:
                normalized = self.normalize_media_url(page.url, u)
                if not normalized or normalized in seen:
                    continue
                if self.is_unwanted_image_src(normalized):
                    continue
                seen.add(normalized)
                out.append(normalized)
            return out[:limit]
        except Exception:
            return []

    def capture_video_cover(self, page, output_dir, name_prefix):
        """视频笔记：截取播放器静止画面作为封面兜底"""
        covers = []
        try:
            self.log("    [视频] 检测到视频笔记，尝试截取播放器封面...", 'info')

            try:
                page.evaluate("""() => {
                    const v = document.querySelector('video');
                    if (v) { v.pause(); v.currentTime = 0; }
                }""")
                time.sleep(1.2)
            except Exception:
                pass

            container_selectors = [
                "video",
                ".xgplayer",
                '[class*="video-container"]',
                '[class*="player"]',
                ".video-wrapper",
            ]
            container = None
            for sel in container_selectors:
                container = page.query_selector(sel)
                if container:
                    break

            if not container:
                self.log("    [视频] 未找到视频容器", 'warning')
                return covers

            cover_name = f"{name_prefix}_cover_video.png"
            cover_path = os.path.join(output_dir, cover_name)
            try:
                container.screenshot(path=cover_path, timeout=10000)
                if self.is_large_image_file(cover_path, min_width=100, min_height=100):
                    covers.append(cover_path)
                    self.log("    [视频] 封面截图成功", 'success')
                else:
                    try:
                        os.remove(cover_path)
                    except Exception:
                        pass
            except Exception as e:
                self.log(f"    [视频] 截图失败: {str(e)}", 'warning')
        except Exception as e:
            self.log(f"    [视频] 整体捕获失败: {str(e)}", 'warning')
        return covers

    def step_through_carousel(self, page, output_dir, name_prefix, cover_paths, seen_urls, limit=10):
        """轮播/Swiper逐帧抓取更多封面"""
        try:
            next_selectors = [
                ".swiper-button-next",
                '[class*="swiper"] [class*="next"]',
                'button[aria-label*="next"]',
                '[class*="arrow"][class*="next"]',
            ]
            bullet_selector = '.swiper-pagination-bullet, [class*="swiper"] [class*="bullet"]'

            def get_active_img_src():
                try:
                    src = page.evaluate("""() => {
                        const active = document.querySelector('.swiper-slide-active');
                        if (active) {
                            const img = active.querySelector('img');
                            if (img) return img.getAttribute('src') || img.getAttribute('data-src') || '';
                        }
                        const container = document.querySelector('[class*="swiper"], .carousel');
                        if (container) {
                            const img = container.querySelector('img');
                            if (img) return img.getAttribute('src') || img.getAttribute('data-src') || '';
                        }
                        return '';
                    }""")
                    return src or ""
                except Exception:
                    return ""

            bullets = page.query_selector_all(bullet_selector)
            if bullets and len(bullets) > 1:
                for b in bullets[:limit]:
                    try:
                        b.click()
                        page.wait_for_timeout(700)
                    except Exception:
                        continue
                    src = get_active_img_src()
                    if not src:
                        continue
                    normalized = self.normalize_media_url(page.url, src)
                    if not normalized or normalized in seen_urls or self.is_unwanted_image_src(normalized):
                        continue
                    cover_name = f"{name_prefix}_cover_{len(cover_paths)+1:02d}.png"
                    cover_path = os.path.join(output_dir, cover_name)
                    if self.download_image_via_page(page, normalized, cover_path):
                        if self.is_large_image_file(cover_path):
                            cover_paths.append(cover_path)
                        else:
                            try:
                                os.remove(cover_path)
                            except Exception:
                                pass
                    seen_urls.add(normalized)
                    if len(cover_paths) >= 10:
                        return cover_paths

            for _ in range(limit):
                src = get_active_img_src()
                if src:
                    normalized = self.normalize_media_url(page.url, src)
                    if normalized and normalized not in seen_urls and not self.is_unwanted_image_src(normalized):
                        cover_name = f"{name_prefix}_cover_{len(cover_paths)+1:02d}.png"
                        cover_path = os.path.join(output_dir, cover_name)
                        if self.download_image_via_page(page, normalized, cover_path):
                            if self.is_large_image_file(cover_path):
                                cover_paths.append(cover_path)
                            else:
                                try:
                                    os.remove(cover_path)
                                except Exception:
                                    pass
                        seen_urls.add(normalized)
                        if len(cover_paths) >= 10:
                            break

                clicked = False
                for sel in next_selectors:
                    btn = page.query_selector(sel)
                    if btn:
                        try:
                            btn.click()
                            clicked = True
                            break
                        except Exception:
                            continue
                if not clicked:
                    try:
                        page.keyboard.press("ArrowRight")
                        clicked = True
                    except Exception:
                        pass
                if not clicked:
                    break
                page.wait_for_timeout(600)
        except Exception:
            pass
        return cover_paths

    def capture_cover_screenshots_from_dom(self, page, output_dir, name_prefix, max_covers=10):
        """兜底：不依赖图片URL，直接对页面中的笔记封面区域截图（适用于 blob/data-src/403 场景）"""
        cover_paths = []
        try:
            self.ensure_media_loaded(page)

            # 1) 尝试定位“主轮播容器”（只选页面顶部的大容器，避免推荐流）
            candidates = page.query_selector_all(
                '.swiper, [class*="swiper"], [class*="Swiper"], .carousel, [class*="carousel"], [class*="Carousel"]'
            ) or []

            best = None
            best_score = -1
            for el in candidates:
                try:
                    box = el.bounding_box()
                    if not box:
                        continue
                    if box.get("y", 10**9) < -200 or box.get("y", 10**9) > 1400:
                        continue
                    if box.get("width", 0) < 260 or box.get("height", 0) < 180:
                        continue
                    img_count = 0
                    try:
                        img_count = len(el.query_selector_all("img") or [])
                    except Exception:
                        img_count = 0
                    if img_count < 1:
                        continue
                    score = img_count * 1_000_000 + (box.get("width", 0) * box.get("height", 0))
                    if score > best_score:
                        best = el
                        best_score = score
                except Exception:
                    continue

            def screenshot_img_element(img_el):
                if not img_el:
                    return None
                cover_name = f"{name_prefix}_cover_{len(cover_paths)+1:02d}.png"
                cover_path = os.path.join(output_dir, cover_name)
                try:
                    img_el.screenshot(path=cover_path, timeout=15000)
                    if self.is_large_image_file(cover_path, min_width=100, min_height=100):
                        cover_paths.append(cover_path)
                        return cover_path
                    try:
                        os.remove(cover_path)
                    except Exception:
                        pass
                except Exception:
                    try:
                        if os.path.exists(cover_path):
                            os.remove(cover_path)
                    except Exception:
                        pass
                return None

            # 2) 有轮播：优先点 bullet 逐张截图
            if best:
                bullets = []
                try:
                    bullets = best.query_selector_all('.swiper-pagination-bullet, [class*="bullet"]') or []
                except Exception:
                    bullets = []

                def get_active_img():
                    try:
                        return best.query_selector(".swiper-slide-active img") or best.query_selector("img")
                    except Exception:
                        return None

                seen_src = set()
                if bullets and len(bullets) > 1:
                    for b in bullets[:max_covers]:
                        try:
                            b.click()
                            page.wait_for_timeout(600)
                        except Exception:
                            pass
                        img_el = get_active_img()
                        try:
                            src = (img_el.get_attribute("src") or img_el.get_attribute("data-src") or "") if img_el else ""
                            if src and src in seen_src:
                                continue
                            if src:
                                seen_src.add(src)
                        except Exception:
                            pass
                        screenshot_img_element(img_el)
                        if len(cover_paths) >= max_covers:
                            break
                else:
                    # 单图/未识别 bullet：先截一张，再尝试 next
                    img_el = get_active_img()
                    screenshot_img_element(img_el)

                    next_btn = None
                    try:
                        next_btn = best.query_selector(
                            ".swiper-button-next, [class*='next'], button[aria-label*='next'], [class*='arrow'][class*='next']"
                        )
                    except Exception:
                        next_btn = None

                    for _ in range(max_covers - len(cover_paths)):
                        if not next_btn:
                            break
                        try:
                            next_btn.click()
                            page.wait_for_timeout(700)
                        except Exception:
                            break
                        img_el = get_active_img()
                        try:
                            src = (img_el.get_attribute("src") or img_el.get_attribute("data-src") or "") if img_el else ""
                            if src and src in seen_src:
                                continue
                            if src:
                                seen_src.add(src)
                        except Exception:
                            pass
                        screenshot_img_element(img_el)
                        if len(cover_paths) >= max_covers:
                            break

                return cover_paths

            # 3) 无轮播：兜底截取页面顶部“最大可见图片”
            imgs = page.query_selector_all("img") or []
            best_img = None
            best_area = -1
            for img in imgs:
                try:
                    box = img.bounding_box()
                    if not box:
                        continue
                    if box.get("y", 10**9) < -200 or box.get("y", 10**9) > 1400:
                        continue
                    if box.get("width", 0) < 220 or box.get("height", 0) < 220:
                        continue
                    area = box.get("width", 0) * box.get("height", 0)
                    if area > best_area:
                        best_area = area
                        best_img = img
                except Exception:
                    continue

            if best_img:
                screenshot_img_element(best_img)
        except Exception:
            pass
        return cover_paths

    def capture_cover_images_ultimate(self, page, output_dir, name_prefix, image_urls=None):
        """封面抓取：元数据优先 + DOM/轮播/视频兜底（最多10张）"""
        cover_paths = []
        image_urls = image_urls or []
        max_covers = 10

        try:
            self.log(f"开始封面抓取 (元数据URL数: {len(image_urls)})", 'info')

            seen_urls = set()

            # 策略1：元数据URL直接下载（最快）
            if image_urls:
                for url in image_urls[:max_covers]:
                    if not url:
                        continue
                    normalized = self.normalize_media_url(page.url, url)
                    if not normalized or normalized in seen_urls or self.is_unwanted_image_src(normalized):
                        continue
                    seen_urls.add(normalized)

                    cover_name = f"{name_prefix}_cover_{len(cover_paths)+1:02d}.png"
                    cover_path = os.path.join(output_dir, cover_name)
                    if self.download_image_via_page(page, normalized, cover_path):
                        if self.is_large_image_file(cover_path, min_width=100, min_height=100):
                            cover_paths.append(cover_path)
                            self.log(f"封面保存成功: {os.path.basename(cover_path)}", 'success')
                        else:
                            try:
                                os.remove(cover_path)
                            except Exception:
                                pass
                    if len(cover_paths) >= max_covers:
                        return cover_paths

                # 已成功下载到封面时，不再“补图”，避免把推荐流/其他模块图片当封面
                if cover_paths:
                    return cover_paths[:max_covers]

            # 视频笔记优先兜底
            is_video = False
            try:
                is_video = bool(page.evaluate("""() => document.querySelector('video') !== null"""))
            except Exception:
                is_video = False

            if is_video and len(cover_paths) < 1:
                video_covers = self.capture_video_cover(page, output_dir, name_prefix)
                for p in video_covers:
                    if p and os.path.exists(p):
                        cover_paths.append(p)
                if cover_paths:
                    return cover_paths[:max_covers]

            # 策略2：仅从“当前笔记”的轮播/内容区提取图片URL（严格避免推荐流污染）
            if len(cover_paths) < 1:
                self.log("尝试从页面提取当前笔记图片...", 'info')
                try:
                    self.ensure_media_loaded(page)
                except Exception:
                    pass
                try:
                    dom_urls = self.extract_note_image_urls_from_dom(page, limit=max_covers)
                except Exception:
                    dom_urls = []

                for url in dom_urls[:max_covers]:
                    if len(cover_paths) >= max_covers:
                        break
                    normalized = self.normalize_media_url(page.url, url)
                    if not normalized or normalized in seen_urls or self.is_unwanted_image_src(normalized):
                        continue
                    seen_urls.add(normalized)
                    cover_name = f"{name_prefix}_cover_{len(cover_paths)+1:02d}.png"
                    cover_path = os.path.join(output_dir, cover_name)
                    if self.download_image_via_page(page, normalized, cover_path):
                        if self.is_large_image_file(cover_path, min_width=100, min_height=100):
                            cover_paths.append(cover_path)
                        else:
                            try:
                                os.remove(cover_path)
                            except Exception:
                                pass

            # 策略3：URL下载失败时，直接截图封面区域兜底（适用于 blob/data-src/403）
            if len(cover_paths) < 1:
                try:
                    self.log("尝试截图方式抓取封面...", "info")
                    screenshot_covers = self.capture_cover_screenshots_from_dom(
                        page, output_dir, name_prefix, max_covers=max_covers
                    )
                    for p in screenshot_covers or []:
                        if p and os.path.exists(p):
                            cover_paths.append(p)
                except Exception:
                    pass

            if cover_paths:
                self.log(f"封面抓取完成: {len(cover_paths)} 张", 'success')
            else:
                self.log("封面抓取失败：未找到有效图片URL", 'warning')

        except Exception as e:
            self.log(f"封面抓取失败: {str(e)}", 'error')

        return cover_paths[:max_covers]

    def validate_fans_number(self, fans_str):
        """验证粉丝数的合理性"""
        try:
            if not fans_str or not isinstance(fans_str, str):
                return False

            # 解析为纯数字
            clean_number = self.parse_fans_number(fans_str)
            if not clean_number:
                return False

            # 转换为整数进行验证
            try:
                number = int(clean_number)
            except ValueError:
                return False

            # 合理性检查
            # 粉丝数应该在 1 到 10亿 之间
            if number < 1 or number > 1000000000:
                return False

            # 排除明显不合理的数字
            unreasonable_numbers = [123, 666, 888, 999, 111, 222, 333, 444, 555, 777]
            if number in unreasonable_numbers and number < 1000:
                return False

            return True

        except Exception as e:
            self.log(f"    粉丝数验证失败: {str(e)}", 'warning')
            return False

    def fans_from_page_text(self, page):
        """从页面文本中提取粉丝数（增强版）"""
        try:
            self.log("  [文本] 从页面全文搜索粉丝数...", 'info')

            # 使用JavaScript进行更智能的文本搜索
            fans_data = page.evaluate("""() => {
                const results = {
                    fans: '',
                    confidence: 0,
                    context: '',
                    method: ''
                };

                // 搜索范围：页面前10000字符（重点关注顶部区域）
                const bodyText = document.body.innerText || '';
                const searchText = bodyText.substring(0, 12000);

                // 多种匹配模式
                const patterns = [
                    // 标准模式：粉丝/关注者 + 数字
                    {
                        regex: /(?:粉丝|关注者|followers)\\s*[:：]?\\s*([0-9]+\\.?[0-9]*[万wWkK]?)/gi,
                        confidence: 80,
                        method: '标准匹配'
                    },
                    // 数字 + 粉丝/关注者
                    {
                        regex: /([0-9]+\\.?[0-9]*[万wWkK]?)\\s*(?:粉丝|关注者|followers)/gi,
                        confidence: 70,
                        method: '倒序匹配'
                    },
                    // 常见表达
                    {
                        regex: /粉丝数\\s*[:：]?\\s*([0-9]+\\.?[0-9]*[万wWkK]?)/gi,
                        confidence: 75,
                        method: '粉丝数匹配'
                    },
                    // 关注者表达
                    {
                        regex: /关注者\\s*[:：]?\\s*([0-9]+\\.?[0-9]*[万wWkK]?)/gi,
                        confidence: 75,
                        method: '关注者匹配'
                    }
                ];

                // 执行匹配
                for (const pattern of patterns) {
                    const matches = searchText.match(pattern.regex);
                    if (matches) {
                        for (const match of matches) {
                            // 提取数字部分
                            const numberMatch = match.match(/([0-9]+\\.?[0-9]*[万wWkK]?)/);
                            if (numberMatch && numberMatch[1]) {
                                const fans = numberMatch[1];

                                // 检查上下文，确保不是无关数据
                                const index = searchText.indexOf(match);
                                const contextStart = Math.max(0, index - 50);
                                const contextEnd = Math.min(searchText.length, index + match.length + 50);
                                const context = searchText.substring(contextStart, contextEnd);

                                // 过滤明显的无关内容
                                if (context.includes('关注') || context.includes('点赞') ||
                                    context.includes('收藏') || context.includes('分享') ||
                                    context.includes('用户') || context.includes('博主') ||
                                    !context.includes('评论')) {  // 排除评论区的数据

                                    results.fans = fans;
                                    results.confidence = pattern.confidence;
                                    results.context = context;
                                    results.method = pattern.method;
                                    return results;
                                }
                            }
                        }
                    }
                }

                // 如果没有找到，尝试更宽松的搜索
                if (!results.fans) {
                    const loosePatterns = [
                        /([0-9]+[万wWkK])\\s*粉/gi,
                        /粉\\s*([0-9]+[万wWkK])/gi,
                        /([0-9]+)\\s*粉/gi
                    ];

                    for (const pattern of loosePatterns) {
                        const matches = searchText.match(pattern);
                        if (matches && matches[0]) {
                            const numberMatch = matches[0].match(/([0-9]+[万wWkK]?)/);
                            if (numberMatch) {
                                results.fans = numberMatch[1];
                                results.confidence = 30;
                                results.method = '宽松匹配';
                                return results;
                            }
                        }
                    }
                }

                return results;
            }""")

            if fans_data and fans_data.get('fans'):
                fans_number = fans_data['fans']
                confidence = fans_data.get('confidence', 0)
                method = fans_data.get('method', '')
                context = fans_data.get('context', '')

                self.log(f"  [文本] 找到粉丝数: {fans_number} (置信度: {confidence}% | 方法: {method})", 'info')
                if context:
                    self.log(f"  [文本] 上下文: {context[:100]}...", 'info')

                # 格式化处理
                fans_number = self.parse_fans_number(fans_number)
                if fans_number and fans_number != "0":
                    return fans_number

            self.log("  [文本] 未找到有效的粉丝数", 'warning')
            return ""

        except Exception as e:
            self.log(f"  [文本] 页面文本搜索失败: {str(e)}", 'error')
            return ""

    def normalize_media_url(self, base_url, url):
        """增强版媒体URL标准化（处理复杂协议、路径、编码问题）"""
        try:
            if not url:
                self.log("    URL为空，跳过处理", 'warning')
                return ""

            # 预处理：清理空白字符和特殊字符
            url = str(url).strip()
            if not url:
                return ""

            # 1. 处理协议相关URL
            if url.startswith("//"):
                # 协议相对URL，使用HTTPS
                normalized_url = "https:" + url
            elif url.startswith("http://") or url.startswith("https://"):
                normalized_url = url
                # 将HTTP升级为HTTPS（更安全）
                if url.startswith("http://") and not url.startswith("http://localhost"):
                    normalized_url = "https://" + url[7:]
            elif url.startswith("data:"):
                # Base64编码的图片，跳过处理
                return ""
            elif url.startswith("blob:"):
                # Blob URL，跳过处理
                return ""
            else:
                # 相对URL，需要与base_url拼接
                if not base_url:
                    self.log("    缺少base_url，无法处理相对URL", 'warning')
                    return ""
                normalized_url = urljoin(base_url, url)

            # 2. URL解码和重编码处理
            try:
                from urllib.parse import unquote, quote, urlparse, urlunparse
                import re

                # 解码URL中的特殊字符（避免双重编码）
                decoded_url = unquote(normalized_url)

                # 解析URL组件
                parsed = urlparse(decoded_url)

                # 3. 重建URL，确保各组件正确
                # 域名部分保持原样（大小写敏感）
                netloc = parsed.netloc

                # 路径部分：规范化斜杠，但保留有意义的编码
                path = parsed.path
                if path:
                    # 移除多余的斜杠
                    path = re.sub(r'/+', '/', path)
                    # 确保路径以斜杠开头（绝对路径）
                    if not path.startswith('/'):
                        path = '/' + path

                # 查询参数：保持原样，但确保格式正确
                query = parsed.query

                # 重建URL
                if path or query:
                    rebuilt = urlunparse((
                        'https',  # 强制使用HTTPS
                        netloc,
                        path or '/',
                        '',  # params
                        query,
                        ''   # fragment
                    ))
                else:
                    rebuilt = f"https://{netloc}/"

                normalized_url = rebuilt

            except Exception as e:
                self.log(f"    URL重建失败，使用原始URL: {str(e)}", 'warning')
                # 如果重建失败，回到简单处理
                pass

            # 4. URL有效性检查
            try:
                parsed_final = urlparse(normalized_url)
                if not parsed_final.netloc:
                    # 如果没有域名，可能是 malformed URL
                    self.log(f"    URL格式无效，缺少域名: {normalized_url[:50]}...", 'warning')
                    return ""

                # 检查是否为允许的域名
                allowed_domains = [
                    'xiaohongshu.com',
                    'xhscdn.com',
                    'sns-img-bd.xhscdn.com',
                    'sns-img-qn.xhscdn.com',
                    'ci.xiaohongshu.com',
                    'edith.xiaohongshu.com'
                ]

                domain_match = any(domain in parsed_final.netloc.lower() for domain in allowed_domains)
                if not domain_match:
                    # 如果不是小红书域名，但仍可以是有效图片URL，记录警告但继续
                    self.log(f"    非小红书域名，但仍尝试处理: {parsed_final.netloc}", 'info')

                # 验证文件扩展名（如果有）
                path_lower = parsed_final.path.lower()
                image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.avif', '.bmp']
                has_image_extension = any(path_lower.endswith(ext) for ext in image_extensions)

                # 如果没有图片扩展名但URL看起来像图片，仍然尝试
                if not has_image_extension and any(keyword in path_lower for keyword in ['image', 'img', 'photo', 'pic']):
                    self.log(f"    URL可能为图片但缺少扩展名，仍尝试处理", 'info')

            except Exception as e:
                self.log(f"    URL有效性检查失败: {str(e)}", 'warning')
                return ""

            # 5. 最终清理
            # 移除URL片段（#后面的部分）
            if '#' in normalized_url:
                normalized_url = normalized_url.split('#')[0]

            # 确保URL不为空且长度合理
            if len(normalized_url) > 2048:  # URL长度限制
                self.log(f"    URL过长，可能是 malformed: {len(normalized_url)} 字符", 'warning')
                return ""

            if not normalized_url:
                self.log("    处理后URL为空", 'warning')
                return ""

            return normalized_url

        except Exception as e:
            self.log(f"    URL标准化过程异常: {str(e)}", 'error')
            # 发生异常时，返回原始URL进行兜底
            try:
                # 至少进行基本的协议检查
                url_str = str(url).strip()
                if url_str and not url_str.startswith(('http://', 'https://', 'data:', 'blob:')):
                    base_str = str(base_url).strip()
                    if base_str:
                        return urljoin(base_str, url_str)
                return url_str
            except Exception:
                return ""

    def is_unwanted_image_src(self, src):
        """排除不需要的图片源"""
        if not src:
            return True
        unwanted_keywords = [
            "avatar", "profile", "logo", "icon", "watermark", "badge",
            "placeholder", "default_avatar", "favicon", "xhslogo", "head"
        ]
        return any(keyword in src.lower() for keyword in unwanted_keywords)

    def download_image_via_page(self, page, image_url, save_path):
        """通过页面下载图片（增强版：重试机制、流式处理、多格式支持）"""
        import time
        import math

        # 更全面的请求头配置
        headers_list = [
            {
                "Referer": page.url,
                "User-Agent": self.http_user_agent,
                "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
                "Accept-Encoding": "gzip, deflate, br",
                "Origin": "https://www.xiaohongshu.com",
                "Sec-Fetch-Dest": "image",
                "Sec-Fetch-Mode": "no-cors",
                "Sec-Fetch-Site": "cross-site",
                "Cache-Control": "no-cache",
                "Pragma": "no-cache"
            },
            {
                "Referer": "https://www.xiaohongshu.com/",
                "User-Agent": self.http_user_agent,
                "Accept": "image/jpeg,image/png,image/gif,image/webp,image/apng,*/*;q=0.9",
                "Accept-Language": "zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2",
                "DNT": "1",
                "Connection": "keep-alive",
                "Upgrade-Insecure-Requests": "1"
            }
        ]

        def finalize_image(path, original_format_hint=""):
            """智能图片格式处理"""
            try:
                with PILImage.open(path) as img:
                    img.load()

                    # 获取原始图片信息
                    original_format = img.format or original_format_hint
                    file_size = os.path.getsize(path)

                    # 决定最佳保存格式
                    if img.mode in ("RGBA", "LA", "P") and img.info.get('transparency') is not None:
                        # 有透明通道，优先保存为PNG
                        final_format = "PNG"
                        converted = img.convert("RGBA")
                    elif img.mode in ("L", "P"):
                        # 灰度或调色板图像，检查是否需要转换为RGB
                        if len(img.getcolors() or []) > 256:
                            final_format = "JPEG"
                            converted = img.convert("RGB")
                        else:
                            final_format = "PNG"
                            converted = img.convert("RGBA") if img.mode == "P" else img
                    else:
                        # RGB图像，根据质量选择格式
                        if file_size > 2 * 1024 * 1024:  # 大于2MB考虑压缩
                            final_format = "JPEG"
                            converted = img.convert("RGB")
                        else:
                            final_format = original_format or "PNG"
                            converted = img.copy()

                    # 保存图片
                    converted.save(path, format=final_format, quality=95, optimize=True)
                    self.log(f"    图片格式优化: {original_format}->{final_format}, 大小: {file_size/1024:.1f}KB", 'info')
                    return True

            except Exception as e:
                self.log(f"    图片格式处理失败: {str(e)}", 'warning')
                # 尝试原始文件保留
                if os.path.exists(path) and os.path.getsize(path) > 0:
                    self.log(f"    保留原始文件", 'info')
                    return True
                try:
                    os.remove(path)
                except Exception:
                    pass
                return False

        def download_with_streaming_response(response, path):
            """流式下载处理大文件"""
            try:
                with open(path, 'wb') as f:
                    # 分块写入，避免内存溢出
                    chunk_size = 8192
                    total_size = 0
                    while True:
                        chunk = response.read_chunk(chunk_size)
                        if not chunk:
                            break
                        f.write(chunk)
                        total_size += len(chunk)
                        # 防止下载过大的文件
                        if total_size > 10 * 1024 * 1024:  # 10MB限制
                            self.log(f"    文件过大，终止下载: {total_size/1024/1024:.1f}MB", 'warning')
                            return False
                return True
            except Exception as e:
                self.log(f"    流式下载失败: {str(e)}", 'error')
                return False

        # 重试下载逻辑
        max_retries = 3
        for attempt in range(max_retries):
            try:
                self.log(f"    尝试下载图片 ({attempt + 1}/{max_retries}): {image_url[:80]}...", 'info')

                # 指数退避延迟
                if attempt > 0:
                    delay = min(math.pow(2, attempt), 10)  # 最大10秒
                    time.sleep(delay)

                # 尝试不同的请求头
                for headers_idx, headers in enumerate(headers_list):
                    try:
                        self.log(f"      使用请求头配置 {headers_idx + 1}", 'info')

                        # 增加超时时间并使用流式下载
                        response = page.context.request.get(
                            image_url,
                            headers=headers,
                            timeout=30000,
                            # 启用流式响应（如果Playwright支持）
                        )

                        if response.ok:
                            # 获取文件扩展名用于格式提示
                            content_type = response.headers.get('content-type', '')
                            format_hint = 'JPEG' if 'jpeg' in content_type else 'PNG' if 'png' in content_type else ''

                            # 检查内容长度
                            content_length = response.headers.get('content-length')
                            if content_length and int(content_length) > 10 * 1024 * 1024:
                                self.log(f"      文件过大: {int(content_length)/1024/1024:.1f}MB", 'warning')
                                continue

                            # 保存文件
                            if hasattr(response, 'body') and callable(getattr(response, 'body')):
                                # Playwright响应对象
                                try:
                                    with open(save_path, 'wb') as f:
                                        f.write(response.body())
                                except Exception as be:
                                    self.log(f"      直接写入失败: {str(be)}", 'warning')
                                    continue
                            else:
                                # 尝试流式下载
                                if not download_with_streaming_response(response, save_path):
                                    continue

                            # 验证下载的文件
                            if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                                if finalize_image(save_path, format_hint):
                                    final_size = os.path.getsize(save_path)
                                    self.log(f"    ✓ 下载成功: {final_size/1024:.1f}KB", 'success')
                                    return True
                            else:
                                self.log(f"      文件保存失败", 'warning')
                                continue

                        else:
                            self.log(f"      HTTP {response.status}: {response.status_text}", 'warning')

                    except Exception as e:
                        self.log(f"      请求头配置 {headers_idx + 1} 失败: {str(e)}", 'warning')
                        continue

                # 第一轮HTTP请求失败，尝试JavaScript fetch
                if attempt == max_retries - 1:
                    self.log(f"    最后尝试JavaScript fetch", 'info')
                    try:
                        js_data = page.evaluate(
                            """async (url) => {
                                try {
                                    const response = await fetch(url, {
                                        credentials: 'include',
                                        mode: 'cors',
                                        cache: 'no-store'
                                    });
                                    if (!response.ok) return null;
                                    const buffer = await response.arrayBuffer();
                                    return {
                                        data: Array.from(new Uint8Array(buffer)),
                                        contentType: response.headers.get('content-type') || ''
                                    };
                                } catch (err) {
                                    console.log('Fetch error:', err);
                                    return null;
                                }
                            }""",
                            image_url,
                        )

                        if js_data and js_data.get('data'):
                            with open(save_path, 'wb') as f:
                                f.write(bytes(js_data['data']))

                            content_type = js_data.get('contentType', '')
                            format_hint = 'JPEG' if 'jpeg' in content_type else 'PNG' if 'png' in content_type else ''

                            if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                                if finalize_image(save_path, format_hint):
                                    final_size = os.path.getsize(save_path)
                                    self.log(f"    ✓ JavaScript下载成功: {final_size/1024:.1f}KB", 'success')
                                    return True

                    except Exception as je:
                        self.log(f"    JavaScript下载失败: {str(je)}", 'error')

            except Exception as e:
                self.log(f"    下载尝试 {attempt + 1} 失败: {str(e)}", 'error')
                continue

        self.log(f"    ✗ 所有下载尝试均失败", 'error')
        # 清理可能的部分下载文件
        try:
            if os.path.exists(save_path):
                os.remove(save_path)
        except Exception:
            pass
        return False

    def is_large_image_file(self, image_path, min_width=160, min_height=160):
        """检查图片文件尺寸"""
        try:
            with PILImage.open(image_path) as im:
                w, h = im.size
                return w >= min_width and h >= min_height
        except Exception:
            return True

    def sanitize_name(self, name):
        """清理文件名"""
        try:
            safe = re.sub(r'[^0-9A-Za-z\u4e00-\u9fff]+', '_', name)
            return safe.strip('_') or "item"
        except Exception:
            return "item"

    def extract_note_id(self, url):
        """提取笔记ID"""
        try:
            match = re.search(r'/explore/([a-zA-Z0-9]+)', url)
            if match:
                return match.group(1)
        except:
            pass
        return ""

    def resize_image(self, image_path, max_width=200, max_height=150):
        """按最大宽高等比例缩放图片，返回(width, height)"""
        try:
            with PILImage.open(image_path) as img:
                width_ratio = max_width / img.width
                height_ratio = max_height / img.height
                ratio = min(width_ratio, height_ratio)
                return int(img.width * ratio), int(img.height * ratio)
        except Exception:
            return max_width, max_height

    def create_ultimate_excel_report(self, results, output_dir):
        """v4.0终极版Excel报告（兼容Pro v2字段 + v4状态）"""
        os.makedirs(output_dir, exist_ok=True)

        def argb(color, default="FFFFFFFF"):
            if not color:
                return default
            value = str(color).strip()
            if value.startswith("#"):
                value = value[1:]
            value = value.upper()
            if len(value) == 6:
                return "FF" + value
            if len(value) == 8:
                return value
            return default

        primary = argb(self.colors.get("primary", "4169E1"))
        success = argb(self.colors.get("success", "28A745"))
        danger = argb(self.colors.get("danger", "DC3545"))
        warning = argb(self.colors.get("warning", "FF9500"))
        white = argb("FFFFFF")
        border_color = argb("CCCCCC")
        link_blue = argb("0066CC")
        fans_color = argb("FF6347")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "小红书监测v3.0"

        # 列宽设置（Pro v2字段 + v3新增状态）
        column_widths = {
            "A": 6,    # 序号
            "B": 12,   # 产品
            "C": 50,   # 链接
            "D": 12,   # 采集状态
            "E": 30,   # 错误信息
            "F": 18,   # 笔记ID
            "G": 35,   # 标题
            "H": 15,   # 作者昵称
            "I": 18,   # 作者ID
            "J": 12,   # 粉丝数
            "K": 40,   # 博主主页
            "L": 16,   # 发布时间
            "M": 10,   # 点赞数
            "N": 10,   # 收藏数
            "O": 10,   # 评论数
            "P": 10,   # 分享数
            "Q": 40,   # 封面链接
            "R": 40,   # 视频链接
            "S": 50,   # 正文
            "T": 35,   # 截图
            "U": 22,   # 封面1
            "V": 22,   # 封面2
            "W": 22,   # 封面3
            "X": 22,   # 封面4
            "Y": 22,   # 封面5
            "Z": 22,   # 封面6
            "AA": 22,  # 封面7
            "AB": 22,  # 封面8
            "AC": 22,  # 封面9
            "AD": 22,  # 封面10
            "AE": 15,  # 登录状态
            "AF": 12,  # 反爬状态
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        headers = [
            "序号", "产品", "链接", "采集状态", "错误信息", "笔记ID", "标题", "作者昵称", "作者ID",
            "粉丝数", "博主主页", "发布时间", "点赞数", "收藏数", "评论数", "分享数",
            "封面链接", "视频链接", "正文", "截图",
            "封面1", "封面2", "封面3", "封面4", "封面5",
            "封面6", "封面7", "封面8", "封面9", "封面10",
            "登录状态", "反爬状态",
        ]

        header_fill = PatternFill(start_color=primary, end_color=primary, fill_type="solid")
        header_font = Font(color=white, bold=True, size=10)
        thin_border = Border(
            left=Side(style="thin", color=border_color),
            right=Side(style="thin", color=border_color),
            top=Side(style="thin", color=border_color),
            bottom=Side(style="thin", color=border_color),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 30

        anchor_cols = ["U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD"]

        current_row = 2
        for result in results:
            row_height = 100

            # 1. 序号
            cell = ws.cell(row=current_row, column=1, value=result.get("序号", ""))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            # 2. 产品
            cell = ws.cell(row=current_row, column=2, value=result.get("产品", ""))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            # 3. 链接
            cell = ws.cell(row=current_row, column=3, value=result.get("链接", ""))
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border

            # 4. 采集状态
            status = result.get("采集状态", "失败")
            cell = ws.cell(row=current_row, column=4, value=status)
            cell.fill = PatternFill(start_color=success if status == "成功" else danger,
                                    end_color=success if status == "成功" else danger,
                                    fill_type="solid")
            cell.font = Font(color=white, bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            # 5. 错误信息
            cell = ws.cell(row=current_row, column=5, value=result.get("错误信息", ""))
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(color=danger, size=9)
            cell.border = thin_border

            # 6. 笔记ID
            cell = ws.cell(row=current_row, column=6, value=result.get("笔记ID", ""))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Consolas", size=9)
            cell.border = thin_border

            # 7. 标题
            cell = ws.cell(row=current_row, column=7, value=result.get("标题", ""))
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, size=9)
            cell.border = thin_border

            # 8. 作者昵称
            cell = ws.cell(row=current_row, column=8, value=result.get("作者昵称", ""))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=9)
            cell.border = thin_border

            # 9. 作者ID
            cell = ws.cell(row=current_row, column=9, value=result.get("作者ID", ""))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Consolas", size=9)
            cell.border = thin_border

            # 10. 粉丝数
            cell = ws.cell(row=current_row, column=10, value=result.get("粉丝数", ""))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, color=fans_color, size=9)
            cell.border = thin_border

            # 11. 博主主页
            cell = ws.cell(row=current_row, column=11, value=result.get("博主主页", ""))
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(size=8, color=link_blue, underline="single")
            cell.border = thin_border

            # 12. 发布时间
            cell = ws.cell(row=current_row, column=12, value=result.get("发布时间", ""))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=9)
            cell.border = thin_border

            # 13-16. 互动数据
            for col, key in [(13, "点赞数"), (14, "收藏数"), (15, "评论数"), (16, "分享数")]:
                cell = ws.cell(row=current_row, column=col, value=result.get(key, ""))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=9)
                cell.border = thin_border

            # 17. 封面链接
            cell = ws.cell(row=current_row, column=17, value=result.get("封面链接", ""))
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(size=8, color=link_blue)
            cell.border = thin_border

            # 18. 视频链接
            cell = ws.cell(row=current_row, column=18, value=result.get("视频链接", ""))
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(size=8, color=link_blue)
            cell.border = thin_border

            # 19. 正文
            cell = ws.cell(row=current_row, column=19, value=result.get("正文", ""))
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.font = Font(size=8)
            cell.border = thin_border

            # 20. 截图
            screenshot_path = result.get("截屏文件", "")
            cell = ws.cell(row=current_row, column=20)
            cell.border = thin_border
            if screenshot_path and os.path.exists(screenshot_path):
                try:
                    width, height = self.resize_image(screenshot_path, max_width=250, max_height=120)
                    img = XLImage(screenshot_path)
                    img.width = width
                    img.height = height
                    img.anchor = f"T{current_row}"
                    ws.add_image(img)
                    row_height = max(row_height, height * 0.75 + 10)
                except Exception as e:
                    cell.value = f"加载失败: {str(e)}"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = "无截图"
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 21-30. 多封面展示（最多10张）
            cover_paths = result.get("封面图列表", []) or []
            for i in range(10):
                col_index = 21 + i
                cell = ws.cell(row=current_row, column=col_index)
                cell.border = thin_border
                if i < len(cover_paths) and os.path.exists(cover_paths[i]):
                    try:
                        width, height = self.resize_image(cover_paths[i], max_width=140, max_height=110)
                        img = XLImage(cover_paths[i])
                        img.width = width
                        img.height = height
                        img.anchor = f"{anchor_cols[i]}{current_row}"
                        ws.add_image(img)
                        row_height = max(row_height, height * 0.75 + 10)
                    except Exception:
                        cell.value = "封面失败"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.value = ""

            # 31. 登录状态
            login_status = result.get("登录状态", "未知")
            cell = ws.cell(row=current_row, column=31, value=login_status)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if login_status == "成功":
                cell.fill = PatternFill(start_color=success, end_color=success, fill_type="solid")
                cell.font = Font(color=white, bold=True, size=10)

            # 32. 反爬状态
            anti_status = result.get("反爬状态", "未知")
            cell = ws.cell(row=current_row, column=32, value=anti_status)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if anti_status == "触发反爬":
                cell.fill = PatternFill(start_color=warning, end_color=warning, fill_type="solid")
                cell.font = Font(color=white, bold=True, size=10)
            elif anti_status == "正常":
                cell.fill = PatternFill(start_color=success, end_color=success, fill_type="solid")
                cell.font = Font(color=white, bold=True, size=10)

            ws.row_dimensions[current_row].height = row_height
            current_row += 1

        # 统计行
        stats_row = current_row + 1
        success_count = sum(1 for r in results if r.get("采集状态") == "成功")
        fail_count = len(results) - success_count
        anti_crawler_count = sum(1 for r in results if r.get("反爬状态") == "触发反爬")

        ws.cell(row=stats_row, column=1, value="统计").font = Font(bold=True, size=11, color=primary)
        ws.cell(
            row=stats_row,
            column=2,
            value=f"总计: {len(results)} 条 | 成功: {success_count} 条 | 失败: {fail_count} 条 | 反爬触发: {anti_crawler_count} 条",
        ).font = Font(bold=True, size=10)

        ws.freeze_panes = "A2"

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_dir, f"小红书监测v3.0_Ultimate_{timestamp}.xlsx")
        wb.save(output_file)

        return output_file, success_count, fail_count, anti_crawler_count

    def monitor_task(self):
        """v3.0主监控任务"""
        try:
            self.is_running = True
            self.stop_flag = False
            self.start_button.config(state='disabled')
            self.stop_button.config(state='normal')

            # 重置统计
            self.stats = {
                'total_processed': 0,
                'success_count': 0,
                'failed_count': 0,
                'anti_crawler_count': 0,
                'start_time': datetime.now()
            }
            self.update_stats()

            csv_file = self.csv_file.get()
            if not os.path.exists(csv_file):
                messagebox.showerror("错误", f"CSV文件不存在: {csv_file}")
                return

            output_dir = self.output_dir.get()
            screenshot_dir = os.path.join(output_dir, 'screenshots')
            os.makedirs(screenshot_dir, exist_ok=True)

            self.log("="*60)
            self.log("v3.0 Ultimate 监测任务开始", 'info')
            self.log(f"输入: {csv_file}", 'info')
            self.log(f"输出: {output_dir}", 'info')
            self.log(f"模式: {'完整模式' if self.capture_mode.get() == 'full' else '快速模式'}", 'info')
            self.log(f"反爬检测: {'启用' if self.enable_anti_detection.get() else '禁用'}", 'info')
            self.log("="*60)

            items = self.read_links_from_csv(csv_file)
            self.log(f"读取到 {len(items)} 个链接", 'info')

            if not items:
                messagebox.showwarning("警告", "CSV中没有链接")
                return

            # 执行抓取
            results = self.capture_with_ultimate_features(items, screenshot_dir)

            if not self.stop_flag:
                self.log("生成v3.0 Excel报告...", 'info')
                output_file, success, fail, anti_crawler = self.create_ultimate_excel_report(results, output_dir)

                # 计算总耗时
                total_time = datetime.now() - self.stats['start_time']
                minutes, seconds = divmod(total_time.total_seconds(), 60)

                self.log("="*60)
                self.log("v3.0 Ultimate 监测任务完成！", 'success')
                self.log(f"统计: 总计{len(results)} | 成功{success} | 失败{fail} | 反爬{anti_crawler}", 'info')
                self.log(f"耗时: {int(minutes)}分{int(seconds)}秒", 'info')
                self.log(f"报告: {os.path.basename(output_file)}", 'success')
                self.log("="*60)

                messagebox.showinfo(
                    "✅ v3.0 Ultimate 完成",
                    f"v3.0 Ultimate 监测完成！\n\n"
                    f"总计: {len(results)} 条\n"
                    f"✅ 成功: {success} 条\n"
                    f"❌ 失败: {fail} 条\n"
                    f"🛡️ 反爬触发: {anti_crawler} 条\n"
                    f"⏱️ 耗时: {int(minutes)}分{int(seconds)}秒\n\n"
                    f"v3.0 新特性:\n"
                    f"✅ 智能反爬检测\n"
                    f"✅ 随机时间间隔\n"
                    f"✅ 多重登录重试\n"
                    f"✅ 修复版封面抓取\n\n"
                    f"报告: {os.path.basename(output_file)}"
                )

        except Exception as e:
            self.log(f"v3.0任务执行失败: {str(e)}", 'error')
            messagebox.showerror("错误", f"v3.0监测失败:\n{str(e)}")

        finally:
            self.is_running = False
            self.start_button.config(state='normal')
            self.stop_button.config(state='disabled')
            self.update_progress(0, 1, "就绪")

    def start_monitoring(self):
        if not self.is_running:
            thread = threading.Thread(target=self.monitor_task, daemon=True)
            thread.start()

    def stop_monitoring(self):
        if self.is_running:
            self.stop_flag = True
            self.log("正在停止v3.0监测任务...", 'warning')

def main():
    root = tk.Tk()

    # 设置窗口图标和位置
    root.update_idletasks()
    width = 1100
    height = 850
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')

    app = UltimateMonitorV3GUI(root)
    root.mainloop()

if __name__ == '__main__':
    main()
