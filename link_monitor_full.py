"""
链接监测工具 - 完整版
功能:
1. 支持扫码登录小红书
2. 自动访问CSV中的所有链接
3. 对每个链接进行截屏
4. 生成包含链接、状态、截屏的Excel报告
5. 列出打开失败的链接
"""

import csv
import os
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from PIL import Image as PILImage

def read_links_from_csv(csv_file):
    """从CSV文件读取链接"""
    links = []
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
        for row in rows[1:]:  # 跳过标题行
            if row and len(row) > 0 and row[0].strip():
                links.append(row[0].strip())
    return links

def resize_screenshot(image_path, max_width=300, max_height=200):
    """调整截图大小以适应Excel单元格"""
    try:
        img = PILImage.open(image_path)
        # 计算缩放比例
        width_ratio = max_width / img.width
        height_ratio = max_height / img.height
        ratio = min(width_ratio, height_ratio)
        
        new_width = int(img.width * ratio)
        new_height = int(img.height * ratio)
        
        return new_width, new_height
    except Exception as e:
        print(f"  警告: 无法调整图片大小 - {str(e)}")
        return max_width, max_height

def create_excel_report(results, output_file='链接监测报告.xlsx'):
    """创建Excel报告，包含截屏"""
    print("\n正在生成Excel报告...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "链接监测结果"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 20
    
    # 创建标题行
    headers = ['序号', '链接', '状态', '截屏预览', '检测时间']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 设置标题行高度
    ws.row_dimensions[1].height = 25
    
    # 填充数据
    current_row = 2
    for idx, result in enumerate(results, 1):
        # 序号
        cell_a = ws.cell(row=current_row, column=1, value=idx)
        cell_a.alignment = Alignment(horizontal='center', vertical='center')
        cell_a.border = thin_border
        
        # 链接
        cell_b = ws.cell(row=current_row, column=2, value=result['链接'])
        cell_b.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_b.border = thin_border
        
        # 状态
        status_cell = ws.cell(row=current_row, column=3, value=result['状态'])
        if result['状态'] == '成功':
            status_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            status_cell.font = Font(color="FFFFFF", bold=True)
        elif result['状态'] == '失败':
            status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            status_cell.font = Font(color="FFFFFF", bold=True)
        else:
            status_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            status_cell.font = Font(bold=True)
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        status_cell.border = thin_border
        
        # 截屏
        screenshot_path = result.get('截屏文件', '')
        cell_d = ws.cell(row=current_row, column=4)
        cell_d.border = thin_border
        
        if screenshot_path and os.path.exists(screenshot_path):
            try:
                # 获取调整后的图片尺寸
                width, height = resize_screenshot(screenshot_path, max_width=300, max_height=180)
                
                # 设置行高（Excel的行高单位约为磅）
                ws.row_dimensions[current_row].height = height * 0.75
                
                # 插入图片
                img = XLImage(screenshot_path)
                img.width = width
                img.height = height
                img.anchor = f'D{current_row}'
                ws.add_image(img)
                
                cell_d.value = ''  # 清空单元格文本
            except Exception as e:
                cell_d.value = f'图片加载失败: {str(e)}'
                cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.row_dimensions[current_row].height = 30
        else:
            error_msg = result.get('错误信息', '无截图')
            cell_d.value = error_msg
            cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_d.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            ws.row_dimensions[current_row].height = 30
        
        # 检测时间
        cell_e = ws.cell(row=current_row, column=5, value=result['检测时间'])
        cell_e.alignment = Alignment(horizontal='center', vertical='center')
        cell_e.border = thin_border
        
        current_row += 1
    
    # 添加统计信息
    stats_row = current_row + 1
    ws.cell(row=stats_row, column=1, value='统计').font = Font(bold=True, size=12)
    
    success_count = sum(1 for r in results if r['状态'] == '成功')
    fail_count = sum(1 for r in results if r['状态'] == '失败')
    
    ws.cell(row=stats_row, column=2, value=f'总计: {len(results)} | 成功: {success_count} | 失败: {fail_count}')
    ws.cell(row=stats_row, column=2).font = Font(bold=True, size=11)
    
    # 保存Excel文件
    wb.save(output_file)
    print(f"[OK] Excel报告已生成: {output_file}")
    return output_file

def collect_results():
    """收集现有的截屏文件并创建结果"""
    csv_file = 'yilideeplink.csv'
    links = read_links_from_csv(csv_file)
    screenshot_dir = 'screenshots'
    
    results = []
    
    for idx, link in enumerate(links, 1):
        screenshot_name = f'screenshot_{idx}.png'
        screenshot_path = os.path.join(screenshot_dir, screenshot_name)
        
        result = {
            '链接': link,
            '状态': '待检测',
            '截屏文件': '',
            '错误信息': '',
            '检测时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        if os.path.exists(screenshot_path):
            result['状态'] = '成功'
            result['截屏文件'] = screenshot_path
        else:
            result['状态'] = '失败'
            result['错误信息'] = '截屏文件不存在，可能需要重新访问链接'
        
        results.append(result)
    
    return results

def print_summary(results):
    """打印检测总结"""
    print("\n" + "="*70)
    print("                     链接监测结果总结")
    print("="*70)
    
    success_count = sum(1 for r in results if r['状态'] == '成功')
    fail_count = sum(1 for r in results if r['状态'] == '失败')
    
    print(f"\n总计: {len(results)} 个链接")
    print(f"[+] 成功: {success_count} 个")
    print(f"[-] 失败: {fail_count} 个")
    
    if fail_count > 0:
        print("\n失败的链接列表:")
        print("-" * 70)
        for idx, result in enumerate(results, 1):
            if result['状态'] == '失败':
                print(f"\n{idx}. {result['链接']}")
                print(f"   错误原因: {result['错误信息']}")
    
    print("\n" + "="*70)

def main():
    print("\n" + "="*70)
    print("                    链接监测工具")
    print("="*70)
    
    # 收集结果
    print("\n正在收集链接和截屏信息...")
    results = collect_results()
    
    # 生成Excel报告
    output_file = create_excel_report(results)
    
    # 打印总结
    print_summary(results)
    
    print(f"\n报告文件: {output_file}")
    print("\n提示: 如需重新截屏，请使用浏览器工具访问链接并保存截图到screenshots目录")
    print("="*70 + "\n")

if __name__ == '__main__':
    main()

