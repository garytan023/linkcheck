"""
链接监测工具 - GUI版本
功能：
1. CSV文件选择
2. 输出文件夹选择
3. 定时运行设置
4. 实时进度显示
5. 日志输出
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import csv
import os
import threading
import schedule
import time
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image as PILImage

class LinkMonitorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("链接监测工具 v2.0")
        self.root.geometry("900x700")
        self.root.resizable(True, True)
        
        # 变量
        self.csv_file = tk.StringVar(value="yilideeplink.csv")
        self.output_dir = tk.StringVar(value="./output")
        self.schedule_enabled = tk.BooleanVar(value=False)
        self.schedule_time = tk.StringVar(value="09:00")
        self.is_running = False
        self.stop_flag = False
        
        self.create_widgets()
        
    def create_widgets(self):
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置行列权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # ===== 文件选择区域 =====
        file_frame = ttk.LabelFrame(main_frame, text="文件设置", padding="10")
        file_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        file_frame.columnconfigure(1, weight=1)
        
        # CSV文件选择
        ttk.Label(file_frame, text="输入CSV:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(file_frame, textvariable=self.csv_file, width=50).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(file_frame, text="浏览...", command=self.browse_csv).grid(row=0, column=2, padx=5)
        
        # 输出文件夹选择
        ttk.Label(file_frame, text="输出目录:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(file_frame, textvariable=self.output_dir, width=50).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(file_frame, text="浏览...", command=self.browse_output).grid(row=1, column=2, padx=5)
        
        # ===== 定时设置区域 =====
        schedule_frame = ttk.LabelFrame(main_frame, text="定时任务", padding="10")
        schedule_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Checkbutton(schedule_frame, text="启用定时运行", 
                       variable=self.schedule_enabled,
                       command=self.toggle_schedule).grid(row=0, column=0, sticky=tk.W)
        
        ttk.Label(schedule_frame, text="运行时间:").grid(row=0, column=1, padx=10)
        self.time_entry = ttk.Entry(schedule_frame, textvariable=self.schedule_time, width=10)
        self.time_entry.grid(row=0, column=2)
        self.time_entry.config(state='disabled')
        
        ttk.Label(schedule_frame, text="(格式: HH:MM，如 09:00)").grid(row=0, column=3, padx=5)
        
        # ===== 控制按钮区域 =====
        button_frame = ttk.Frame(main_frame, padding="10")
        button_frame.grid(row=2, column=0, columnspan=3, pady=10)
        
        self.start_button = ttk.Button(button_frame, text="立即开始监测", 
                                      command=self.start_monitoring, width=20)
        self.start_button.grid(row=0, column=0, padx=5)
        
        self.stop_button = ttk.Button(button_frame, text="停止", 
                                     command=self.stop_monitoring, 
                                     state='disabled', width=15)
        self.stop_button.grid(row=0, column=1, padx=5)
        
        ttk.Button(button_frame, text="清空日志", 
                  command=self.clear_log, width=15).grid(row=0, column=2, padx=5)
        
        # ===== 进度条 =====
        progress_frame = ttk.Frame(main_frame)
        progress_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        progress_frame.columnconfigure(0, weight=1)
        
        self.progress = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        self.progress_label = ttk.Label(progress_frame, text="就绪")
        self.progress_label.grid(row=1, column=0, pady=2)
        
        # ===== 日志区域 =====
        log_frame = ttk.LabelFrame(main_frame, text="运行日志", padding="5")
        log_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(4, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, width=80, 
                                                 wrap=tk.WORD, state='disabled')
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # ===== 状态栏 =====
        self.status_bar = ttk.Label(main_frame, text="就绪", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
    def browse_csv(self):
        """浏览并选择CSV文件"""
        filename = filedialog.askopenfilename(
            title="选择CSV文件",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.csv_file.set(filename)
            self.log(f"已选择CSV文件: {filename}")
    
    def browse_output(self):
        """浏览并选择输出目录"""
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.output_dir.set(directory)
            self.log(f"已选择输出目录: {directory}")
    
    def toggle_schedule(self):
        """切换定时任务状态"""
        if self.schedule_enabled.get():
            self.time_entry.config(state='normal')
            self.log("定时任务已启用")
        else:
            self.time_entry.config(state='disabled')
            self.log("定时任务已禁用")
    
    def log(self, message):
        """添加日志"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_message = f"[{timestamp}] {message}\n"
        
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, log_message)
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')
        
        # 更新状态栏
        self.status_bar.config(text=message)
    
    def clear_log(self):
        """清空日志"""
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        self.log("日志已清空")
    
    def update_progress(self, current, total, message=""):
        """更新进度条"""
        progress = (current / total) * 100 if total > 0 else 0
        self.progress['value'] = progress
        self.progress_label.config(text=f"{message} ({current}/{total}) - {progress:.0f}%")
        self.root.update_idletasks()
    
    def read_links_from_csv(self, csv_file):
        """从CSV文件读取链接"""
        links = []
        try:
            with open(csv_file, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows = list(reader)
                for row in rows[1:]:  # 跳过标题行
                    if row and len(row) > 0 and row[0].strip():
                        links.append(row[0].strip())
        except Exception as e:
            self.log(f"错误: 无法读取CSV文件 - {str(e)}")
            raise
        return links
    
    def resize_screenshot(self, image_path, max_width=300, max_height=200):
        """调整截图大小"""
        try:
            img = PILImage.open(image_path)
            width_ratio = max_width / img.width
            height_ratio = max_height / img.height
            ratio = min(width_ratio, height_ratio)
            new_width = int(img.width * ratio)
            new_height = int(img.height * ratio)
            return new_width, new_height
        except Exception as e:
            self.log(f"警告: 无法调整图片大小 - {str(e)}")
            return max_width, max_height
    
    def create_excel_report(self, results, output_dir):
        """创建Excel报告"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "链接监测结果"
            
            # 设置列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 45
            ws.column_dimensions['E'].width = 20
            
            # 创建标题行
            headers = ['序号', '链接', '状态', '截屏预览', '检测时间']
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            ws.row_dimensions[1].height = 25
            
            # 填充数据
            current_row = 2
            for idx, result in enumerate(results, 1):
                cell_a = ws.cell(row=current_row, column=1, value=idx)
                cell_a.alignment = Alignment(horizontal='center', vertical='center')
                cell_a.border = thin_border
                
                cell_b = ws.cell(row=current_row, column=2, value=result['链接'])
                cell_b.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell_b.border = thin_border
                
                status_cell = ws.cell(row=current_row, column=3, value=result['状态'])
                if result['状态'] == '成功':
                    status_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    status_cell.font = Font(color="FFFFFF", bold=True)
                else:
                    status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    status_cell.font = Font(color="FFFFFF", bold=True)
                status_cell.alignment = Alignment(horizontal='center', vertical='center')
                status_cell.border = thin_border
                
                screenshot_path = result.get('截屏文件', '')
                cell_d = ws.cell(row=current_row, column=4)
                cell_d.border = thin_border
                
                if screenshot_path and os.path.exists(screenshot_path):
                    try:
                        width, height = self.resize_screenshot(screenshot_path, max_width=300, max_height=180)
                        ws.row_dimensions[current_row].height = height * 0.75
                        img = XLImage(screenshot_path)
                        img.width = width
                        img.height = height
                        img.anchor = f'D{current_row}'
                        ws.add_image(img)
                    except Exception as e:
                        cell_d.value = f'图片加载失败: {str(e)}'
                        cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.row_dimensions[current_row].height = 30
                else:
                    cell_d.value = result.get('错误信息', '无截图')
                    cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell_d.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                    ws.row_dimensions[current_row].height = 30
                
                cell_e = ws.cell(row=current_row, column=5, value=result['检测时间'])
                cell_e.alignment = Alignment(horizontal='center', vertical='center')
                cell_e.border = thin_border
                
                current_row += 1
            
            # 统计信息
            stats_row = current_row + 1
            success_count = sum(1 for r in results if r['状态'] == '成功')
            fail_count = sum(1 for r in results if r['状态'] == '失败')
            
            ws.cell(row=stats_row, column=1, value='统计').font = Font(bold=True, size=12)
            ws.cell(row=stats_row, column=2, value=f'总计: {len(results)} | 成功: {success_count} | 失败: {fail_count}')
            ws.cell(row=stats_row, column=2).font = Font(bold=True, size=11)
            
            # 保存文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(output_dir, f'链接监测报告_{timestamp}.xlsx')
            wb.save(output_file)
            
            return output_file, success_count, fail_count
            
        except Exception as e:
            self.log(f"错误: 生成Excel报告失败 - {str(e)}")
            raise
    
    def monitor_task(self):
        """执行监测任务"""
        try:
            self.is_running = True
            self.stop_flag = False
            self.start_button.config(state='disabled')
            self.stop_button.config(state='normal')
            
            # 验证文件
            csv_file = self.csv_file.get()
            if not os.path.exists(csv_file):
                messagebox.showerror("错误", f"CSV文件不存在: {csv_file}")
                return
            
            # 创建输出目录
            output_dir = self.output_dir.get()
            os.makedirs(output_dir, exist_ok=True)
            
            # 创建截图目录
            screenshot_dir = os.path.join(output_dir, 'screenshots')
            os.makedirs(screenshot_dir, exist_ok=True)
            
            self.log("="*60)
            self.log("开始链接监测任务")
            self.log(f"输入文件: {csv_file}")
            self.log(f"输出目录: {output_dir}")
            self.log("="*60)
            
            # 读取链接
            self.log("正在读取链接...")
            links = self.read_links_from_csv(csv_file)
            self.log(f"找到 {len(links)} 个链接")
            
            if not links:
                messagebox.showwarning("警告", "CSV文件中没有找到链接")
                return
            
            # 使用Playwright访问链接并截图
            self.log("正在启动浏览器...")
            results = self.capture_with_playwright(links, screenshot_dir)
            
            if not self.stop_flag:
                # 生成报告
                self.log("正在生成Excel报告...")
                output_file, success_count, fail_count = self.create_excel_report(results, output_dir)
                
                self.log("="*60)
                self.log("监测任务完成！")
                self.log(f"总计: {len(results)} 个链接")
                self.log(f"成功: {success_count} 个")
                self.log(f"失败: {fail_count} 个")
                self.log(f"报告文件: {output_file}")
                self.log("="*60)
                
                messagebox.showinfo("完成", f"监测完成！\n\n总计: {len(results)}\n成功: {success_count}\n失败: {fail_count}\n\n报告已保存至:\n{output_file}")
            
        except Exception as e:
            self.log(f"错误: {str(e)}")
            messagebox.showerror("错误", f"监测过程中发生错误:\n{str(e)}")
        
        finally:
            self.is_running = False
            self.start_button.config(state='normal')
            self.stop_button.config(state='disabled')
            self.update_progress(0, 1, "就绪")
    
    def capture_with_playwright(self, links, screenshot_dir):
        """使用Playwright捕获截图"""
        from playwright.sync_api import sync_playwright
        
        results = []
        
        self.log("浏览器将打开，如需登录请在首次访问时完成")
        self.log("-" * 60)
        
        with sync_playwright() as p:
            # 启动浏览器
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(viewport={'width': 1920, 'height': 1080})
            page = context.new_page()
            
            # 先访问主页
            self.log("正在打开小红书主页...")
            page.goto('https://www.xiaohongshu.com/')
            self.log("等待10秒，请完成登录（如已登录可忽略）...")
            time.sleep(10)
            
            # 访问每个链接
            for idx, link in enumerate(links, 1):
                if self.stop_flag:
                    self.log("任务已被用户停止")
                    break
                
                self.update_progress(idx, len(links), f"正在访问链接 {idx}")
                self.log(f"[{idx}/{len(links)}] 正在访问: {link}")
                
                result = {
                    '链接': link,
                    '状态': '失败',
                    '截屏文件': '',
                    '错误信息': '',
                    '检测时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                try:
                    # 访问链接
                    page.goto(link, timeout=30000)
                    self.log(f"  等待页面加载...")
                    time.sleep(3)
                    
                    # 截图
                    screenshot_name = f'screenshot_{idx}.png'
                    screenshot_path = os.path.join(screenshot_dir, screenshot_name)
                    page.screenshot(path=screenshot_path)
                    
                    result['状态'] = '成功'
                    result['截屏文件'] = screenshot_path
                    self.log(f"  [OK] 截图已保存: {screenshot_name}")
                    
                except Exception as e:
                    result['错误信息'] = f'访问失败: {str(e)}'
                    self.log(f"  [FAIL] 错误: {str(e)}")
                
                results.append(result)
            
            # 关闭浏览器
            browser.close()
            self.log("浏览器已关闭")
        
        return results
    
    def start_monitoring(self):
        """开始监测"""
        if not self.is_running:
            thread = threading.Thread(target=self.monitor_task, daemon=True)
            thread.start()
    
    def stop_monitoring(self):
        """停止监测"""
        if self.is_running:
            self.stop_flag = True
            self.log("正在停止任务...")

def main():
    root = tk.Tk()
    app = LinkMonitorGUI(root)
    root.mainloop()

if __name__ == '__main__':
    main()

