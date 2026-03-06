"""
WPP MD 小红书链接监测工具 - Pro专业版
增强功能：封面抓取、多图支持、悬停交互、详细内容抓取
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import csv
import os
import threading
import time
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image as PILImage
from playwright.sync_api import sync_playwright
import json
from urllib.parse import urljoin
import re
import sys
import string

def resolve_playwright_browsers_path():
    """Resolve bundled Playwright browsers path for packaged exe."""
    candidates = []
    # PyInstaller解包目录
    if getattr(sys, "_MEIPASS", None):
        candidates.append(os.path.join(sys._MEIPASS, "browsers"))
    # 当前目录（与exe同目录）
    candidates.append(os.path.join(os.getcwd(), "browsers"))
    # 用户目录缓存
    candidates.append(os.path.join(os.path.expanduser('~'), '.playwright-browsers'))
    for path in candidates:
        if path and os.path.exists(path):
            return path
    return None

# 设置Playwright浏览器路径（优先使用打包内置的browsers目录）
if not os.environ.get('PLAYWRIGHT_BROWSERS_PATH'):
    pw_path = resolve_playwright_browsers_path()
    if pw_path:
        os.environ['PLAYWRIGHT_BROWSERS_PATH'] = pw_path

class ProLinkMonitorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("WPP MD 小红书链接监测工具 Pro")
        self.root.geometry("950x800")
        
        # 配色
        self.colors = {
            'bg': '#5B6FB5',
            'primary': '#5B7BF5',
            'secondary': '#6B5FB5',
            'white': '#FFFFFF',
            'text_light': '#E8EEFF',
            'text_dark': '#2D3748',
            'success': '#48BB78',
            'danger': '#F56565',
        }
        
        # 变量
        self.csv_file = tk.StringVar(value="yilideeplink.csv")
        self.output_dir = tk.StringVar(value="./output")
        self.schedule_enabled = tk.BooleanVar(value=False)
        self.schedule_time = tk.StringVar(value="09:00")
        self.login_wait_time = tk.StringVar(value="30")
        self.capture_mode = tk.StringVar(value="full")  # simple/full
        self.all_links_mode = tk.BooleanVar(value=False)  # 是否对所有链接仅截图校验
        self.is_running = False
        self.stop_flag = False
        self.http_user_agent = (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36"
        )
        # 存储Playwright登录态，扫码一次后可复用
        self.storage_state_path = os.path.join(os.getcwd(), "xhs_pro_storage.json")
        # 持久化浏览器用户数据目录，防止每次都要扫码
        self.user_data_dir = os.path.join(os.getcwd(), "xhs_pro_profile")
        # 控制粉丝数获取的兜底策略，默认启用API兜底以提高准确性
        self.enable_api_fans = True
        self.enable_homepage_fans = True
        
        self.root.configure(bg=self.colors['bg'])
        self.create_ui()
        
    def create_ui(self):
        # 主容器
        main_canvas = tk.Canvas(self.root, bg=self.colors['bg'], highlightthickness=0)
        scrollbar = tk.Scrollbar(self.root, orient="vertical", command=main_canvas.yview)
        scrollable_frame = tk.Frame(main_canvas, bg=self.colors['bg'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        
        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=930)
        main_canvas.configure(yscrollcommand=scrollbar.set)

        # 允许鼠标滚轮上下滑动
        def _on_mousewheel(event):
            delta = -1 * int(event.delta / 120) if event.delta else 0
            main_canvas.yview_scroll(delta, "units")
            return "break"
        main_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        # 兼容触控板/其他平台
        main_canvas.bind_all("<Button-4>", lambda e: main_canvas.yview_scroll(-1, "units"))
        main_canvas.bind_all("<Button-5>", lambda e: main_canvas.yview_scroll(1, "units"))
        
        main_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        content = tk.Frame(scrollable_frame, bg=self.colors['bg'])
        content.pack(fill=tk.BOTH, expand=True, padx=25, pady=25)
        
        # 标题
        title_frame = tk.Frame(content, bg=self.colors['bg'])
        title_frame.pack(pady=(0, 20))
        
        logo_canvas = tk.Canvas(title_frame, width=70, height=70, bg=self.colors['bg'], highlightthickness=0)
        logo_canvas.pack()
        logo_canvas.create_oval(10, 10, 60, 60, fill=self.colors['primary'], outline='')
        logo_canvas.create_oval(22, 22, 48, 48, fill=self.colors['white'], outline='')
        logo_canvas.create_oval(30, 30, 40, 40, fill=self.colors['primary'], outline='')
        
        tk.Label(
            title_frame,
            text="WPP MD 小红书链接监测工具 Pro",
            font=('Microsoft YaHei UI', 20, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['white']
        ).pack(pady=(10, 2))
        
        tk.Label(
            title_frame,
            text="Professional Link Monitoring Tool - 增强版",
            font=('Arial', 9),
            bg=self.colors['bg'],
            fg=self.colors['text_light']
        ).pack()
        
        # 文件配置
        file_card = self.create_card(content, "📁 文件配置")
        
        csv_row = tk.Frame(file_card, bg='white')
        csv_row.pack(fill=tk.X, pady=5)
        
        tk.Label(csv_row, text="输入CSV文件:", font=('Microsoft YaHei UI', 9), bg='white', fg=self.colors['text_dark']).pack(side=tk.LEFT)
        
        csv_entry_frame = tk.Frame(csv_row, bg='#F7FAFC')
        csv_entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 5))
        
        tk.Entry(
            csv_entry_frame,
            textvariable=self.csv_file,
            font=('Consolas', 9),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT
        ).pack(fill=tk.X, padx=8, pady=6)
        
        tk.Button(
            csv_row,
            text="浏览...",
            command=self.browse_csv,
            font=('Microsoft YaHei UI', 9),
            bg=self.colors['primary'],
            fg='white',
            relief=tk.FLAT,
            padx=15,
            pady=4,
            cursor='hand2'
        ).pack(side=tk.RIGHT)
        
        output_row = tk.Frame(file_card, bg='white')
        output_row.pack(fill=tk.X, pady=5)
        
        tk.Label(output_row, text="输出目录:", font=('Microsoft YaHei UI', 9), bg='white', fg=self.colors['text_dark']).pack(side=tk.LEFT, padx=(0, 15))
        
        output_entry_frame = tk.Frame(output_row, bg='#F7FAFC')
        output_entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 5))
        
        tk.Entry(
            output_entry_frame,
            textvariable=self.output_dir,
            font=('Consolas', 9),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT
        ).pack(fill=tk.X, padx=8, pady=6)
        
        tk.Button(
            output_row,
            text="浏览...",
            command=self.browse_output,
            font=('Microsoft YaHei UI', 9),
            bg=self.colors['primary'],
            fg='white',
            relief=tk.FLAT,
            padx=15,
            pady=4,
            cursor='hand2'
        ).pack(side=tk.RIGHT)
        
        # 抓取模式（新增）
        mode_card = self.create_card(content, "🎯 抓取模式 (Pro)")
        
        mode_row = tk.Frame(mode_card, bg='white')
        mode_row.pack(fill=tk.X, pady=5)
        
        tk.Radiobutton(
            mode_row,
            text="快速模式（仅截图）",
            variable=self.capture_mode,
            value="simple",
            font=('Microsoft YaHei UI', 9),
            bg='white',
            fg=self.colors['text_dark'],
            activebackground='white',
            selectcolor='white',
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=(0, 20))
        
        tk.Radiobutton(
            mode_row,
            text="完整模式（封面+详情+多图）",
            variable=self.capture_mode,
            value="full",
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='white',
            fg=self.colors['primary'],
            activebackground='white',
            selectcolor='white',
            cursor='hand2'
        ).pack(side=tk.LEFT)
        
        mode_desc = tk.Label(
            mode_card,
            text="✨ 完整模式会抓取：封面图、标题、作者、点赞数、评论数、收藏数、多图内容",
            font=('Arial', 8),
            bg='white',
            fg='#718096'
        )
        mode_desc.pack(pady=(5, 0))

        # 链接类型选项
        link_mode_card = self.create_card(content, "🌐 链接类型")
        link_row = tk.Frame(link_mode_card, bg='white')
        link_row.pack(fill=tk.X, pady=5)
        tk.Radiobutton(
            link_row,
            text="小红书专用（需登录抓取详情）",
            variable=self.all_links_mode,
            value=False,
            font=('Microsoft YaHei UI', 9),
            bg='white',
            fg=self.colors['text_dark'],
            activebackground='white',
            selectcolor='white',
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=(0, 20))
        tk.Radiobutton(
            link_row,
            text="所有链接（仅截图+有效性检测）",
            variable=self.all_links_mode,
            value=True,
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='white',
            fg=self.colors['primary'],
            activebackground='white',
            selectcolor='white',
            cursor='hand2'
        ).pack(side=tk.LEFT)
        tk.Label(
            link_mode_card,
            text="选择“所有链接”后跳过登录，对每个链接做快捷访问与截图校验。",
            font=('Arial', 8),
            bg='white',
            fg='#718096',
            anchor='w',
            justify='left'
        ).pack(fill=tk.X, pady=(4, 0))
        
        # 定时任务
        schedule_card = self.create_card(content, "⏰ 定时任务")
        
        schedule_row = tk.Frame(schedule_card, bg='white')
        schedule_row.pack(fill=tk.X, pady=5)
        
        self.schedule_check = tk.Checkbutton(
            schedule_row,
            text="启用定时运行",
            variable=self.schedule_enabled,
            command=self.toggle_schedule,
            font=('Microsoft YaHei UI', 9),
            bg='white',
            fg=self.colors['text_dark'],
            activebackground='white',
            selectcolor='white',
            cursor='hand2'
        )
        self.schedule_check.pack(side=tk.LEFT)
        
        tk.Label(schedule_row, text="运行时间:", font=('Microsoft YaHei UI', 9), bg='white', fg=self.colors['text_dark']).pack(side=tk.LEFT, padx=(20, 8))
        
        time_frame = tk.Frame(schedule_row, bg='#F7FAFC')
        time_frame.pack(side=tk.LEFT)
        
        self.time_entry = tk.Entry(
            time_frame,
            textvariable=self.schedule_time,
            font=('Consolas', 9),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            width=8,
            state='disabled'
        )
        self.time_entry.pack(padx=8, pady=6)
        
        tk.Label(schedule_row, text="(格式: HH:MM)", font=('Arial', 8), bg='white', fg='#A0AEC0').pack(side=tk.LEFT, padx=(5, 0))
        
        # 登录设置
        login_card = self.create_card(content, "🔐 登录设置")
        
        login_row = tk.Frame(login_card, bg='white')
        login_row.pack(fill=tk.X, pady=5)
        
        tk.Label(
            login_row,
            text="扫码等待时间:",
            font=('Microsoft YaHei UI', 9),
            bg='white',
            fg=self.colors['text_dark']
        ).pack(side=tk.LEFT)
        
        wait_time_frame = tk.Frame(login_row, bg='#F7FAFC')
        wait_time_frame.pack(side=tk.LEFT, padx=(10, 5))
        
        tk.Entry(
            wait_time_frame,
            textvariable=self.login_wait_time,
            font=('Consolas', 9),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            width=6
        ).pack(padx=8, pady=6)
        
        tk.Label(
            login_row,
            text="秒 (建议30-60秒)",
            font=('Arial', 8),
            bg='white',
            fg='#A0AEC0'
        ).pack(side=tk.LEFT, padx=(5, 0))
        
        # 控制按钮
        button_frame = tk.Frame(content, bg=self.colors['bg'])
        button_frame.pack(pady=15)
        
        self.start_button = tk.Button(
            button_frame,
            text="▶  立即开始监测",
            command=self.start_monitoring,
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            relief=tk.FLAT,
            padx=30,
            pady=12,
            cursor='hand2'
        )
        self.start_button.pack(side=tk.LEFT, padx=5)
        
        self.stop_button = tk.Button(
            button_frame,
            text="⬛  停止",
            command=self.stop_monitoring,
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg=self.colors['danger'],
            fg='white',
            relief=tk.FLAT,
            padx=25,
            pady=12,
            state='disabled',
            cursor='hand2'
        )
        self.stop_button.pack(side=tk.LEFT, padx=5)
        
        # 进度
        progress_card = self.create_card(content, "📊 执行进度")
        
        style = ttk.Style()
        style.theme_use('clam')
        style.configure(
            "Custom.Horizontal.TProgressbar",
            troughcolor='#E2E8F0',
            bordercolor='#E2E8F0',
            background=self.colors['primary'],
            lightcolor=self.colors['primary'],
            darkcolor=self.colors['primary']
        )
        
        self.progress = ttk.Progressbar(
            progress_card,
            mode='determinate',
            style="Custom.Horizontal.TProgressbar",
            length=400
        )
        self.progress.pack(fill=tk.X, pady=(5, 8))
        
        self.progress_label = tk.Label(
            progress_card,
            text="就绪",
            font=('Microsoft YaHei UI', 9),
            bg='white',
            fg=self.colors['text_dark']
        )
        self.progress_label.pack()
        
        # 日志
        log_card = self.create_card(content, "📝 运行日志")
        
        log_toolbar = tk.Frame(log_card, bg='white')
        log_toolbar.pack(fill=tk.X, pady=(0, 5))
        
        tk.Button(
            log_toolbar,
            text="清空日志",
            command=self.clear_log,
            font=('Microsoft YaHei UI', 8),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            padx=10,
            pady=3,
            cursor='hand2'
        ).pack(side=tk.RIGHT)
        
        self.log_text = scrolledtext.ScrolledText(
            log_card,
            height=12,
            font=('Consolas', 8),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            wrap=tk.WORD,
            state='disabled'
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
        
        # 状态栏
        status_frame = tk.Frame(content, bg='#4B5FA5', height=35)
        status_frame.pack(fill=tk.X, pady=(10, 0))
        status_frame.pack_propagate(False)
        
        self.status_bar = tk.Label(
            status_frame,
            text="● 就绪",
            font=('Microsoft YaHei UI', 9),
            bg='#4B5FA5',
            fg=self.colors['text_light'],
            anchor=tk.W
        )
        self.status_bar.pack(side=tk.LEFT, padx=15, fill=tk.Y)
        
    def create_card(self, parent, title):
        card_outer = tk.Frame(parent, bg='white', relief=tk.FLAT)
        card_outer.pack(fill=tk.X, pady=(0, 12))
        
        title_frame = tk.Frame(card_outer, bg='white')
        title_frame.pack(fill=tk.X, padx=15, pady=(10, 5))
        
        tk.Label(
            title_frame,
            text=title,
            font=('Microsoft YaHei UI', 11, 'bold'),
            bg='white',
            fg=self.colors['text_dark']
        ).pack(side=tk.LEFT)
        
        tk.Frame(card_outer, bg='#E2E8F0', height=1).pack(fill=tk.X, padx=15)
        
        content = tk.Frame(card_outer, bg='white')
        content.pack(fill=tk.BOTH, expand=True, padx=15, pady=(8, 10))
        
        return content
    
    def browse_csv(self):
        filename = filedialog.askopenfilename(
            title="选择CSV文件",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.csv_file.set(filename)
            self.log(f"[选择] CSV: {os.path.basename(filename)}")
    
    def browse_output(self):
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.output_dir.set(directory)
            self.log(f"[选择] 输出: {directory}")
    
    def toggle_schedule(self):
        if self.schedule_enabled.get():
            self.time_entry.config(state='normal', bg='white')
            self.log("[定时] 已启用")
        else:
            self.time_entry.config(state='disabled', bg='#F7FAFC')
            self.log("[定时] 已禁用")
    
    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {message}\n"
        
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, log_message)
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')
        
        self.status_bar.config(text=f"● {message}")
    
    def clear_log(self):
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        self.log("[清空] 日志已清空")
    
    def update_progress(self, current, total, message=""):
        progress = (current / total) * 100 if total > 0 else 0
        self.progress['value'] = progress
        self.progress_label.config(text=f"{message} ({current}/{total}) • {progress:.0f}%")
        self.root.update_idletasks()
    
    def read_items_from_csv(self, csv_file):
        """读取CSV，支持列: 产品, 序号, 链接；兼容单列链接"""
        items = []
        try:
            with open(csv_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                if reader.fieldnames and ('链接' in reader.fieldnames or reader.fieldnames[0]):
                    for idx, row in enumerate(reader, 1):
                        link = (row.get('链接') or row.get(reader.fieldnames[0]) or '').strip()
                        if not link:
                            continue
                        product = (row.get('产品') or '').strip()
                        seq = (row.get('序号') or '').strip() or str(idx)
                        items.append({'产品': product, '序号': seq, '链接': link})
                        continue
        except Exception:
            pass

        # 兼容旧格式：单列链接
        if not items:
            try:
                with open(csv_file, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                    for idx, row in enumerate(rows[1:], 1):
                        if row and len(row) > 0 and row[0].strip():
                            items.append({'产品': '', '序号': str(idx), '链接': row[0].strip()})
            except Exception:
                pass
        return items

    def ensure_media_loaded(self, page):
        """Scroll page to trigger lazy-loaded images."""
        try:
            page.evaluate("() => window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(800)
            page.evaluate("() => window.scrollTo(0, document.body.scrollHeight / 2)")
            page.wait_for_timeout(400)
            page.evaluate("() => window.scrollTo(0, 0)")
            page.wait_for_timeout(200)
        except Exception:
            pass

    def normalize_media_url(self, base_url, url):
        """Normalize protocol-relative or relative URLs to absolute."""
        try:
            if not url:
                return ""
            if url.startswith("//"):
                return "https:" + url
            if url.startswith("http://") or url.startswith("https://"):
                return url
            return urljoin(base_url, url)
        except Exception:
            return url or ""

    def is_unwanted_image_src(self, src: str) -> bool:
        """Heuristics to skip avatars/logos/watermarks."""
        if not src:
            return True
        lowered = src.lower()
        unwanted_keywords = [
            "avatar", "profile", "logo", "icon", "watermark", "badge",
            "placeholder", "default_avatar", "default-avatar", "favicon",
            "xhslogo", "/head/", "/avatar/", "logos", "brand"
        ]
        return any(k in lowered for k in unwanted_keywords)

    def element_is_large_enough(self, elem) -> bool:
        """Check DOM element visual size to avoid tiny icons."""
        try:
            box = elem.bounding_box()
            if not box:
                return False
            return (box.get("width", 0) >= 180) and (box.get("height", 0) >= 180)
        except Exception:
            return False

    def is_large_image_file(self, image_path: str, min_width: int = 160, min_height: int = 160) -> bool:
        """Check saved image file dimensions to avoid wrong covers."""
        try:
            with PILImage.open(image_path) as im:
                w, h = im.size
                return w >= min_width and h >= min_height
        except Exception:
            # If we cannot open the file, keep it to avoid over-filtering
            return True

    def sanitize_name(self, name: str) -> str:
        try:
            safe = re.sub(r'[^0-9A-Za-z\u4e00-\u9fff]+', '_', name)
            return safe.strip('_') or "item"
        except Exception:
            return "item"

    def capture_video_cover(self, page, output_dir, name_prefix):
        """Capture video screenshot - only screenshot of paused video player container."""
        covers = []
        
        try:
            self.log("    [视频] 视频笔记 - 截取播放器静止画面...")
            
            # 1. 暂停视频并定位到第0秒（静止状态）
            try:
                page.evaluate("""() => {
                    const v = document.querySelector('video');
                    if (v) { 
                        v.pause(); 
                        v.currentTime = 0; 
                    }
                }""")
                time.sleep(1.5)  # 等待视频暂停并回到第0秒
            except Exception:
                pass
            
            # 2. 直接截取视频容器（包含封面/首帧）
            container_selectors = [
                'video',  # 优先直接截video元素
                '.xgplayer',  # 小红书播放器类名
                '[class*="video-container"]',
                '[class*="player"]',
                '.video-wrapper',
            ]
            
            container = None
            for sel in container_selectors:
                container = page.query_selector(sel)
                if container:
                    break
            
            if container:
                try:
                    cover_name = f'{name_prefix}_cover_video.png'
                    cover_path = os.path.join(output_dir, cover_name)
                    container.screenshot(path=cover_path, timeout=10000)
                    if self.is_large_image_file(cover_path, min_width=100, min_height=100):
                        covers.append(cover_path)
                        self.log(f"    [视频] ✓ 截图成功")
                    else:
                        try:
                            os.remove(cover_path)
                        except Exception:
                            pass
                except Exception as e:
                    self.log(f"    [视频] 截图失败: {str(e)}")
            else:
                self.log("    [视频] ✗ 未找到视频容器")
                
        except Exception as e:
            self.log(f"    [视频] 整体捕获失败: {str(e)}")
        
        return covers

    def step_through_carousel(self, page, output_dir, name_prefix, cover_paths, seen_urls, limit=10):
        """Attempt to step a swiper/carousel to capture multiple frames."""
        try:
            next_selectors = [
                '.swiper-button-next',
                '[class*="swiper"] [class*="next"]',
                'button[aria-label*="next"]',
                '[class*="arrow"][class*="next"]',
            ]
            bullet_selector = '.swiper-pagination-bullet, [class*="swiper"] [class*="bullet"]'

            def get_active_img_src():
                try:
                    src = page.evaluate("""() => {
                        const active = document.querySelector('.swiper-slide-active');
                        if (active) {
                            const img = active.querySelector('img');
                            if (img) return img.getAttribute('src') || img.getAttribute('data-src') || '';
                        }
                        const container = document.querySelector('[class*="swiper"], .carousel');
                        if (container) {
                            const img = container.querySelector('img');
                            if (img) return img.getAttribute('src') || img.getAttribute('data-src') || '';
                        }
                        return '';
                    }""")
                    return src or ""
                except Exception:
                    return ""

            # 优先使用分页圆点，逐个点击可稳定触发懒加载
            bullets = page.query_selector_all(bullet_selector)
            if bullets and len(bullets) > 1:
                for i, b in enumerate(bullets[:limit]):
                    try:
                        b.click()
                        page.wait_for_timeout(700)
                    except Exception:
                        continue
                    src = get_active_img_src()
                    if src:
                        normalized = self.normalize_media_url(page.url, src)
                        if normalized and normalized not in seen_urls and not self.is_unwanted_image_src(normalized):
                            cover_name = f'{name_prefix}_cover_{len(cover_paths)+1}.png'
                            cover_path = os.path.join(output_dir, cover_name)
                            if self.download_image_via_page(page, normalized, cover_path):
                                if self.is_large_image_file(cover_path):
                                    cover_paths.append(cover_path)
                                else:
                                    try:
                                        os.remove(cover_path)
                                    except Exception:
                                        pass
                            seen_urls.add(normalized)
                            if len(cover_paths) >= 10:
                                return cover_paths

            # 再尝试点击下一张按钮
            for step in range(limit):
                src = get_active_img_src()
                if src:
                    normalized = self.normalize_media_url(page.url, src)
                    if normalized and normalized not in seen_urls and not self.is_unwanted_image_src(normalized):
                        cover_name = f'{name_prefix}_cover_{len(cover_paths)+1}.png'
                        cover_path = os.path.join(output_dir, cover_name)
                        if self.download_image_via_page(page, normalized, cover_path):
                            if self.is_large_image_file(cover_path):
                                cover_paths.append(cover_path)
                            else:
                                try:
                                    os.remove(cover_path)
                                except Exception:
                                    pass
                        seen_urls.add(normalized)
                        if len(cover_paths) >= 10:
                            break

                # Click next
                clicked = False
                for sel in next_selectors:
                    btn = page.query_selector(sel)
                    if btn:
                        try:
                            btn.click()
                            clicked = True
                            break
                        except Exception:
                            continue
                # 如果没有next按钮，尝试键盘右箭头
                if not clicked:
                    try:
                        page.keyboard.press("ArrowRight")
                        clicked = True
                    except Exception:
                        pass
                if not clicked:
                    break
                page.wait_for_timeout(600)
        except Exception:
            pass
        return cover_paths

    def fetch_user_fans_via_api(self, page, author_id):
        """Call Xiaohongshu API to obtain accurate fan counts."""
        try:
            if not author_id:
                return ""
            api_url = f"https://edith.xiaohongshu.com/api/sns/web/v1/user/otherinfo?user_id={author_id}"
            headers = {
                "Referer": page.url,
                "User-Agent": self.http_user_agent,
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                "Origin": "https://www.xiaohongshu.com",
                "Sec-Fetch-Mode": "cors",
                "Sec-Fetch-Site": "same-site",
            }
            response = page.context.request.get(api_url, headers=headers, timeout=15000)
            if response.ok:
                data = response.json().get("data") or {}
                fans_value = (
                    data.get("fans")
                    or data.get("fansCount")
                    or data.get("fans_count")
                    or data.get("followers")
                    or data.get("follows")
                )
                if fans_value is not None:
                    return str(fans_value)
            else:
                self.log(f"  [API] 请求失败: {response.status} {response.status_text}")
        except Exception as e:
            self.log(f"  [API] 粉丝数获取失败: {str(e)}")
        return ""

    def fetch_user_fans_via_homepage(self, context, author_home_url):
        """Open author homepage in a new tab to fetch fans count as fallback."""
        if not author_home_url:
            return ""
        try:
            self.log("  [主页] 打开博主主页获取粉丝数...")
            home_page = context.new_page()
            home_page.goto(author_home_url, wait_until='domcontentloaded', timeout=20000)
            time.sleep(2)
            fans_text = home_page.evaluate("""() => {
                const bodyText = document.body.innerText.substring(0, 8000);
                const patterns = [
                    /(粉丝|关注者)[^\\d]{0,3}([0-9]+\\.?[0-9]*[万wWkK]?)/i,
                    /([0-9]+\\.?[0-9]*[万wWkK]?)\\s*(粉丝|关注者)/i,
                    /粉丝数[^\\d]{0,3}([0-9]+\\.?[0-9]*[万wWkK]?)/i,
                ];
                for (const p of patterns) {
                    const m = bodyText.match(p);
                    if (m) {
                        return m[2] || m[1] || '';
                    }
                }
                // 常见粉丝元素
                const selectors = [
                    '.fans', '.followers', '.user-data .item', '[class*=\"fans\"]', '[class*=\"Followers\"]'
                ];
                for (const sel of selectors) {
                    const el = document.querySelector(sel);
                    if (!el) continue;
                    const text = (el.innerText || el.textContent || '').trim();
                    if (text && /[0-9]/.test(text) && text.length < 50) {
                        const numMatch = text.match(/([0-9]+\\.?[0-9]*[万wWkK]?)/);
                        if (numMatch) return numMatch[1];
                    }
                }
                return '';
            }""")
            try:
                home_page.close()
            except Exception:
                pass
            return fans_text or ""
        except Exception as e:
            self.log(f"  [主页] 粉丝数获取失败: {str(e)}")
            return ""
    
    def extract_note_id(self, url):
        """从URL提取笔记ID"""
        try:
            import re
            match = re.search(r'/explore/([a-zA-Z0-9]+)', url)
            if match:
                return match.group(1)
        except:
            pass
        return ''
    


    def extract_detailed_content(self, page, url):
        """Extract detailed note info for Pro mode."""
        details = {
            "笔记ID": self.extract_note_id(url),
            "标题": "",
            "作者昵称": "",
            "作者ID": "",
            "粉丝数": "",
            "博主主页": "",
            "发布时间": "",
            "点赞数": "",
            "收藏数": "",
            "评论数": "",
            "分享数": "",
            "封面链接": "",
            "视频链接": "",
            "正文": "",
            "封面图数量": 0
        }
        image_urls = []
        note_id = details["笔记ID"]
        
        try:
            self.log("  [等待] 页面加载...")
            time.sleep(2.5)
            
            try:
                page_data = page.evaluate(r"""(noteId) => {
                    const pick = (...values) => {
                        for (const val of values) {
                            if (val !== undefined && val !== null && val !== '') {
                                return val;
                            }
                        }
                        return '';
                    };
                    
                    const result = {
                        title: '',
                        authorName: '',
                        authorId: '',
                        authorHome: '',
                        fans: '',
                        publishTime: '',
                        likes: '',
                        collects: '',
                        comments: '',
                        shares: '',
                        coverUrl: '',
                        videoUrl: '',
                        content: '',
                        images: []
                    };
                    
                    const globalState = window.__INITIAL_STATE__ || window.__REDUX_STATE__ || {};
                    const noteStore = globalState.note || globalState.noteDetail || (globalState.NoteView && globalState.NoteView.note) || {};
                    const detailMap = noteStore.noteDetailMap || noteStore.notes || noteStore.noteInfoMap || {};
                    let detail = null;
                    if (noteId && detailMap[noteId]) {
                        detail = detailMap[noteId];
                    } else {
                        const values = Object.values(detailMap);
                        if (values.length > 0) detail = values[0];
                    }
                    
                    if (detail) {
                        const note = detail.note || detail;
                        const noteUserId = note.user_id || note.userId;
                        const user = detail.user || note.user || (globalState.user && globalState.user.userDetailMap && globalState.user.userDetailMap[noteUserId]) || {};
                        const stats = detail.noteStat || detail.noteStats || detail.interactInfo || note.interactInfo || detail.statistics || {};
                        
                        result.title = (note.title || note.note_title || '').trim();
                        result.content = (note.desc || note.content || note.note_desc || '').trim();
                        result.publishTime = pick(note.time, note.publish_time, note.create_time, note.first_publish_time);
                        
                        result.authorName = pick(user.nickname, user.nick_name, user.name);
                        const authorId = pick(user.user_id, user.userid, user.id, noteUserId);
                        if (authorId) {
                            result.authorId = authorId.toString();
                            result.authorHome = `https://www.xiaohongshu.com/user/profile/${authorId}`;
                        }
                        
                        const fansValue = pick(user.fans, user.fans_num, user.fansNum, user.fansCount, user.followers);
                        if (fansValue !== '') result.fans = fansValue.toString();
                        
                        result.likes = pick(stats.likes, stats.likeCount, stats.like_cnt, stats.likedCount, stats.like_num, stats.noteLikeCount);
                        result.collects = pick(stats.collectCount, stats.collect_cnt, stats.collects, stats.favoriteCount, stats.saveCount, stats.collect_num);
                        result.comments = pick(stats.commentCount, stats.comment_cnt, stats.comments, stats.comment_num);
                        result.shares = pick(stats.shareCount, stats.share_cnt, stats.share, stats.share_num);
                        
                        const imageList = note.imageList || note.images || detail.imageList || [];
                        result.images = imageList
                            .map(img => (img && (img.url || img.url_default || img.img_url || img.original)) || '')
                            .filter(Boolean);
                        if (result.images.length > 0) {
                            result.coverUrl = result.images[0];
                        }
                        
                        const video = note.video || note.videoInfo || note.videoPlay || {};
                        const media = video.media || {};
                        const stream = media.stream || {};
                        result.videoUrl = pick(video.url, video.playUrl, video.play_url, stream.h264, stream.ld, stream.sd);
                    }
                    
                    return result;
                }""", note_id)
                
                if page_data:
                    def set_if_empty(key, value):
                        if value and not details[key]:
                            if key == "标题":
                                details[key] = str(value).strip()[:200]
                            elif key == "正文":
                                details[key] = str(value).strip()[:500]
                            else:
                                details[key] = str(value).strip()
                    
                    set_if_empty("标题", page_data.get("title", ""))
                    set_if_empty("作者昵称", page_data.get("authorName", ""))
                    set_if_empty("作者ID", page_data.get("authorId", ""))
                    set_if_empty("博主主页", page_data.get("authorHome", ""))
                    set_if_empty("粉丝数", page_data.get("fans", ""))
                    set_if_empty("发布时间", page_data.get("publishTime", ""))
                    set_if_empty("点赞数", page_data.get("likes", ""))
                    set_if_empty("收藏数", page_data.get("collects", ""))
                    set_if_empty("评论数", page_data.get("comments", ""))
                    set_if_empty("分享数", page_data.get("shares", ""))
                    set_if_empty("封面链接", page_data.get("coverUrl", ""))
                    set_if_empty("视频链接", page_data.get("videoUrl", ""))
                    set_if_empty("正文", page_data.get("content", ""))
                    
                    image_urls = page_data.get("images", []) or []
                    if image_urls:
                        details["封面图数量"] = len(image_urls)
                        if not details["封面链接"]:
                            details["封面链接"] = image_urls[0]
                        self.log(f"  [State] 找到{len(image_urls)}个图片URL")
                    else:
                        self.log(f"  [State] 未找到图片URL (可能是视频)")
                    
                    self.log(f"  [成功] 标题: {details['标题'][:30] if details['标题'] else '无'}")
                    self.log(f"  [成功] 作者: {details['作者昵称']} (粉丝: {details['粉丝数'] or '未知'})")
            except Exception as e:
                self.log(f"  [State抓取] 失败: {str(e)}")
            
            try:
                dom_data = page.evaluate("""() => {
                    const data = {
                        title: '',
                        authorName: '',
                        authorId: '',
                        authorHome: '',
                        fans: '',
                        publishTime: '',
                        likes: '',
                        collects: '',
                        comments: '',
                        shares: '',
                        coverUrl: '',
                        videoUrl: '',
                        content: '',
                        images: []
                    };
                    
                    const titleEl = document.querySelector('#detail-title, [id*="title"], .title');
                    if (titleEl) data.title = titleEl.innerText || titleEl.textContent || '';
                    
                    const authorEl = document.querySelector('.author-name, .username, .user-nickname, [class*="author"][class*="name"]');
                    if (authorEl) data.authorName = authorEl.innerText || authorEl.textContent || '';
                    
                    const authorLink = document.querySelector('a[href*="/user/profile"]');
                    if (authorLink) {
                        data.authorHome = authorLink.href || '';
                        const match = data.authorHome.match(/profile\\/([a-zA-Z0-9]+)/);
                        if (match) data.authorId = match[1];
                    }
                    
                    const fansEl = document.querySelector('.fans-count, [class*="fans"], .follower-count');
                    if (fansEl) data.fans = fansEl.innerText || fansEl.textContent || '';
                    
                    const timeEl = document.querySelector('.publish-time, .date, [class*="time"]');
                    if (timeEl) data.publishTime = timeEl.innerText || timeEl.textContent || '';
                    
                    const interactEls = document.querySelectorAll('[class*="interact"] span, [class*="count"]');
                    interactEls.forEach(el => {
                        const text = el.innerText || el.textContent || '';
                        const className = el.className || '';
                        if (!data.likes && (className.includes('like') || el.closest('[class*="like"]'))) data.likes = text;
                        if (!data.collects && (className.includes('collect') || el.closest('[class*="collect"]'))) data.collects = text;
                        if (!data.comments && (className.includes('comment') || el.closest('[class*="comment"]'))) data.comments = text;
                        if (!data.shares && (className.includes('share') || el.closest('[class*="share"]'))) data.shares = text;
                    });
                    
                    const contentEl = document.querySelector('#detail-desc, .note-content, .content, [class*="desc"]');
                    if (contentEl) data.content = (contentEl.innerText || contentEl.textContent || '').slice(0, 500);
                    
                    const coverImg = document.querySelector('img[src*="sns-img"]');
                    if (coverImg) data.coverUrl = coverImg.src || '';
                    
                    const video = document.querySelector('video source, video');
                    if (video) data.videoUrl = video.src || (video.querySelector('source') && video.querySelector('source').src) || '';
                    
                    data.images = Array.from(document.querySelectorAll('img[src*="sns-img"]')).map(img => img.src || '').filter(Boolean);
                    
                    return data;
                }""")
                
                if dom_data:
                    for key, target in [
                        ("title", "标题"),
                        ("authorName", "作者昵称"),
                        ("authorId", "作者ID"),
                        ("authorHome", "博主主页"),
                        ("fans", "粉丝数"),
                        ("publishTime", "发布时间"),
                        ("likes", "点赞数"),
                        ("collects", "收藏数"),
                        ("comments", "评论数"),
                        ("shares", "分享数"),
                        ("coverUrl", "封面链接"),
                        ("videoUrl", "视频链接"),
                        ("content", "正文"),
                    ]:
                        if dom_data.get(key) and not details[target]:
                            value = dom_data.get(key)
                            if target == "标题":
                                value = value[:200]
                            if target == "正文":
                                value = value[:500]
                            details[target] = value.strip()
                    
                    if not image_urls:
                        image_urls = dom_data.get("images", []) or []
                        if image_urls:
                            details["封面图数量"] = len(image_urls)
                            if not details["封面链接"]:
                                details["封面链接"] = image_urls[0]
                            self.log(f"  [DOM] 找到{len(image_urls)}个图片URL")
                        else:
                            self.log(f"  [DOM] 未找到图片URL")
            except Exception as e:
                self.log(f"  [DOM抓取] 失败: {str(e)}")
            
            if not details["粉丝数"]:
                try:
                    self.log("  [悬停] 尝试获取粉丝数...")
                    
                    # 扩展作者元素选择器 - 确保视频和图片笔记都能正确选中
                    author_selectors = [
                        # 视频笔记特有的选择器
                        '.video-info .author-name',
                        '.side-bar .user-name',
                        '[class*="video"] .author',
                        # 图片笔记选择器
                        '.author-container .name',
                        '.author-wrapper .name',
                        'a[href*="/user/profile"] .name',
                        '.user-nickname',
                        '.note-author .name',
                        # 通用容器
                        'a[href*="/user/profile"]',
                        '.author-wrapper',
                        '.user-info',
                        '.author-container',
                    ]
                    
                    author_element = None
                    used_selector = ""
                    for selector in author_selectors:
                        try:
                            author_element = page.query_selector(selector)
                            if author_element:
                                used_selector = selector
                                self.log(f"  [悬停] 使用选择器: {selector}")
                                break
                        except Exception:
                            continue
                    
                    if author_element:
                        self.log(f"  [悬停] 找到作者元素，开始悬停...")
                        # 滚动到可见区域
                        try:
                            author_element.scroll_into_view_if_needed()
                        except Exception:
                            pass
                        
                        # 悬停操作
                        try:
                            box = author_element.bounding_box()
                            if box:
                                # 鼠标移动到元素中心
                                page.mouse.move(box['x'] + box['width']/2, box['y'] + box['height']/2)
                            else:
                                author_element.hover()
                        except Exception:
                            try:
                                author_element.hover()
                            except Exception:
                                pass
                        
                        # 等待悬浮卡片出现（加长等待时间）
                        time.sleep(3.5)
                        
                        # 读取悬浮卡片中的粉丝数
                        fans_data = page.evaluate("""() => {
                            const result = {
                                text: '',
                                source: '',
                                allPopups: []
                            };
                            
                            // 1. 收集所有可能的悬浮卡片（用于调试）
                            const cardSelectors = [
                                '.user-card-container',
                                '.user-info-card', 
                                '[class*="UserCard"]',
                                '[class*="popup"]',
                                '[class*="popover"]',
                                '[class*="tooltip"]',
                                '.author-card',
                                '[class*="user-card"]',
                            ];
                            
                            for (const sel of cardSelectors) {
                                const cards = document.querySelectorAll(sel);
                                for (const card of cards) {
                                    if (card.offsetParent !== null) {
                                        const cardText = card.innerText || '';
                                        result.allPopups.push({
                                            selector: sel,
                                            hasText: cardText.length > 0,
                                            hasFans: cardText.includes('粉丝'),
                                            text: cardText.substring(0, 100)
                                        });
                                        
                                        if (cardText.includes('粉丝')) {
                                            result.text = cardText;
                                            result.source = sel;
                                            return result;
                                        }
                                    }
                                }
                            }
                            
                            // 2. 查找所有包含"粉丝"的可见元素（最后渲染的优先）
                            const elements = Array.from(document.querySelectorAll('div, span, p, a'));
                            for (let i = elements.length - 1; i >= 0; i--) {
                                const el = elements[i];
                                const text = el.innerText || '';
                                if (el.offsetParent !== null && 
                                    text.includes('粉丝') && 
                                    /[\\d]+/.test(text) &&
                                    text.length < 200) {
                                    result.text = text;
                                    result.source = 'generic-element';
                                    return result;
                                }
                            }
                            
                            return result;
                        }""")
                        
                        fans_text = fans_data.get('text', '') if isinstance(fans_data, dict) else fans_data
                        
                        # 调试输出
                        if isinstance(fans_data, dict):
                            all_popups = fans_data.get('allPopups', [])
                            if all_popups:
                                self.log(f"  [调试] 找到{len(all_popups)}个弹窗元素")
                                for popup in all_popups[:3]:  # 只显示前3个
                                    self.log(f"    - {popup.get('selector','')}: hasFans={popup.get('hasFans',False)}")
                            else:
                                self.log(f"  [调试] 未找到任何弹窗元素")
                            
                            if fans_data.get('source'):
                                self.log(f"  [调试] 粉丝数来源: {fans_data.get('source')}")
                        
                        if fans_text:
                            self.log(f"  [调试] 原始文本: {fans_text[:150]}")
                            
                            # 支持多种格式：粉丝: 1.2万 / 1.2万 粉丝 / 粉丝 1.2万
                            patterns = [
                                r'(粉丝|关注者|followers)[:\s：]*([0-9]+\.?[0-9]*[万wWkK]?)',
                                r'([0-9]+\.?[0-9]*[万wWkK]?)\s*(粉丝|关注者|followers)',
                            ]
                            
                            matched = False
                            for pattern in patterns:
                                match = re.search(pattern, fans_text, re.IGNORECASE)
                                if match:
                                    # 判断哪个组是数字
                                    if re.match(r'[0-9]', match.group(1)):
                                        num_part = match.group(1)
                                    else:
                                        num_part = match.group(2)
                                    details["粉丝数"] = num_part
                                    self.log(f"  [悬停] ✓ 粉丝数: {details['粉丝数']}")
                                    matched = True
                                    break
                            
                            if not matched:
                                self.log(f"  [悬停] ✗ 提取失败，原始文本: {fans_text[:100]}")
                        else:
                            self.log("  [悬停] ✗ 未找到包含粉丝的文本")
                    else:
                        self.log("  [悬停] 未找到作者元素")
                except Exception as e:
                    self.log(f"  [悬停] 失败: {str(e)}")
            
            # 全局文本搜索作为兜底
            if not details["粉丝数"]:
                try:
                    self.log("  [文本] 从页面全文搜索粉丝数...")
                    fans_text = page.evaluate("""() => {
                        const bodyText = document.body.innerText.substring(0, 8000); 
                        const patterns = [
                            /(?:粉丝|关注者)[:\\s]*([0-9]+\\.?[0-9]*[万wWkK]?)/i,
                            /([0-9]+\\.?[0-9]*[万wWkK]?)\\s*(?:粉丝|关注者)/i,
                        ];
                        for (const pattern of patterns) {
                            const match = bodyText.match(pattern);
                            if (match) return match[1];
                        }
                        return '';
                    }""")
                    if fans_text:
                        details["粉丝数"] = fans_text
                        self.log(f"  [文本] 粉丝数: {details['粉丝数']}")
                except Exception:
                    pass

            # API兜底，提升粉丝数准确性（已登录状态下请求）
            if self.enable_api_fans and not details["粉丝数"] and details.get("作者ID"):
                api_fans = self.fetch_user_fans_via_api(page, details.get("作者ID"))
                if api_fans:
                    details["粉丝数"] = api_fans
                    self.log(f"  [API] 粉丝数: {api_fans}")

            # 博主主页兜底（已登录或需要额外信息时）
            if self.enable_homepage_fans and not details["粉丝数"] and details.get("博主主页"):
                home_fans = self.fetch_user_fans_via_homepage(page.context, details.get("博主主页"))
                if home_fans:
                    details["粉丝数"] = home_fans
                    self.log(f"  [主页] 粉丝数: {home_fans}")
        
        except Exception as e:
            self.log(f"  [错误] 详情抓取失败: {str(e)}")
        
        return details, image_urls


    def download_image_via_page(self, page, image_url, save_path):
        """Download cover image bytes through page fetch."""
        headers = {
            "Referer": page.url,
            "User-Agent": self.http_user_agent,
            "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
        }
        try:
            response = page.context.request.get(image_url, headers=headers, timeout=20000)
            if response.ok:
                with open(save_path, 'wb') as f:
                    f.write(response.body())
                return True
        except Exception:
            pass

        try:
            data = page.evaluate(
                """async (url) => {
                    try {
                        const response = await fetch(url, { credentials: 'include' });
                        if (!response.ok) return null;
                        const buffer = await response.arrayBuffer();
                        return Array.from(new Uint8Array(buffer));
                    } catch (err) {
                        return null;
                    }
                }""",
                image_url,
            )
            if data:
                with open(save_path, 'wb') as f:
                    f.write(bytes(data))
                return True
        except Exception as e:
            self.log(f"    [封面] 下载失败: {str(e)}")
        return False


    def capture_cover_images(self, page, output_dir, name_prefix, image_urls=None):
        """Capture cover images (supports multiple) and return local paths."""
        cover_paths = []
        image_urls = image_urls or []
        
        try:
            self.log(f"    [封面] 开始抓取 (元数据URL数: {len(image_urls)})")
            
            # 策略1: 元数据URL直接下载
            if image_urls and len(image_urls) > 0:
                self.log(f"    [封面] 发现{len(image_urls)}个元数据图片URL")
                seen = set()
                for idx, url in enumerate(image_urls[:10], 1):
                    if not url or url in seen:
                        continue
                    normalized = self.normalize_media_url(page.url, url)
                    if normalized in seen or self.is_unwanted_image_src(normalized):
                        continue
                    seen.add(normalized)
                    cover_name = f'{name_prefix}_cover_{len(cover_paths)+1}.png'
                    cover_path = os.path.join(output_dir, cover_name)
                    
                    self.log(f"      [{idx}] 下载: {normalized[:60]}...")
                    if self.download_image_via_page(page, normalized, cover_path):
                        if self.is_large_image_file(cover_path, min_width=100, min_height=100):
                            cover_paths.append(cover_path)
                            self.log(f"      [{idx}] ✓ 保存成功")
                        else:
                            self.log(f"      [{idx}] ✗ 图片太小，跳过")
                            try:
                                os.remove(cover_path)
                            except Exception:
                                pass
                    else:
                        self.log(f"      [{idx}] ✗ 下载失败")
                    
                    if len(cover_paths) >= 10:
                        break
                
                if cover_paths:
                    self.log(f"    [封面] 元数据下载完成: {len(cover_paths)}张")
                    # 不要直接return，继续尝试获取更多
                else:
                    self.log("    [封面] 元数据下载未成功，尝试其他方式...")
            
            # 检查是否为视频笔记（优先判断）
            is_video = page.evaluate("""() => {
                return document.querySelector('video') !== null;
            }""")
            
            # 如果是视频笔记且元数据未成功，优先用视频方法
            if is_video and len(cover_paths) < 1:
                self.log("    [封面] 检测到视频笔记，优先使用视频抓取...")
                video_covers = self.capture_video_cover(page, output_dir, name_prefix)
                if video_covers:
                    cover_paths.extend(video_covers)
                    self.log(f"    [封面] 视频方法成功: {len(video_covers)}张")
                    # 视频抓取成功就直接返回，不再用DOM
                    if len(cover_paths) >= 1:
                        return cover_paths
            
            # 如果不是视频或视频抓取失败，尝试DOM抓取
            if len(cover_paths) < 1:
                self.log("    [封面] 尝试DOM直接抓取...")
                self.ensure_media_loaded(page)
                
                # 使用JS直接获取所有图片信息，更可靠
                img_info = page.evaluate("""() => {
                    const images = [];
                    const seen = new Set();
                    
                    // 小红书轮播图（优先级最高）
                    document.querySelectorAll('.swiper-slide img, .carousel img, [class*="Swiper"] img').forEach(img => {
                        const src = img.src || img.getAttribute('data-src') || '';
                        if (src && !seen.has(src)) {
                            seen.add(src);
                            images.push({
                                src: src,
                                width: img.naturalWidth || img.width || 0,
                                height: img.naturalHeight || img.height || 0,
                                type: 'swiper'
                            });
                        }
                    });
                    
                    // 笔记内容图
                    document.querySelectorAll('img[src*="sns-img"], img[src*="xhscdn"], .note-content img').forEach(img => {
                        const src = img.src || img.getAttribute('data-src') || '';
                        if (src && !seen.has(src)) {
                            seen.add(src);
                            images.push({
                                src: src,
                                width: img.naturalWidth || img.width || 0,
                                height: img.naturalHeight || img.height || 0,
                                type: 'content'
                            });
                        }
                    });
                    
                    // 所有其他img
                    document.querySelectorAll('img').forEach(img => {
                        const src = img.src || img.getAttribute('data-src') || '';
                        if (src && !seen.has(src)) {
                            seen.add(src);
                            images.push({
                                src: src,
                                width: img.naturalWidth || img.width || 0,
                                height: img.naturalHeight || img.height || 0,
                                type: 'other'
                            });
                        }
                    });
                    
                    return images;
                }""")
                
                self.log(f"    [DOM] 找到{len(img_info)}个候选图片")
                
                seen_urls = set()
                for idx, info in enumerate(img_info[:15], 1):
                    src = info.get('src', '')
                    if not src:
                        continue
                    
                    normalized = self.normalize_media_url(page.url, src)
                    if normalized in seen_urls or self.is_unwanted_image_src(normalized):
                        continue
                    
                    seen_urls.add(normalized)
                    cover_name = f'{name_prefix}_cover_{len(cover_paths)+1}.png'
                    cover_path = os.path.join(output_dir, cover_name)
                    
                    # 下载图片
                    if self.download_image_via_page(page, normalized, cover_path):
                        if self.is_large_image_file(cover_path, min_width=100, min_height=100):
                            cover_paths.append(cover_path)
                            if len(cover_paths) == 1:
                                self.log(f"    [DOM] 开始保存图片...")
                            self.log(f"      [{len(cover_paths)}] {info.get('type','')} {info.get('width',0)}x{info.get('height',0)}")
                        else:
                            try:
                                os.remove(cover_path)
                            except Exception:
                                pass
                    
                    if len(cover_paths) >= 10:
                        break

            # Step through carousel if still few covers (非视频笔记才尝试轮播)
            if len(cover_paths) < 2 and not is_video:
                self.log("    [封面] 尝试轮播逐帧...")
                try:
                    self.step_through_carousel(page, output_dir, name_prefix, cover_paths, seen_urls if 'seen_urls' in locals() else set(), limit=6)
                except Exception as e:
                    self.log(f"    [封面] 轮播失败: {str(e)}")

            # 如果还有空间，尝试收集页面中明显的背景封面元素
            if len(cover_paths) < 3:
                try:
                    bg_elems = page.query_selector_all('[style*="background-image"], .xhs-image, .cover')
                    for el in bg_elems:
                        if len(cover_paths) >= 5:
                            break
                        try:
                            bg = page.evaluate("(el)=>getComputedStyle(el).backgroundImage", el)
                            if not isinstance(bg, str) or "url(" not in bg:
                                continue
                            match = re.search(r'url\\(["\\\']?(.*?)["\\\']?\\)', bg)
                            if not match:
                                continue
                            src = self.normalize_media_url(page.url, match.group(1))
                            if src and src not in seen_urls and not self.is_unwanted_image_src(src):
                                cover_name = f'{name_prefix}_cover_{len(cover_paths)+1}.png'
                                cover_path = os.path.join(output_dir, cover_name)
                                if self.download_image_via_page(page, src, cover_path):
                                    if self.is_large_image_file(cover_path):
                                        cover_paths.append(cover_path)
                                    else:
                                        try:
                                            os.remove(cover_path)
                                        except Exception:
                                            pass
                                seen_urls.add(src)
                        except Exception:
                            continue
                except Exception:
                    pass
        
            if cover_paths:
                self.log(f"    [封面] ✓ 成功保存 {len(cover_paths)} 张")
            else:
                self.log("    [封面] ✗ 未找到任何有效封面图")
        
        except Exception as e:
            self.log(f"    [错误] 封面抓取失败: {str(e)}")
        
        return cover_paths

    def capture_with_playwright_pro(self, items, screenshot_dir, all_links_mode=False):
        """Pro版本的抓取功能"""
        results = []
        covers_dir = os.path.join(screenshot_dir, 'covers')
        os.makedirs(covers_dir, exist_ok=True)
        
        self.log("[浏览器] 正在启动...")
        
        with sync_playwright() as p:
            try:
                os.makedirs(self.user_data_dir, exist_ok=True)
                context = p.chromium.launch_persistent_context(
                    user_data_dir=self.user_data_dir,
                    headless=False,
                    viewport={'width': 1920, 'height': 1080},
                    user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                )
                self.log(f"[登录] 使用持久化用户目录: {self.user_data_dir}")
            except Exception as e:
                self.log(f"[错误] 无法启动浏览器 - {str(e)}")
                raise
            
            page = context.pages[0] if context.pages else context.new_page()
            page.set_default_timeout(60000)
            
            login_success = False
            if not all_links_mode:
                self.log("[主页] 正在打开小红书...")
                try:
                    page.goto('https://www.xiaohongshu.com/', wait_until='domcontentloaded', timeout=30000)
                except Exception:
                    page.goto('https://www.xiaohongshu.com/', timeout=30000)
                time.sleep(3)
                
                try:
                    wait_seconds = int(self.login_wait_time.get())
                    wait_seconds = max(5, min(wait_seconds, 300))
                except:
                    wait_seconds = 30

                # 先通过持久化Cookie初步判断
                try:
                    cookies = context.cookies()
                    has_auth_cookie = any(
                        c.get("name") in ["a1", "web_session", "a1.sig"] and "xiaohongshu.com" in (c.get("domain") or "")
                        for c in cookies
                    )
                    if has_auth_cookie:
                        login_success = True
                        self.log("[✓就绪] 检测到历史Cookie登录态，跳过扫码倒计时")
                except Exception:
                    pass

                # 如果已经有登录态，直接跳过等待
                try:
                    precheck_login = page.evaluate("""() => {
                        const avatar = document.querySelector('.avatar, .user-avatar, [class*="Avatar"]');
                        const userInfo = document.querySelector('.user-info, .login-info, [class*="UserInfo"]');
                        const loginBtn = document.querySelector('.login-btn, [class*="login-button"], button:has-text("登录")');
                        const hasCookie = document.cookie.includes('web_session') || 
                                          document.cookie.includes('xsecappid') || 
                                          document.cookie.includes('a1');
                        return (avatar || userInfo) && hasCookie && !loginBtn;
                    }""")
                    if precheck_login:
                        login_success = True
                        self.log("[✓就绪] 已检测到历史登录态，无需扫码")
                except Exception:
                    precheck_login = False
                
                if not login_success:
                    self.log(f"[登录] 等待{wait_seconds}秒扫码...")
                    
                    for remaining in range(wait_seconds, 0, -1):
                        if self.stop_flag:
                            break
                        if remaining % 5 == 0 or remaining <= 5:
                            self.log(f"  [倒计时] {remaining}秒")
                        
                        if remaining % 5 == 0:
                            try:
                                is_logged_in = page.evaluate("""() => {
                                    const avatar = document.querySelector('.avatar, .user-avatar, [class*="Avatar"]');
                                    const userInfo = document.querySelector('.user-info, .login-info, [class*="UserInfo"]');
                                    const loginBtn = document.querySelector('.login-btn, [class*="login-button"], button:has-text("登录")');
                                    const hasCookie = document.cookie.includes('web_session') || 
                                                    document.cookie.includes('xsecappid') || 
                                                    document.cookie.includes('a1');
                                    if ((avatar || userInfo) && hasCookie && !loginBtn) {
                                        return true;
                                    }
                                    return false;
                                }""")
                                
                                if is_logged_in:
                                    login_success = True
                                    self.log(f"  [✓成功] 检测到登录成功！")
                                    time.sleep(2)
                                    break
                            except Exception:
                                pass

                        time.sleep(1)
                if not login_success:
                    try:
                        is_logged_in_final = page.evaluate("""() => {
                            const avatar = document.querySelector('.avatar, .user-avatar, [class*="Avatar"]');
                                const userInfo = document.querySelector('.user-info, .login-info, [class*="UserInfo"]');
                                const loginBtn = document.querySelector('.login-btn, [class*="login-button"], button:has-text("登录")');
                                const hasCookie = document.cookie.includes('web_session') || 
                                                  document.cookie.includes('xsecappid') || 
                                                  document.cookie.includes('a1');
                                return (avatar || userInfo) && hasCookie && !loginBtn;
                        }""")
                        if is_logged_in_final:
                            login_success = True
                            self.log(f"  [✓成功] 检测到登录成功！")
                            time.sleep(2)
                    except Exception:
                        pass
            
            else:
                login_success = True
                page.goto('about:blank')
            
            if not self.stop_flag:
                if login_success:
                    self.log("[✓就绪] 登录成功，开始访问链接...")
                    try:
                        context.storage_state(path=self.storage_state_path)
                        self.log("[登录] 登录状态已保存，下次可直接复用免扫码")
                    except Exception as e:
                        self.log(f"[登录] 状态保存失败: {str(e)}")
                else:
                    self.log("[✗警告] 未检测到登录成功！")
                    self.log("[✗警告] 请确保扫码登录完成，否则每个链接都会弹出登录窗口")
                    self.log("[继续] 将尝试访问链接...")
                    self.log("[等待] 给10秒时间补救，请立即扫码...")
                    for i in range(10, 0, -1):
                        self.log(f"  [倒计时] {i}秒")
                        time.sleep(1)
                        if self.stop_flag:
                            break
            
            is_full_mode = self.capture_mode.get() == "full"
            
            for idx, item in enumerate(items, 1):
                if self.stop_flag:
                    self.log("[停止] 用户中断")
                    break
                
                link = item.get('链接', '')
                product = item.get('产品', '')
                seq_val = item.get('序号', str(idx))
                name_prefix = self.sanitize_name(f"{product}_{seq_val}") if product else self.sanitize_name(f"item_{seq_val}")

                self.update_progress(idx, len(items), f"处理链接 {idx}")
                self.log(f"[{idx}/{len(items)}] {link[:80]}...")
                is_xhs = (not all_links_mode) and ("xiaohongshu.com" in link)
                
                result = {
                    '序号': seq_val,
                    '产品': product,
                    '链接': link,
                    '采集状态': '失败',
                    '错误信息': '',
                    'HTTP状态': '',
                    '笔记ID': '',
                    '标题': '',
                    '作者昵称': '',
                    '作者ID': '',
                    '粉丝数': '',
                    '博主主页': '',
                    '发布时间': '',
                    '点赞数': '',
                    '收藏数': '',
                    '评论数': '',
                    '分享数': '',
                    '封面链接': '',
                    '视频链接': '',
                    '正文': '',
                    '截屏文件': '',
                    '封面图列表': [],
                    '封面图数量': 0,
                    '检测时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                try:
                    # 非小红书通用模式：只访问并截图，简单校验HTTP状态
                    if not is_xhs:
                        resp = page.goto(link, timeout=45000, wait_until='domcontentloaded')
                        status_code = resp.status if resp else None
                        result['HTTP状态'] = status_code or ''
                        if status_code and status_code >= 400:
                            raise Exception(f"HTTP {status_code}")
                        time.sleep(2)
                        screenshot_name = f'{name_prefix}_screenshot.png'
                        screenshot_path = os.path.join(screenshot_dir, screenshot_name)
                        try:
                            page.screenshot(path=screenshot_path, full_page=True, timeout=60000)
                            result['截屏文件'] = screenshot_path
                        except Exception as se:
                            self.log(f"  [截图超时] 使用body兜底: {str(se)}")
                            page.locator("body").screenshot(path=screenshot_path, timeout=60000)
                            result['截屏文件'] = screenshot_path
                        result['采集状态'] = '成功'
                        self.log(f"  [✓] 通用链接完成 | 状态: {status_code or '未知'}")
                    else:
                        # 小红书逻辑
                        max_retries = 2
                        resp = None
                        for attempt in range(max_retries):
                            try:
                                resp = page.goto(link, timeout=45000, wait_until='domcontentloaded')
                                break
                            except Exception as e:
                                if attempt == max_retries - 1:
                                    raise e
                                self.log(f"  [重试] 加载超时，正在重试 ({attempt+1}/{max_retries})...")
                                time.sleep(2)
                        
                        result['HTTP状态'] = resp.status if resp else ''
                        time.sleep(3)
                        
                        has_login_popup = page.evaluate("""() => {
                            const loginModal = document.querySelector('.login-modal, [class*="LoginModal"], [class*="login-dialog"]');
                            const qrCode = document.querySelector('[class*="qrcode"], [class*="QRCode"]');
                            const loginBox = document.querySelector('.login-box, [class*="login-container"]');
                            return !!(loginModal || qrCode || loginBox);
                        }""")
                        
                        if has_login_popup:
                            self.log(f"  [✗错误] 检测到登录弹窗！初始登录未成功！")
                            self.log(f"  [提示] 请在开始前确保扫码登录成功")
                            self.log(f"  [等待] 给30秒时间扫码...")
                            for i in range(30, 0, -5):
                                self.log(f"    剩余 {i} 秒...")
                                time.sleep(5)
                                still_has_popup = page.evaluate("""() => {
                                    const loginModal = document.querySelector('.login-modal, [class*="LoginModal"]');
                                    const qrCode = document.querySelector('[class*="qrcode"], [class*="QRCode"]');
                                    return !!(loginModal || qrCode);
                                }""")
                                if not still_has_popup:
                                    self.log(f"  [✓成功] 登录成功，继续处理...")
                                    try:
                                        context.storage_state(path=self.storage_state_path)
                                        self.log("[登录] 登录状态已保存，下次可直接复用免扫码")
                                    except Exception:
                                        pass
                                    break
                            else:
                                self.log(f"  [✗失败] 仍未登录，跳过此链接...")
                                raise Exception("需要登录但用户未完成登录")
                        
                        screenshot_name = f'{name_prefix}_screenshot.png'
                        screenshot_path = os.path.join(screenshot_dir, screenshot_name)
                        try:
                            page.screenshot(path=screenshot_path, full_page=False, timeout=60000)
                            result['截屏文件'] = screenshot_path
                        except Exception as se:
                            self.log(f"  [截图超时] 使用body兜底: {str(se)}")
                            try:
                                page.locator("body").screenshot(path=screenshot_path, timeout=60000)
                                result['截屏文件'] = screenshot_path
                            except Exception as se2:
                                self.log(f"  [截图失败] {str(se2)}")
                        
                        if is_full_mode:
                            self.log(f"  [完整模式] 开始深度抓取...")
                            details, image_urls = self.extract_detailed_content(page, link)
                            result.update(details)
                            cover_paths = self.capture_cover_images(page, covers_dir, name_prefix, image_urls)
                            result['封面图列表'] = cover_paths
                            result['封面图数量'] = len(cover_paths) if cover_paths else len(image_urls)
                            if not result.get('封面链接') and image_urls:
                                result['封面链接'] = image_urls[0]
                            
                            self.log(f"  [数据] 点赞:{details.get('点赞数','')} 收藏:{details.get('收藏数','')} 评论:{details.get('评论数','')}")
                            self.log(f"  [封面] {len(cover_paths) or len(image_urls)}张 | 链接: {result.get('封面链接','无')[:50]}")
                        
                        result['采集状态'] = '成功'
                        self.log(f"  [✓] 链接{idx}完成")
                    
                except Exception as e:
                    result['错误信息'] = str(e)
                    self.log(f"  [✗] {str(e)}")
                
                results.append(result)
                
                # 链接之间增加延迟，避免触发反爬
                if idx < len(items):
                    time.sleep(2)
            
            try:
                context.close()
            except Exception:
                pass
            self.log("[关闭] 浏览器已关闭")
        
        return results
    
    def resize_image(self, image_path, max_width=200, max_height=150):
        try:
            img = PILImage.open(image_path)
            width_ratio = max_width / img.width
            height_ratio = max_height / img.height
            ratio = min(width_ratio, height_ratio)
            return int(img.width * ratio), int(img.height * ratio)
        except:
            return max_width, max_height
    
    def create_excel_report_pro(self, results, output_dir):
        """Pro版本的Excel报告 - 完整版"""
        os.makedirs(output_dir, exist_ok=True)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "小红书监测Pro"
        
        # 列宽设置（基础字段 + 截图 + 多封面）
        column_widths = {
            'A': 6,    # 序号
            'B': 12,   # 产品
            'C': 50,   # 链接
            'D': 12,   # 采集状态
            'E': 30,   # 错误信息
            'F': 18,   # 笔记ID
            'G': 35,   # 标题
            'H': 15,   # 作者昵称
            'I': 18,   # 作者ID
            'J': 12,   # 粉丝数
            'K': 40,   # 博主主页
            'L': 16,   # 发布时间
            'M': 10,   # 点赞数
            'N': 10,   # 收藏数
            'O': 10,   # 评论数
            'P': 10,   # 分享数
            'Q': 40,   # 封面链接
            'R': 40,   # 视频链接
            'S': 50,   # 正文
            'T': 35,   # 截图
            'U': 22,   # 封面1
            'V': 22,   # 封面2
            'W': 22,   # 封面3
            'X': 22,   # 封面4
            'Y': 22,   # 封面5
            'Z': 22,   # 封面6
            'AA': 22,  # 封面7
            'AB': 22,  # 封面8
            'AC': 22,  # 封面9
            'AD': 22,  # 封面10
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 表头
        headers = [
            '序号', '产品', '链接', '采集状态', '错误信息', '笔记ID', '标题', '作者昵称', '作者ID',
            '粉丝数', '博主主页', '发布时间', '点赞数', '收藏数', '评论数', '分享数',
            '封面链接', '视频链接', '正文', '截图',
            '封面1', '封面2', '封面3', '封面4', '封面5',
            '封面6', '封面7', '封面8', '封面9', '封面10'
        ]
        
        header_fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 30
        
        # 数据行
        current_row = 2
        for result in results:
            row_height = 100
            
            # 1. 序号
            cell = ws.cell(row=current_row, column=1, value=result.get('序号', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
            # 2. 产品
            cell = ws.cell(row=current_row, column=2, value=result.get('产品', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            # 3. 链接
            cell = ws.cell(row=current_row, column=3, value=result['链接'])
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
            
            # 4. 采集状态
            status = result.get('采集状态', '失败')
            cell = ws.cell(row=current_row, column=4, value=status)
            if status == '成功':
                cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=10)
            else:
                cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
            # 5. 错误信息
            cell = ws.cell(row=current_row, column=5, value=result.get('错误信息', ''))
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.font = Font(color='DC3545', size=9)
            cell.border = thin_border
            
            # 6. 笔记ID
            cell = ws.cell(row=current_row, column=6, value=result.get('笔记ID', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Consolas', size=9)
            cell.border = thin_border
            
            # 7. 标题
            cell = ws.cell(row=current_row, column=7, value=result.get('标题', ''))
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.font = Font(bold=True, size=9)
            cell.border = thin_border
            
            # 8. 作者昵称
            cell = ws.cell(row=current_row, column=8, value=result.get('作者昵称', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=9)
            cell.border = thin_border
            
            # 9. 作者ID
            cell = ws.cell(row=current_row, column=9, value=result.get('作者ID', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Consolas', size=9)
            cell.border = thin_border
            
            # 10. 粉丝数
            cell = ws.cell(row=current_row, column=10, value=result.get('粉丝数', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True, color='FF6347', size=9)
            cell.border = thin_border
            
            # 11. 博主主页
            cell = ws.cell(row=current_row, column=11, value=result.get('博主主页', ''))
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.font = Font(size=8, color='0066CC', underline='single')
            cell.border = thin_border
            
            # 12. 发布时间
            cell = ws.cell(row=current_row, column=12, value=result.get('发布时间', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=9)
            cell.border = thin_border
            
            # 13. 点赞数
            cell = ws.cell(row=current_row, column=13, value=result.get('点赞数', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=9)
            cell.border = thin_border
            
            # 14. 收藏数
            cell = ws.cell(row=current_row, column=14, value=result.get('收藏数', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=9)
            cell.border = thin_border
            
            # 15. 评论数
            cell = ws.cell(row=current_row, column=15, value=result.get('评论数', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=9)
            cell.border = thin_border
            
            # 16. 分享数
            cell = ws.cell(row=current_row, column=16, value=result.get('分享数', ''))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=9)
            cell.border = thin_border
            
            # 17. 封面链接
            cell = ws.cell(row=current_row, column=17, value=result.get('封面链接', ''))
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.font = Font(size=8, color='0066CC')
            cell.border = thin_border
            
            # 18. 视频链接
            cell = ws.cell(row=current_row, column=18, value=result.get('视频链接', ''))
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.font = Font(size=8, color='0066CC')
            cell.border = thin_border
            
            # 19. 正文
            cell = ws.cell(row=current_row, column=19, value=result.get('正文', ''))
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.font = Font(size=8)
            cell.border = thin_border

            # 20. 截图（先放截屏）
            screenshot_path = result.get('截屏文件', '')
            cell = ws.cell(row=current_row, column=20)
            cell.border = thin_border
            if screenshot_path and os.path.exists(screenshot_path):
                try:
                    width, height = self.resize_image(screenshot_path, max_width=250, max_height=120)
                    img = XLImage(screenshot_path)
                    img.width = width
                    img.height = height
                    img.anchor = f'T{current_row}'
                    ws.add_image(img)
                    row_height = max(row_height, height * 0.75 + 10)
                except Exception as e:
                    cell.value = f'加载失败: {str(e)}'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.value = '无截图'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 21-30. 多封面展示（最多10张）
            cover_paths = result.get('封面图列表', []) or []
            max_covers = 10
            for i in range(max_covers):
                col_index = 21 + i
                cell = ws.cell(row=current_row, column=col_index)
                cell.border = thin_border
                if i < len(cover_paths) and os.path.exists(cover_paths[i]):
                    try:
                        width, height = self.resize_image(cover_paths[i], max_width=140, max_height=110)
                        img = XLImage(cover_paths[i])
                        img.width = width
                        img.height = height
                        anchor_cols = ['U','V','W','X','Y','Z','AA','AB','AC','AD']
                        anchor_col = anchor_cols[i]
                        img.anchor = f'{anchor_col}{current_row}'
                        ws.add_image(img)
                        row_height = max(row_height, height * 0.75 + 10)
                    except Exception as e:
                        cell.value = f'封面失败'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.value = ''
            
            ws.row_dimensions[current_row].height = row_height
            current_row += 1
        
        # 统计行
        stats_row = current_row + 1
        success_count = sum(1 for r in results if r.get('采集状态') == '成功')
        fail_count = len(results) - success_count
        
        ws.cell(row=stats_row, column=1, value='统计').font = Font(bold=True, size=11, color='4169E1')
        ws.cell(row=stats_row, column=2, value=f'总计: {len(results)} 条 | 成功: {success_count} 条 | 失败: {fail_count} 条').font = Font(bold=True, size=10)
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_dir, f'小红书监测Pro_{timestamp}.xlsx')
        wb.save(output_file)
        
        return output_file, success_count, fail_count, 0
    
    def monitor_task(self):
        try:
            self.is_running = True
            self.stop_flag = False
            self.start_button.config(state='disabled')
            self.stop_button.config(state='normal')
            
            csv_file = self.csv_file.get()
            if not os.path.exists(csv_file):
                messagebox.showerror("错误", f"CSV文件不存在: {csv_file}")
                return
            
            output_dir = self.output_dir.get()
            screenshot_dir = os.path.join(output_dir, 'screenshots')
            os.makedirs(screenshot_dir, exist_ok=True)
            
            self.log("="*50)
            self.log("[任务] 开始Pro监测")
            self.log(f"[输入] {csv_file}")
            self.log(f"[输出] {output_dir}")
            self.log(f"[模式] {'完整模式' if self.capture_mode.get() == 'full' else '快速模式'}")
            self.log("="*50)
            
            items = self.read_items_from_csv(csv_file)
            self.log(f"[链接] 找到 {len(items)} 个")
            
            if not items:
                messagebox.showwarning("警告", "CSV中没有链接")
                return
            
            results = self.capture_with_playwright_pro(
                items, screenshot_dir, all_links_mode=self.all_links_mode.get()
            )
            
            if not self.stop_flag:
                self.log("[报告] 生成Pro版Excel...")
                output_file, success, fail, _ = self.create_excel_report_pro(results, output_dir)
                
                self.log("="*50)
                self.log("[完成] Pro监测任务完成！")
                self.log(f"[统计] 总:{len(results)} | 成功:{success} | 失败:{fail}")
                self.log(f"[文件] {os.path.basename(output_file)}")
                self.log("="*50)
                
                # 显示详细数据
                has_fans = sum(1 for r in results if r.get('粉丝数'))
                has_content = sum(1 for r in results if r.get('正文'))
                
                messagebox.showinfo(
                    "✓ 完成",
                    f"Pro监测完成！\n\n"
                    f"总计: {len(results)} 条\n"
                    f"成功: {success} 条\n"
                    f"失败: {fail} 条\n"
                    f"粉丝数: {has_fans} 条\n"
                    f"正文: {has_content} 条\n\n"
                    f"报告: {os.path.basename(output_file)}"
                )
        
        except Exception as e:
            self.log(f"[错误] {str(e)}")
            messagebox.showerror("错误", f"监测失败:\n{str(e)}")
        
        finally:
            self.is_running = False
            self.start_button.config(state='normal')
            self.stop_button.config(state='disabled')
            self.update_progress(0, 1, "就绪")
    
    def start_monitoring(self):
        if not self.is_running:
            thread = threading.Thread(target=self.monitor_task, daemon=True)
            thread.start()
    
    def stop_monitoring(self):
        if self.is_running:
            self.stop_flag = True
            self.log("[停止] 正在停止...")

def main():
    root = tk.Tk()
    
    root.update_idletasks()
    width = 950
    height = 800
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    app = ProLinkMonitorGUI(root)
    root.mainloop()

if __name__ == '__main__':
    main()
