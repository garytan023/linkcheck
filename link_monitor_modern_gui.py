"""
WPP MD 小红书链接监测工具
现代化苹果风格界面 - 蓝白配色
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import csv
import os
import threading
import time
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image as PILImage
from playwright.sync_api import sync_playwright

class ModernLinkMonitorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("WPP MD 小红书链接监测")
        self.root.geometry("1000x750")
        
        # 苹果风格配色
        self.colors = {
            'primary_blue': '#007AFF',      # 苹果蓝
            'light_blue': '#E8F4FF',        # 浅蓝背景
            'white': '#FFFFFF',             # 纯白
            'gray_bg': '#F5F5F7',           # 浅灰背景
            'text_dark': '#1D1D1F',         # 深色文字
            'text_gray': '#86868B',         # 灰色文字
            'success': '#34C759',           # 成功绿
            'danger': '#FF3B30',            # 危险红
            'border': '#D1D1D6'             # 边框灰
        }
        
        # 设置整体背景
        self.root.configure(bg=self.colors['white'])
        
        # 变量
        self.csv_file = tk.StringVar(value="yilideeplink.csv")
        self.output_dir = tk.StringVar(value="./output")
        self.schedule_enabled = tk.BooleanVar(value=False)
        self.schedule_time = tk.StringVar(value="09:00")
        self.is_running = False
        self.stop_flag = False
        
        self.create_modern_widgets()
        
    def create_modern_widgets(self):
        # ===== 顶部标题栏 =====
        header_frame = tk.Frame(self.root, bg=self.colors['primary_blue'], height=80)
        header_frame.pack(fill=tk.X, side=tk.TOP)
        header_frame.pack_propagate(False)
        
        # Logo和标题
        title_container = tk.Frame(header_frame, bg=self.colors['primary_blue'])
        title_container.pack(expand=True)
        
        title_label = tk.Label(
            title_container,
            text="WPP MD 小红书链接监测",
            font=('SF Pro Display', 24, 'bold'),
            bg=self.colors['primary_blue'],
            fg=self.colors['white']
        )
        title_label.pack(pady=5)
        
        subtitle_label = tk.Label(
            title_container,
            text="Professional Link Monitoring Tool",
            font=('SF Pro Display', 11),
            bg=self.colors['primary_blue'],
            fg=self.colors['light_blue']
        )
        subtitle_label.pack()
        
        # ===== 主内容区 =====
        main_container = tk.Frame(self.root, bg=self.colors['gray_bg'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # ===== 文件设置卡片 =====
        file_card = self.create_card(main_container, "📁 文件设置")
        file_card.pack(fill=tk.X, pady=(0, 15))
        
        # CSV文件
        csv_row = tk.Frame(file_card, bg=self.colors['white'])
        csv_row.pack(fill=tk.X, pady=8)
        
        tk.Label(
            csv_row,
            text="输入CSV文件",
            font=('SF Pro Text', 11),
            bg=self.colors['white'],
            fg=self.colors['text_dark']
        ).pack(side=tk.LEFT, padx=(0, 15))
        
        csv_entry = tk.Entry(
            csv_row,
            textvariable=self.csv_file,
            font=('SF Pro Text', 10),
            bg=self.colors['gray_bg'],
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            bd=0
        )
        csv_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, padx=5)
        
        csv_btn = tk.Button(
            csv_row,
            text="浏览",
            command=self.browse_csv,
            font=('SF Pro Text', 10, 'bold'),
            bg=self.colors['primary_blue'],
            fg=self.colors['white'],
            relief=tk.FLAT,
            bd=0,
            padx=20,
            pady=8,
            cursor='hand2'
        )
        csv_btn.pack(side=tk.LEFT, padx=(5, 0))
        
        # 输出目录
        output_row = tk.Frame(file_card, bg=self.colors['white'])
        output_row.pack(fill=tk.X, pady=8)
        
        tk.Label(
            output_row,
            text="输出目录    ",
            font=('SF Pro Text', 11),
            bg=self.colors['white'],
            fg=self.colors['text_dark']
        ).pack(side=tk.LEFT, padx=(0, 15))
        
        output_entry = tk.Entry(
            output_row,
            textvariable=self.output_dir,
            font=('SF Pro Text', 10),
            bg=self.colors['gray_bg'],
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            bd=0
        )
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, padx=5)
        
        output_btn = tk.Button(
            output_row,
            text="浏览",
            command=self.browse_output,
            font=('SF Pro Text', 10, 'bold'),
            bg=self.colors['primary_blue'],
            fg=self.colors['white'],
            relief=tk.FLAT,
            bd=0,
            padx=20,
            pady=8,
            cursor='hand2'
        )
        output_btn.pack(side=tk.LEFT, padx=(5, 0))
        
        # ===== 定时任务卡片 =====
        schedule_card = self.create_card(main_container, "⏰ 定时任务")
        schedule_card.pack(fill=tk.X, pady=(0, 15))
        
        schedule_row = tk.Frame(schedule_card, bg=self.colors['white'])
        schedule_row.pack(fill=tk.X, pady=8)
        
        check_style = ttk.Style()
        check_style.configure('Modern.TCheckbutton', 
                             background=self.colors['white'],
                             font=('SF Pro Text', 11))
        
        schedule_check = ttk.Checkbutton(
            schedule_row,
            text="启用定时运行",
            variable=self.schedule_enabled,
            command=self.toggle_schedule,
            style='Modern.TCheckbutton'
        )
        schedule_check.pack(side=tk.LEFT, padx=(0, 20))
        
        tk.Label(
            schedule_row,
            text="运行时间:",
            font=('SF Pro Text', 11),
            bg=self.colors['white'],
            fg=self.colors['text_gray']
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        self.time_entry = tk.Entry(
            schedule_row,
            textvariable=self.schedule_time,
            font=('SF Pro Text', 10),
            bg=self.colors['gray_bg'],
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            bd=0,
            width=10,
            state='disabled'
        )
        self.time_entry.pack(side=tk.LEFT, ipady=6, padx=5)
        
        tk.Label(
            schedule_row,
            text="(格式: HH:MM)",
            font=('SF Pro Text', 9),
            bg=self.colors['white'],
            fg=self.colors['text_gray']
        ).pack(side=tk.LEFT, padx=(5, 0))
        
        # ===== 控制按钮区 =====
        button_frame = tk.Frame(main_container, bg=self.colors['gray_bg'])
        button_frame.pack(pady=15)
        
        self.start_button = tk.Button(
            button_frame,
            text="▶ 立即开始监测",
            command=self.start_monitoring,
            font=('SF Pro Text', 13, 'bold'),
            bg=self.colors['primary_blue'],
            fg=self.colors['white'],
            relief=tk.FLAT,
            bd=0,
            padx=35,
            pady=15,
            cursor='hand2'
        )
        self.start_button.pack(side=tk.LEFT, padx=5)
        
        self.stop_button = tk.Button(
            button_frame,
            text="⬛ 停止",
            command=self.stop_monitoring,
            font=('SF Pro Text', 13, 'bold'),
            bg=self.colors['danger'],
            fg=self.colors['white'],
            relief=tk.FLAT,
            bd=0,
            padx=25,
            pady=15,
            state='disabled',
            cursor='hand2'
        )
        self.stop_button.pack(side=tk.LEFT, padx=5)
        
        clear_button = tk.Button(
            button_frame,
            text="🗑 清空日志",
            command=self.clear_log,
            font=('SF Pro Text', 13),
            bg=self.colors['gray_bg'],
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            bd=0,
            padx=25,
            pady=15,
            cursor='hand2'
        )
        clear_button.pack(side=tk.LEFT, padx=5)
        
        # ===== 进度卡片 =====
        progress_card = self.create_card(main_container, "📊 执行进度")
        progress_card.pack(fill=tk.X, pady=(0, 15))
        
        # 进度条
        style = ttk.Style()
        style.theme_use('clam')
        style.configure(
            "Blue.Horizontal.TProgressbar",
            troughcolor=self.colors['gray_bg'],
            bordercolor=self.colors['gray_bg'],
            background=self.colors['primary_blue'],
            lightcolor=self.colors['primary_blue'],
            darkcolor=self.colors['primary_blue']
        )
        
        self.progress = ttk.Progressbar(
            progress_card,
            mode='determinate',
            style="Blue.Horizontal.TProgressbar"
        )
        self.progress.pack(fill=tk.X, pady=(8, 5))
        
        self.progress_label = tk.Label(
            progress_card,
            text="就绪",
            font=('SF Pro Text', 10),
            bg=self.colors['white'],
            fg=self.colors['text_gray']
        )
        self.progress_label.pack(pady=(0, 8))
        
        # ===== 日志卡片 =====
        log_card = self.create_card(main_container, "📝 运行日志")
        log_card.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = scrolledtext.ScrolledText(
            log_card,
            height=12,
            font=('Consolas', 9),
            bg=self.colors['gray_bg'],
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            bd=0,
            wrap=tk.WORD,
            state='disabled'
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # ===== 底部状态栏 =====
        status_frame = tk.Frame(self.root, bg=self.colors['light_blue'], height=35)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        status_frame.pack_propagate(False)
        
        self.status_bar = tk.Label(
            status_frame,
            text="● 就绪",
            font=('SF Pro Text', 10),
            bg=self.colors['light_blue'],
            fg=self.colors['primary_blue'],
            anchor=tk.W
        )
        self.status_bar.pack(side=tk.LEFT, padx=20, fill=tk.Y)
        
    def create_card(self, parent, title):
        """创建卡片式容器"""
        card_frame = tk.Frame(
            parent,
            bg=self.colors['white'],
            relief=tk.FLAT,
            bd=0
        )
        
        # 标题
        title_label = tk.Label(
            card_frame,
            text=title,
            font=('SF Pro Display', 13, 'bold'),
            bg=self.colors['white'],
            fg=self.colors['text_dark'],
            anchor=tk.W
        )
        title_label.pack(fill=tk.X, padx=15, pady=(12, 8))
        
        # 分隔线
        separator = tk.Frame(card_frame, bg=self.colors['border'], height=1)
        separator.pack(fill=tk.X, padx=15)
        
        # 内容容器
        content_frame = tk.Frame(card_frame, bg=self.colors['white'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=(8, 12))
        
        return content_frame
    
    def browse_csv(self):
        filename = filedialog.askopenfilename(
            title="选择CSV文件",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.csv_file.set(filename)
            self.log(f"✓ 已选择CSV文件: {filename}")
    
    def browse_output(self):
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.output_dir.set(directory)
            self.log(f"✓ 已选择输出目录: {directory}")
    
    def toggle_schedule(self):
        if self.schedule_enabled.get():
            self.time_entry.config(state='normal', bg=self.colors['white'])
            self.log("⏰ 定时任务已启用")
        else:
            self.time_entry.config(state='disabled', bg=self.colors['gray_bg'])
            self.log("⏰ 定时任务已禁用")
    
    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {message}\n"
        
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, log_message)
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')
        
        self.status_bar.config(text=f"● {message}")
    
    def clear_log(self):
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        self.log("🗑 日志已清空")
    
    def update_progress(self, current, total, message=""):
        progress = (current / total) * 100 if total > 0 else 0
        self.progress['value'] = progress
        self.progress_label.config(text=f"{message} ({current}/{total}) • {progress:.0f}%")
        self.root.update_idletasks()
    
    def read_links_from_csv(self, csv_file):
        links = []
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
            for row in rows[1:]:
                if row and len(row) > 0 and row[0].strip():
                    links.append(row[0].strip())
        return links
    
    def resize_screenshot(self, image_path, max_width=300, max_height=200):
        try:
            img = PILImage.open(image_path)
            width_ratio = max_width / img.width
            height_ratio = max_height / img.height
            ratio = min(width_ratio, height_ratio)
            return int(img.width * ratio), int(img.height * ratio)
        except:
            return max_width, max_height
    
    def create_excel_report(self, results, output_dir):
        os.makedirs(output_dir, exist_ok=True)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "链接监测结果"
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 20
        
        headers = ['序号', '链接', '状态', '截屏预览', '检测时间']
        header_fill = PatternFill(start_color="007AFF", end_color="007AFF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 25
        
        current_row = 2
        for idx, result in enumerate(results, 1):
            ws.cell(row=current_row, column=1, value=idx).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=1).border = thin_border
            
            ws.cell(row=current_row, column=2, value=result['链接']).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.cell(row=current_row, column=2).border = thin_border
            
            status_cell = ws.cell(row=current_row, column=3, value=result['状态'])
            if result['状态'] == '成功':
                status_cell.fill = PatternFill(start_color="34C759", end_color="34C759", fill_type="solid")
                status_cell.font = Font(color="FFFFFF", bold=True)
            else:
                status_cell.fill = PatternFill(start_color="FF3B30", end_color="FF3B30", fill_type="solid")
                status_cell.font = Font(color="FFFFFF", bold=True)
            status_cell.alignment = Alignment(horizontal='center', vertical='center')
            status_cell.border = thin_border
            
            screenshot_path = result.get('截屏文件', '')
            cell_d = ws.cell(row=current_row, column=4)
            cell_d.border = thin_border
            
            if screenshot_path and os.path.exists(screenshot_path):
                try:
                    width, height = self.resize_screenshot(screenshot_path, max_width=300, max_height=180)
                    ws.row_dimensions[current_row].height = height * 0.75
                    img = XLImage(screenshot_path)
                    img.width = width
                    img.height = height
                    img.anchor = f'D{current_row}'
                    ws.add_image(img)
                except:
                    cell_d.value = '图片加载失败'
                    ws.row_dimensions[current_row].height = 30
            else:
                cell_d.value = result.get('错误信息', '无截图')
                cell_d.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                ws.row_dimensions[current_row].height = 30
            
            cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            ws.cell(row=current_row, column=5, value=result['检测时间']).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=5).border = thin_border
            
            current_row += 1
        
        stats_row = current_row + 1
        success_count = sum(1 for r in results if r['状态'] == '成功')
        fail_count = sum(1 for r in results if r['状态'] == '失败')
        
        ws.cell(row=stats_row, column=1, value='统计').font = Font(bold=True, size=12)
        ws.cell(row=stats_row, column=2, value=f'总计: {len(results)} | 成功: {success_count} | 失败: {fail_count}')
        ws.cell(row=stats_row, column=2).font = Font(bold=True, size=11)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_dir, f'链接监测报告_{timestamp}.xlsx')
        wb.save(output_file)
        
        return output_file, success_count, fail_count
    
    def capture_with_playwright(self, links, screenshot_dir):
        results = []
        
        self.log("🌐 正在启动浏览器...")
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(viewport={'width': 1920, 'height': 1080})
            page = context.new_page()
            
            self.log("🏠 正在打开小红书主页...")
            page.goto('https://www.xiaohongshu.com/')
            self.log("⏳ 等待10秒，请完成登录（如已登录可忽略）...")
            time.sleep(10)
            
            for idx, link in enumerate(links, 1):
                if self.stop_flag:
                    self.log("⚠ 任务已被用户停止")
                    break
                
                self.update_progress(idx, len(links), f"正在访问链接 {idx}")
                self.log(f"[{idx}/{len(links)}] 访问: {link[:60]}...")
                
                result = {
                    '链接': link,
                    '状态': '失败',
                    '截屏文件': '',
                    '错误信息': '',
                    '检测时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                try:
                    page.goto(link, timeout=30000)
                    time.sleep(3)
                    
                    screenshot_name = f'screenshot_{idx}.png'
                    screenshot_path = os.path.join(screenshot_dir, screenshot_name)
                    page.screenshot(path=screenshot_path)
                    
                    result['状态'] = '成功'
                    result['截屏文件'] = screenshot_path
                    self.log(f"  ✓ 截图已保存: {screenshot_name}")
                    
                except Exception as e:
                    result['错误信息'] = f'访问失败: {str(e)}'
                    self.log(f"  ✗ 错误: {str(e)}")
                
                results.append(result)
            
            browser.close()
            self.log("🔒 浏览器已关闭")
        
        return results
    
    def monitor_task(self):
        try:
            self.is_running = True
            self.stop_flag = False
            self.start_button.config(state='disabled', bg=self.colors['text_gray'])
            self.stop_button.config(state='normal', bg=self.colors['danger'])
            
            csv_file = self.csv_file.get()
            if not os.path.exists(csv_file):
                messagebox.showerror("错误", f"CSV文件不存在: {csv_file}")
                return
            
            output_dir = self.output_dir.get()
            screenshot_dir = os.path.join(output_dir, 'screenshots')
            os.makedirs(screenshot_dir, exist_ok=True)
            
            self.log("="*50)
            self.log("🚀 开始链接监测任务")
            self.log(f"📁 输入: {csv_file}")
            self.log(f"📁 输出: {output_dir}")
            self.log("="*50)
            
            links = self.read_links_from_csv(csv_file)
            self.log(f"📋 找到 {len(links)} 个链接")
            
            if not links:
                messagebox.showwarning("警告", "CSV文件中没有找到链接")
                return
            
            results = self.capture_with_playwright(links, screenshot_dir)
            
            if not self.stop_flag:
                self.log("📊 正在生成Excel报告...")
                output_file, success_count, fail_count = self.create_excel_report(results, output_dir)
                
                self.log("="*50)
                self.log("✅ 监测任务完成！")
                self.log(f"📊 总计: {len(results)} | 成功: {success_count} | 失败: {fail_count}")
                self.log(f"📄 报告: {output_file}")
                self.log("="*50)
                
                messagebox.showinfo(
                    "完成",
                    f"监测完成！\n\n"
                    f"总计: {len(results)}\n"
                    f"成功: {success_count}\n"
                    f"失败: {fail_count}\n\n"
                    f"报告已保存至:\n{output_file}"
                )
        
        except Exception as e:
            self.log(f"❌ 错误: {str(e)}")
            messagebox.showerror("错误", f"监测过程中发生错误:\n{str(e)}")
        
        finally:
            self.is_running = False
            self.start_button.config(state='normal', bg=self.colors['primary_blue'])
            self.stop_button.config(state='disabled', bg=self.colors['text_gray'])
            self.update_progress(0, 1, "就绪")
    
    def start_monitoring(self):
        if not self.is_running:
            thread = threading.Thread(target=self.monitor_task, daemon=True)
            thread.start()
    
    def stop_monitoring(self):
        if self.is_running:
            self.stop_flag = True
            self.log("⚠ 正在停止任务...")

def main():
    root = tk.Tk()
    
    # 设置窗口图标和样式
    try:
        root.iconbitmap(default='')
    except:
        pass
    
    app = ModernLinkMonitorGUI(root)
    
    # 窗口居中
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    root.mainloop()

if __name__ == '__main__':
    main()








