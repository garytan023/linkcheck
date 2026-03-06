import csv
import os
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import time

def read_links_from_csv(csv_file):
    """从CSV文件读取链接"""
    links = []
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
        # 跳过标题行
        for row in rows[1:]:
            if row and len(row) > 0 and row[0].strip():
                links.append(row[0].strip())
    return links

def create_excel_report(results, output_file='链接监测报告.xlsx'):
    """创建Excel报告"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "链接监测结果"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 20
    
    # 创建标题行
    headers = ['序号', '链接', '状态', '截屏', '检测时间']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 填充数据
    current_row = 2
    for idx, result in enumerate(results, 1):
        ws.cell(row=current_row, column=1, value=idx)
        ws.cell(row=current_row, column=2, value=result['链接'])
        
        # 状态列
        status_cell = ws.cell(row=current_row, column=3, value=result['状态'])
        if result['状态'] == '成功':
            status_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            status_cell.font = Font(color="FFFFFF", bold=True)
        else:
            status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            status_cell.font = Font(color="FFFFFF", bold=True)
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 截屏列
        screenshot_path = result.get('截屏文件', '')
        if screenshot_path and os.path.exists(screenshot_path):
            ws.cell(row=current_row, column=4, value='查看截图')
            # 设置行高以容纳图片（约150像素高度）
            ws.row_dimensions[current_row].height = 150
            
            # 插入图片
            try:
                img = XLImage(screenshot_path)
                # 调整图片大小
                img.width = 200
                img.height = 140
                # 图片定位到D列
                img.anchor = f'D{current_row}'
                ws.add_image(img)
            except Exception as e:
                ws.cell(row=current_row, column=4, value=f'图片加载失败: {str(e)}')
        else:
            ws.cell(row=current_row, column=4, value=result.get('错误信息', '无截图'))
        
        # 检测时间
        ws.cell(row=current_row, column=5, value=result['检测时间'])
        
        current_row += 1
    
    # 保存Excel文件
    wb.save(output_file)
    print(f"\n✓ Excel报告已生成: {output_file}")
    return output_file

def print_instructions():
    """打印使用说明"""
    print("\n" + "="*60)
    print("链接监测工具 - 使用说明")
    print("="*60)
    print("\n步骤:")
    print("1. 浏览器将自动打开小红书登录页面")
    print("2. 请使用小红书APP扫描二维码登录")
    print("3. 登录成功后，按任意键继续...")
    print("4. 工具将自动访问每个链接并截屏")
    print("5. 完成后将生成Excel报告")
    print("\n" + "="*60 + "\n")

def main():
    csv_file = 'yilideeplink.csv'
    
    # 打印说明
    print_instructions()
    
    # 读取链接
    print(f"正在读取 {csv_file}...")
    links = read_links_from_csv(csv_file)
    print(f"找到 {len(links)} 个链接\n")
    
    if not links:
        print("错误: 未找到任何链接")
        return
    
    # 创建截图目录
    screenshot_dir = 'screenshots'
    if not os.path.exists(screenshot_dir):
        os.makedirs(screenshot_dir)
    
    print("\n请在浏览器中完成扫码登录...")
    print("登录页面截屏已保存为: screenshots/login_page.png")
    print("\n完成登录后，请按Enter键继续...")
    
    # 这里需要等待用户手动确认已登录
    # 实际的浏览器操作将通过MCP工具完成
    
    results = []
    
    # 模拟结果（实际将通过浏览器工具获取）
    print("\n开始检测链接...")
    for idx, link in enumerate(links, 1):
        print(f"\n[{idx}/{len(links)}] 正在访问: {link}")
        
        result = {
            '链接': link,
            '状态': '待检测',
            '截屏文件': '',
            '错误信息': '',
            '检测时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # 这里将被实际的浏览器操作替换
        screenshot_name = f'screenshot_{idx}.png'
        screenshot_path = os.path.join(screenshot_dir, screenshot_name)
        
        # 检查截屏是否存在
        if os.path.exists(screenshot_path):
            result['状态'] = '成功'
            result['截屏文件'] = screenshot_path
            print(f"  ✓ 成功 - 截屏已保存")
        else:
            result['状态'] = '失败'
            result['错误信息'] = '截屏文件不存在'
            print(f"  ✗ 失败 - 截屏文件不存在")
        
        results.append(result)
    
    # 生成Excel报告
    print("\n正在生成Excel报告...")
    output_file = create_excel_report(results)
    
    # 显示总结
    print("\n" + "="*60)
    print("检测完成！")
    print("="*60)
    success_count = sum(1 for r in results if r['状态'] == '成功')
    fail_count = len(results) - success_count
    print(f"总计: {len(results)} 个链接")
    print(f"成功: {success_count} 个")
    print(f"失败: {fail_count} 个")
    
    if fail_count > 0:
        print("\n失败的链接:")
        for idx, result in enumerate(results, 1):
            if result['状态'] != '成功':
                print(f"  {idx}. {result['链接']}")
                print(f"     错误: {result['错误信息']}")
    
    print("\n" + "="*60)

if __name__ == '__main__':
    main()


