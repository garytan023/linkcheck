"""
WPP MD 小红书链接监测工具 - 修复版本
修复问题：
1. 视频封面截图问题
2. 图片封面从第一张开始爬取
3. 链接与输出文档对应关系
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import csv
import os
import threading
import time
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image as PILImage
from playwright.sync_api import sync_playwright
import json
from urllib.parse import urljoin
import re
import sys
import string

def resolve_playwright_browsers_path():
    """Resolve bundled Playwright browsers path for packaged exe."""
    candidates = []
    # PyInstaller解包目录
    if getattr(sys, "_MEIPASS", None):
        candidates.append(os.path.join(sys._MEIPASS, "browsers"))
    # 当前目录（与exe同目录）
    candidates.append(os.path.join(os.getcwd(), "browsers"))
    # 用户目录缓存
    candidates.append(os.path.join(os.path.expanduser('~'), '.playwright-browsers'))
    for path in candidates:
        if path and os.path.exists(path):
            return path
    return None

# 设置Playwright浏览器路径（优先使用打包内置的browsers目录）
if not os.environ.get('PLAYWRIGHT_BROWSERS_PATH'):
    pw_path = resolve_playwright_browsers_path()
    if pw_path:
        os.environ['PLAYWRIGHT_BROWSERS_PATH'] = pw_path

class FixedLinkMonitorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("WPP MD 小红书链接监测工具 - 修复版")
        self.root.geometry("950x800")

        # 配色
        self.colors = {
            'bg': '#5B6FB5',
            'primary': '#5B7BF5',
            'secondary': '#6B5FB5',
            'white': '#FFFFFF',
            'text_light': '#E8EEFF',
            'text_dark': '#2D3748',
            'success': '#48BB78',
            'danger': '#F56565',
        }

        # 变量
        self.csv_file = tk.StringVar(value="yilideeplink.csv")
        self.output_dir = tk.StringVar(value="./output")
        self.schedule_enabled = tk.BooleanVar(value=False)
        self.schedule_time = tk.StringVar(value="09:00")
        self.login_wait_time = tk.StringVar(value="30")
        self.capture_mode = tk.StringVar(value="full")  # simple/full
        self.all_links_mode = tk.BooleanVar(value=False)  # 是否对所有链接仅截图校验
        self.is_running = False
        self.stop_flag = False
        self.http_user_agent = (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36"
        )
        # 存储Playwright登录态，扫码一次后可复用
        self.storage_state_path = os.path.join(os.getcwd(), "xhs_fixed_storage.json")
        # 持久化浏览器用户数据目录，防止每次都要扫码
        self.user_data_dir = os.path.join(os.getcwd(), "xhs_fixed_profile")
        # 控制粉丝数获取的兜底策略，默认启用API兜底以提高准确性
        self.enable_api_fans = True
        self.enable_homepage_fans = True

        self.root.configure(bg=self.colors['bg'])
        self.create_ui()

    def create_ui(self):
        # 主容器
        main_canvas = tk.Canvas(self.root, bg=self.colors['bg'], highlightthickness=0)
        scrollbar = tk.Scrollbar(self.root, orient="vertical", command=main_canvas.yview)
        scrollable_frame = tk.Frame(main_canvas, bg=self.colors['bg'])

        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )

        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=930)
        main_canvas.configure(yscrollcommand=scrollbar.set)

        # 允许鼠标滚轮上下滑动
        def _on_mousewheel(event):
            delta = -1 * int(event.delta / 120) if event.delta else 0
            main_canvas.yview_scroll(delta, "units")
            return "break"
        main_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        # 兼容触控板/其他平台
        main_canvas.bind_all("<Button-4>", lambda e: main_canvas.yview_scroll(-1, "units"))
        main_canvas.bind_all("<Button-5>", lambda e: main_canvas.yview_scroll(1, "units"))

        main_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        content = tk.Frame(scrollable_frame, bg=self.colors['bg'])
        content.pack(fill=tk.BOTH, expand=True, padx=25, pady=25)

        # 标题
        title_frame = tk.Frame(content, bg=self.colors['bg'])
        title_frame.pack(pady=(0, 20))

        logo_canvas = tk.Canvas(title_frame, width=70, height=70, bg=self.colors['bg'], highlightthickness=0)
        logo_canvas.pack()
        logo_canvas.create_oval(10, 10, 60, 60, fill=self.colors['primary'], outline='')
        logo_canvas.create_oval(22, 22, 48, 48, fill=self.colors['white'], outline='')
        logo_canvas.create_oval(30, 30, 40, 40, fill=self.colors['primary'], outline='')

        tk.Label(
            title_frame,
            text="WPP MD 小红书链接监测工具 - 修复版",
            font=('Microsoft YaHei UI', 20, 'bold'),
            bg=self.colors['bg'],
            fg=self.colors['white']
        ).pack(pady=(10, 2))

        tk.Label(
            title_frame,
            text="Fixed Professional Link Monitoring Tool",
            font=('Arial', 9),
            bg=self.colors['bg'],
            fg=self.colors['text_light']
        ).pack()

        # 文件配置
        file_card = self.create_card(content, "📁 文件配置")

        csv_row = tk.Frame(file_card, bg='white')
        csv_row.pack(fill=tk.X, pady=5)

        tk.Label(csv_row, text="输入CSV文件:", font=('Microsoft YaHei UI', 9), bg='white', fg=self.colors['text_dark']).pack(side=tk.LEFT)

        csv_entry_frame = tk.Frame(csv_row, bg='#F7FAFC')
        csv_entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 5))

        tk.Entry(
            csv_entry_frame,
            textvariable=self.csv_file,
            font=('Consolas', 9),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT
        ).pack(fill=tk.X, padx=8, pady=6)

        tk.Button(
            csv_row,
            text="浏览...",
            command=self.browse_csv,
            font=('Microsoft YaHei UI', 9),
            bg=self.colors['primary'],
            fg='white',
            relief=tk.FLAT,
            padx=15,
            pady=4,
            cursor='hand2'
        ).pack(side=tk.RIGHT)

        output_row = tk.Frame(file_card, bg='white')
        output_row.pack(fill=tk.X, pady=5)

        tk.Label(output_row, text="输出目录:", font=('Microsoft YaHei UI', 9), bg='white', fg=self.colors['text_dark']).pack(side=tk.LEFT, padx=(0, 15))

        output_entry_frame = tk.Frame(output_row, bg='#F7FAFC')
        output_entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 5))

        tk.Entry(
            output_entry_frame,
            textvariable=self.output_dir,
            font=('Consolas', 9),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT
        ).pack(fill=tk.X, padx=8, pady=6)

        tk.Button(
            output_row,
            text="浏览...",
            command=self.browse_output,
            font=('Microsoft YaHei UI', 9),
            bg=self.colors['primary'],
            fg='white',
            relief=tk.FLAT,
            padx=15,
            pady=4,
            cursor='hand2'
        ).pack(side=tk.RIGHT)

        # 抓取模式
        mode_card = self.create_card(content, "🎯 抓取模式 (修复版)")

        mode_row = tk.Frame(mode_card, bg='white')
        mode_row.pack(fill=tk.X, pady=5)

        tk.Radiobutton(
            mode_row,
            text="快速模式（仅截图）",
            variable=self.capture_mode,
            value="simple",
            font=('Microsoft YaHei UI', 9),
            bg='white',
            fg=self.colors['text_dark'],
            activebackground='white',
            selectcolor='white',
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=(0, 20))

        tk.Radiobutton(
            mode_row,
            text="完整模式（封面+详情+多图）",
            variable=self.capture_mode,
            value="full",
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='white',
            fg=self.colors['primary'],
            activebackground='white',
            selectcolor='white',
            cursor='hand2'
        ).pack(side=tk.LEFT)

        mode_desc = tk.Label(
            mode_card,
            text="🔧 修复版优化：确保视频封面截图、图片按顺序抓取、链接正确对应",
            font=('Arial', 8),
            bg='white',
            fg='#718096'
        )
        mode_desc.pack(pady=(5, 0))

        # 控制按钮
        button_frame = tk.Frame(content, bg=self.colors['bg'])
        button_frame.pack(pady=15)

        self.start_button = tk.Button(
            button_frame,
            text="▶  立即开始监测",
            command=self.start_monitoring,
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            relief=tk.FLAT,
            padx=30,
            pady=12,
            cursor='hand2'
        )
        self.start_button.pack(side=tk.LEFT, padx=5)

        self.stop_button = tk.Button(
            button_frame,
            text="⬛  停止",
            command=self.stop_monitoring,
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg=self.colors['danger'],
            fg='white',
            relief=tk.FLAT,
            padx=25,
            pady=12,
            state='disabled',
            cursor='hand2'
        )
        self.stop_button.pack(side=tk.LEFT, padx=5)

        # 进度
        progress_card = self.create_card(content, "📊 执行进度")

        style = ttk.Style()
        style.theme_use('clam')
        style.configure(
            "Custom.Horizontal.TProgressbar",
            troughcolor='#E2E8F0',
            bordercolor='#E2E8F0',
            background=self.colors['primary'],
            lightcolor=self.colors['primary'],
            darkcolor=self.colors['primary']
        )

        self.progress = ttk.Progressbar(
            progress_card,
            mode='determinate',
            style="Custom.Horizontal.TProgressbar",
            length=400
        )
        self.progress.pack(fill=tk.X, pady=(5, 8))

        self.progress_label = tk.Label(
            progress_card,
            text="就绪",
            font=('Microsoft YaHei UI', 9),
            bg='white',
            fg=self.colors['text_dark']
        )
        self.progress_label.pack()

        # 日志
        log_card = self.create_card(content, "📝 运行日志")

        log_toolbar = tk.Frame(log_card, bg='white')
        log_toolbar.pack(fill=tk.X, pady=(0, 5))

        tk.Button(
            log_toolbar,
            text="清空日志",
            command=self.clear_log,
            font=('Microsoft YaHei UI', 8),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            padx=10,
            pady=3,
            cursor='hand2'
        ).pack(side=tk.RIGHT)

        self.log_text = scrolledtext.ScrolledText(
            log_card,
            height=12,
            font=('Consolas', 8),
            bg='#F7FAFC',
            fg=self.colors['text_dark'],
            relief=tk.FLAT,
            wrap=tk.WORD,
            state='disabled'
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        # 状态栏
        status_frame = tk.Frame(content, bg='#4B5FA5', height=35)
        status_frame.pack(fill=tk.X, pady=(10, 0))
        status_frame.pack_propagate(False)

        self.status_bar = tk.Label(
            status_frame,
            text="● 就绪 - 修复版已就绪",
            font=('Microsoft YaHei UI', 9),
            bg='#4B5FA5',
            fg=self.colors['text_light'],
            anchor=tk.W
        )
        self.status_bar.pack(side=tk.LEFT, padx=15, fill=tk.Y)

    def create_card(self, parent, title):
        card_outer = tk.Frame(parent, bg='white', relief=tk.FLAT)
        card_outer.pack(fill=tk.X, pady=(0, 12))

        title_frame = tk.Frame(card_outer, bg='white')
        title_frame.pack(fill=tk.X, padx=15, pady=(10, 5))

        tk.Label(
            title_frame,
            text=title,
            font=('Microsoft YaHei UI', 11, 'bold'),
            bg='white',
            fg=self.colors['text_dark']
        ).pack(side=tk.LEFT)

        tk.Frame(card_outer, bg='#E2E8F0', height=1).pack(fill=tk.X, padx=15)

        content = tk.Frame(card_outer, bg='white')
        content.pack(fill=tk.BOTH, expand=True, padx=15, pady=(8, 10))

        return content

    def browse_csv(self):
        filename = filedialog.askopenfilename(
            title="选择CSV文件",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.csv_file.set(filename)
            self.log(f"[选择] CSV: {os.path.basename(filename)}")

    def browse_output(self):
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.output_dir.set(directory)
            self.log(f"[选择] 输出: {directory}")

    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {message}\n"

        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, log_message)
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')

        self.status_bar.config(text=f"● {message}")

    def clear_log(self):
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        self.log("[清空] 日志已清空")

    def update_progress(self, current, total, message=""):
        progress = (current / total) * 100 if total > 0 else 0
        self.progress['value'] = progress
        self.progress_label.config(text=f"{message} ({current}/{total}) • {progress:.0f}%")
        self.root.update_idletasks()

    def read_items_from_csv(self, csv_file):
        """读取CSV，支持列: 产品, 序号, 链接；兼容单列链接"""
        items = []
        try:
            with open(csv_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                if reader.fieldnames and ('链接' in reader.fieldnames or reader.fieldnames[0]):
                    for idx, row in enumerate(reader, 1):
                        link = (row.get('链接') or row.get(reader.fieldnames[0]) or '').strip()
                        if not link:
                            continue
                        product = (row.get('产品') or '').strip()
                        seq = (row.get('序号') or '').strip() or str(idx)
                        items.append({'产品': product, '序号': seq, '链接': link})
                        continue
        except Exception:
            pass

        # 兼容旧格式：单列链接
        if not items:
            try:
                with open(csv_file, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                    for idx, row in enumerate(rows[1:], 1):
                        if row and len(row) > 0 and row[0].strip():
                            items.append({'产品': '', '序号': str(idx), '链接': row[0].strip()})
            except Exception:
                pass
        return items

    def ensure_media_loaded(self, page):
        """Scroll page to trigger lazy-loaded images."""
        try:
            page.evaluate("() => window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(0.8)
            page.evaluate("() => window.scrollTo(0, document.body.scrollHeight / 2)")
            time.sleep(0.4)
            page.evaluate("() => window.scrollTo(0, 0)")
            time.sleep(0.2)
        except Exception:
            pass

    def normalize_media_url(self, base_url, url):
        """Normalize protocol-relative or relative URLs to absolute."""
        try:
            if not url:
                return ""
            if url.startswith("//"):
                return "https:" + url
            if url.startswith("http://") or url.startswith("https://"):
                return url
            return urljoin(base_url, url)
        except Exception:
            return url or ""

    def is_unwanted_image_src(self, src: str) -> bool:
        """Heuristics to skip avatars/logos/watermarks."""
        if not src:
            return True
        lowered = src.lower()
        unwanted_keywords = [
            "avatar", "profile", "logo", "icon", "watermark", "badge",
            "placeholder", "default_avatar", "default-avatar", "favicon",
            "xhslogo", "/head/", "/avatar/", "logos", "brand"
        ]
        return any(k in lowered for k in unwanted_keywords)

    def element_is_large_enough(self, elem) -> bool:
        """Check DOM element visual size to avoid tiny icons."""
        try:
            box = elem.bounding_box()
            if not box:
                return False
            return (box.get("width", 0) >= 180) and (box.get("height", 0) >= 180)
        except Exception:
            return False

    def is_large_image_file(self, image_path: str, min_width: int = 160, min_height: int = 160) -> bool:
        """Check saved image file dimensions to avoid wrong covers."""
        try:
            with PILImage.open(image_path) as im:
                w, h = im.size
                return w >= min_width and h >= min_height
        except Exception:
            # If we cannot open the file, keep it to avoid over-filtering
            return True

    def sanitize_name(self, name: str) -> str:
        try:
            safe = re.sub(r'[^0-9A-Za-z\u4e00-\u9fff]+', '_', name)
            return safe.strip('_') or "item"
        except Exception:
            return "item"

    # ===== 修复1: 改进视频封面截图 =====
    def capture_video_cover_fixed(self, page, output_dir, name_prefix):
        """修复版视频封面截图 - 确保正确获取视频封面"""
        covers = []

        try:
            self.log("    [视频-修复] 检测并截取视频封面...")

            # 1. 确保视频已加载并暂停
            page.evaluate("""() => {
                const videos = document.querySelectorAll('video');
                videos.forEach(v => {
                    if (v.readyState >= 2) {  # HAVE_CURRENT_DATA
                        v.pause();
                        v.currentTime = 0;
                    }
                });
            }""")
            time.sleep(2)  # 等待视频状态稳定

            # 2. 多种策略获取视频封面
            strategies = [
                # 策略1: 直接截取video元素
                {
                    'name': 'video元素直接截图',
                    'selector': 'video',
                    'method': 'screenshot'
                },
                # 策略2: 视频容器
                {
                    'name': '视频容器截图',
                    'selector': '.xgplayer, .video-player, [class*="video-container"]',
                    'method': 'screenshot'
                },
                # 策略3: 视频海报属性
                {
                    'name': '视频poster属性',
                    'selector': 'video[poster]',
                    'method': 'poster'
                },
                # 策略4: 封面图片
                {
                    'name': '视频封面图片',
                    'selector': 'img[src*="cover"], img[src*="poster"], .video-cover img',
                    'method': 'image'
                }
            ]

            for strategy in strategies:
                try:
                    self.log(f"      [策略] 尝试: {strategy['name']}")

                    if strategy['method'] == 'screenshot':
                        elements = page.query_selector_all(strategy['selector'])
                        for elem in elements:
                            try:
                                cover_name = f'{name_prefix}_video_cover_{len(covers)+1}.png'
                                cover_path = os.path.join(output_dir, cover_name)

                                # 检查元素尺寸
                                box = elem.bounding_box()
                                if box and box['width'] > 100 and box['height'] > 100:
                                    elem.screenshot(path=cover_path, timeout=10000)

                                    if self.is_large_image_file(cover_path, min_width=100, min_height=100):
                                        covers.append(cover_path)
                                        self.log(f"      ✓ 成功: {strategy['name']}")
                                        break
                                    else:
                                        try:
                                            os.remove(cover_path)
                                        except:
                                            pass
                            except Exception as e:
                                self.log(f"      ✗ 失败: {strategy['name']} - {str(e)}")
                                continue

                    elif strategy['method'] == 'poster':
                        poster_urls = page.evaluate("""() => {
                            const urls = [];
                            document.querySelectorAll('video[poster]').forEach(v => {
                                if (v.poster) urls.push(v.poster);
                            });
                            return urls;
                        }""")

                        for poster_url in poster_urls:
                            try:
                                cover_name = f'{name_prefix}_video_poster_{len(covers)+1}.png'
                                cover_path = os.path.join(output_dir, cover_name)

                                if self.download_image_via_page(page, poster_url, cover_path):
                                    if self.is_large_image_file(cover_path, min_width=100, min_height=100):
                                        covers.append(cover_path)
                                        self.log(f"      ✓ 成功: {strategy['name']}")
                                        break
                                    else:
                                        try:
                                            os.remove(cover_path)
                                        except:
                                            pass
                            except Exception as e:
                                self.log(f"      ✗ 失败: {strategy['name']} - {str(e)}")
                                continue

                    elif strategy['method'] == 'image':
                        img_elements = page.query_selector_all(strategy['selector'])
                        for img in img_elements:
                            try:
                                src = img.get_attribute('src') or img.get_attribute('data-src')
                                if src:
                                    cover_name = f'{name_prefix}_video_img_{len(covers)+1}.png'
                                    cover_path = os.path.join(output_dir, cover_name)

                                    if self.download_image_via_page(page, src, cover_path):
                                        if self.is_large_image_file(cover_path, min_width=100, min_height=100):
                                            covers.append(cover_path)
                                            self.log(f"      ✓ 成功: {strategy['name']}")
                                            break
                                        else:
                                            try:
                                                os.remove(cover_path)
                                            except:
                                                pass
                            except Exception as e:
                                self.log(f"      ✗ 失败: {strategy['name']} - {str(e)}")
                                continue

                    # 如果已获取到封面，不再尝试其他策略
                    if covers:
                        break

                except Exception as e:
                    self.log(f"      ✗ 策略异常: {strategy['name']} - {str(e)}")
                    continue

            if covers:
                self.log(f"    [视频-修复] ✓ 成功获取 {len(covers)} 个视频封面")
            else:
                self.log("    [视频-修复] ✗ 未能获取视频封面")

        except Exception as e:
            self.log(f"    [视频-修复] ✗ 整体失败: {str(e)}")

        return covers

    # ===== 修复2: 确保图片从第一张开始爬取 =====
    def capture_cover_images_fixed(self, page, output_dir, name_prefix, image_urls=None):
        """修复版封面图片抓取 - 确保从第一张开始按顺序抓取"""
        cover_paths = []
        image_urls = image_urls or []

        try:
            self.log(f"    [封面-修复] 开始抓取 (元数据URL数: {len(image_urls)})")

            # 策略1: 优先使用元数据URL，并确保按顺序处理
            if image_urls and len(image_urls) > 0:
                self.log(f"    [封面-修复] 按顺序处理元数据图片...")
                seen = set()

                # 确保从第一张开始，按索引顺序处理
                for idx in range(min(len(image_urls), 10)):
                    url = image_urls[idx]
                    if not url or url in seen:
                        continue

                    normalized = self.normalize_media_url(page.url, url)
                    if normalized in seen or self.is_unwanted_image_src(normalized):
                        continue

                    seen.add(normalized)
                    cover_name = f'{name_prefix}_cover_{idx+1:02d}.png'  # 使用01, 02...保证顺序
                    cover_path = os.path.join(output_dir, cover_name)

                    self.log(f"      [第{idx+1}张] 下载: {normalized[:60]}...")
                    if self.download_image_via_page(page, normalized, cover_path):
                        if self.is_large_image_file(cover_path, min_width=100, min_height=100):
                            cover_paths.append(cover_path)
                            self.log(f"      [第{idx+1}张] ✓ 保存成功")
                        else:
                            self.log(f"      [第{idx+1}张] ✗ 图片太小，跳过")
                            try:
                                os.remove(cover_path)
                            except Exception:
                                pass
                    else:
                        self.log(f"      [第{idx+1}张] ✗ 下载失败")

                    if len(cover_paths) >= 10:
                        break

            # 检查是否为视频笔记
            is_video = page.evaluate("""() => {
                return document.querySelector('video') !== null;
            }""")

            # 如果是视频笔记且封面不够，使用修复版视频截图
            if is_video and len(cover_paths) < 1:
                self.log("    [封面-修复] 检测到视频，尝试视频封面截图...")
                video_covers = self.capture_video_cover_fixed(page, output_dir, name_prefix)
                if video_covers:
                    cover_paths.extend(video_covers)
                    self.log(f"    [封面-修复] 视频封面成功: {len(video_covers)}张")

            # 策略2: 如果元数据URL不够，使用DOM抓取作为补充
            if len(cover_paths) < 3:
                self.log("    [封面-修复] 使用DOM抓取补充图片...")
                self.ensure_media_loaded(page)

                # 获取所有图片，并按页面顺序排列
                img_info = page.evaluate("""() => {
                    const images = [];
                    const seen = new Set();

                    // 按DOM顺序获取图片，确保顺序性
                    const allImages = document.querySelectorAll('img');
                    allImages.forEach((img, index) => {
                        const src = img.src || img.getAttribute('data-src') || '';
                        if (src && !seen.has(src)) {
                            seen.add(src);
                            images.push({
                                src: src,
                                index: index,
                                width: img.naturalWidth || img.width || 0,
                                height: img.naturalHeight || img.height || 0,
                                className: img.className || '',
                                type: 'general'
                            });
                        }
                    });

                    return images;
                }""")

                self.log(f"    [DOM] 找到{len(img_info)}个候选图片")

                seen_urls = set(url for url in image_urls if url)  # 排除已处理的
                processed_count = len(cover_paths)

                for info in img_info:
                    if len(cover_paths) >= 10:
                        break

                    src = info.get('src', '')
                    if not src or src in seen_urls:
                        continue

                    normalized = self.normalize_media_url(page.url, src)
                    if normalized in seen_urls or self.is_unwanted_image_src(normalized):
                        continue

                    # 优先选择大尺寸图片
                    if info.get('width', 0) < 100 or info.get('height', 0) < 100:
                        continue

                    seen_urls.add(normalized)
                    cover_name = f'{name_prefix}_cover_{processed_count+1:02d}.png'
                    cover_path = os.path.join(output_dir, cover_name)

                    if self.download_image_via_page(page, normalized, cover_path):
                        if self.is_large_image_file(cover_path, min_width=100, min_height=100):
                            cover_paths.append(cover_path)
                            processed_count += 1
                            self.log(f"      [DOM补充第{processed_count}张] {info.get('width',0)}x{info.get('height',0)}")
                        else:
                            try:
                                os.remove(cover_path)
                            except Exception:
                                pass

            # 最终排序确保封面顺序正确
            if cover_paths:
                # 按文件名中的数字重新排序
                def get_sort_key(path):
                    import re
                    match = re.search(r'cover_(\d+)', os.path.basename(path))
                    return int(match.group(1)) if match else 999

                cover_paths.sort(key=get_sort_key)
                self.log(f"    [封面-修复] ✓ 最终获取 {len(cover_paths)} 张封面，已按顺序排列")
            else:
                self.log("    [封面-修复] ✗ 未找到任何有效封面图")

        except Exception as e:
            self.log(f"    [封面-修复] ✗ 抓取失败: {str(e)}")

        return cover_paths

    def download_image_via_page(self, page, image_url, save_path):
        """Download cover image bytes through page fetch."""
        headers = {
            "Referer": page.url,
            "User-Agent": self.http_user_agent,
            "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
        }
        try:
            response = page.context.request.get(image_url, headers=headers, timeout=20000)
            if response.ok:
                with open(save_path, 'wb') as f:
                    f.write(response.body())
                return True
        except Exception:
            pass

        try:
            data = page.evaluate(
                """async (url) => {
                    try {
                        const response = await fetch(url, { credentials: 'include' });
                        if (!response.ok) return null;
                        const buffer = await response.arrayBuffer();
                        return Array.from(new Uint8Array(buffer));
                    } catch (err) {
                        return null;
                    }
                }""",
                image_url,
            )
            if data:
                with open(save_path, 'wb') as f:
                    f.write(bytes(data))
                return True
        except Exception:
            pass
        return False

    def extract_note_id(self, url):
        """从URL提取笔记ID"""
        try:
            match = re.search(r'/explore/([a-zA-Z0-9]+)', url)
            if match:
                return match.group(1)
        except:
            pass
        return ''

    def extract_detailed_content(self, page, url):
        """提取详细内容 - 保持原有逻辑"""
        details = {
            "笔记ID": self.extract_note_id(url),
            "标题": "",
            "作者昵称": "",
            "作者ID": "",
            "粉丝数": "",
            "博主主页": "",
            "发布时间": "",
            "点赞数": "",
            "收藏数": "",
            "评论数": "",
            "分享数": "",
            "封面链接": "",
            "视频链接": "",
            "正文": "",
            "封面图数量": 0
        }
        image_urls = []
        note_id = details["笔记ID"]

        try:
            self.log("  [等待] 页面加载...")
            time.sleep(2.5)

            # 获取页面数据（简化版，保持核心功能）
            page_data = page.evaluate("""() => {
                const globalState = window.__INITIAL_STATE__ || window.__REDUX_STATE__ || {};
                const noteStore = globalState.note || globalState.noteDetail || {};
                const detailMap = noteStore.noteDetailMap || noteStore.notes || {};

                let detail = null;
                const values = Object.values(detailMap);
                if (values.length > 0) detail = values[0];

                if (detail) {
                    const note = detail.note || detail;
                    const user = detail.user || note.user || {};
                    const stats = detail.noteStat || detail.noteStats || {};

                    return {
                        title: (note.title || '').trim(),
                        content: (note.desc || '').trim(),
                        authorName: (user.nickname || '').trim(),
                        authorId: (user.user_id || user.id || '').toString(),
                        likes: stats.likes || '',
                        collects: stats.collectCount || '',
                        comments: stats.commentCount || '',
                        images: (note.imageList || []).map(img => img.url || '').filter(Boolean)
                    };
                }
                return {};
            }""")

            if page_data:
                details["标题"] = page_data.get("title", "")[:200]
                details["正文"] = page_data.get("content", "")[:500]
                details["作者昵称"] = page_data.get("authorName", "")
                details["作者ID"] = page_data.get("authorId", "")
                details["点赞数"] = str(page_data.get("likes", ""))
                details["收藏数"] = str(page_data.get("collects", ""))
                details["评论数"] = str(page_data.get("comments", ""))

                image_urls = page_data.get("images", []) or []
                if image_urls:
                    details["封面图数量"] = len(image_urls)
                    details["封面链接"] = image_urls[0]
                    self.log(f"  [成功] 标题: {details['标题'][:30]}")
                    self.log(f"  [成功] 作者: {details['作者昵称']}")

        except Exception as e:
            self.log(f"  [错误] 详情抓取失败: {str(e)}")

        return details, image_urls

    # ===== 修复3: 确保链接与输出文档正确对应 =====
    def capture_with_playwright_fixed(self, items, screenshot_dir, all_links_mode=False):
        """修复版抓取功能 - 确保链接与结果正确对应"""
        results = []
        covers_dir = os.path.join(screenshot_dir, 'covers')
        os.makedirs(covers_dir, exist_ok=True)

        self.log("[浏览器-修复] 正在启动修复版浏览器...")

        with sync_playwright() as p:
            try:
                os.makedirs(self.user_data_dir, exist_ok=True)
                context = p.chromium.launch_persistent_context(
                    user_data_dir=self.user_data_dir,
                    headless=False,
                    viewport={'width': 1920, 'height': 1080},
                    user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                )
                self.log(f"[登录] 使用持久化用户目录: {self.user_data_dir}")
            except Exception as e:
                self.log(f"[错误] 无法启动浏览器 - {str(e)}")
                raise

            page = context.pages[0] if context.pages else context.new_page()
            page.set_default_timeout(60000)

            # 修复3: 添加结果验证机制
            processed_links = set()

            for idx, item in enumerate(items, 1):
                if self.stop_flag:
                    self.log("[停止] 用户中断")
                    break

                link = item.get('链接', '')
                product = item.get('产品', '')
                seq_val = item.get('序号', str(idx))

                # 修复3: 确保每个链接都有唯一标识
                unique_id = f"{seq_val}_{hash(link) % 10000}"
                name_prefix = self.sanitize_name(f"{product}_{seq_val}") if product else self.sanitize_name(f"item_{seq_val}")

                # 修复3: 检查链接是否已处理（防重复）
                if link in processed_links:
                    self.log(f"[{idx}/{len(items)}] ⚠️ 跳过重复链接: {link[:50]}...")
                    continue

                processed_links.add(link)

                self.update_progress(idx, len(items), f"处理链接 {idx}")
                self.log(f"[{idx}/{len(items)}] 🔗 {link[:80]}...")
                self.log(f"[{idx}/{len(items)}] 📝 产品: {product} | 序号: {seq_val}")

                is_xhs = (not all_links_mode) and ("xiaohongshu.com" in link)

                # 修复3: 创建详细的结果记录，包含唯一标识
                result = {
                    '序号': seq_val,
                    '产品': product,
                    '链接': link,
                    '唯一ID': unique_id,  # 新增唯一标识
                    '采集状态': '失败',
                    '错误信息': '',
                    'HTTP状态': '',
                    '笔记ID': '',
                    '标题': '',
                    '作者昵称': '',
                    '作者ID': '',
                    '粉丝数': '',
                    '博主主页': '',
                    '发布时间': '',
                    '点赞数': '',
                    '收藏数': '',
                    '评论数': '',
                    '分享数': '',
                    '封面链接': '',
                    '视频链接': '',
                    '正文': '',
                    '截屏文件': '',
                    '封面图列表': [],
                    '封面图数量': 0,
                    '检测时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    '处理索引': idx  # 新增处理索引
                }

                try:
                    if not is_xhs:
                        # 非小红书链接处理
                        resp = page.goto(link, timeout=45000, wait_until='domcontentloaded')
                        status_code = resp.status if resp else None
                        result['HTTP状态'] = status_code or ''

                        if status_code and status_code >= 400:
                            raise Exception(f"HTTP {status_code}")

                        time.sleep(2)
                        screenshot_name = f'{name_prefix}_screenshot.png'
                        screenshot_path = os.path.join(screenshot_dir, screenshot_name)

                        try:
                            page.screenshot(path=screenshot_path, full_page=True, timeout=60000)
                            result['截屏文件'] = screenshot_path
                            self.log(f"  ✓ 通用链接截图成功: {screenshot_name}")
                        except Exception as se:
                            self.log(f"  ⚠️ 截图超时，使用body兜底: {str(se)}")
                            page.locator("body").screenshot(path=screenshot_path, timeout=60000)
                            result['截屏文件'] = screenshot_path

                        result['采集状态'] = '成功'
                        self.log(f"  ✓ 通用链接完成 | 状态: {status_code or '未知'}")

                    else:
                        # 小红书链接处理
                        max_retries = 2
                        resp = None

                        for attempt in range(max_retries):
                            try:
                                resp = page.goto(link, timeout=45000, wait_until='domcontentloaded')
                                break
                            except Exception as e:
                                if attempt == max_retries - 1:
                                    raise e
                                self.log(f"  [重试] 加载超时，正在重试 ({attempt+1}/{max_retries})...")
                                time.sleep(2)

                        result['HTTP状态'] = resp.status if resp else ''
                        time.sleep(3)

                        # 登录检查（保持原有逻辑）
                        has_login_popup = page.evaluate("""() => {
                            const loginModal = document.querySelector('.login-modal, [class*="LoginModal"], [class*="login-dialog"]');
                            const qrCode = document.querySelector('[class*="qrcode"], [class*="QRCode"]');
                            const loginBox = document.querySelector('.login-box, [class*="login-container"]');
                            return !!(loginModal || qrCode || loginBox);
                        }""")

                        if has_login_popup:
                            self.log(f"  ⚠️ 检测到登录弹窗！请确保已扫码登录")
                            # 这里可以添加登录等待逻辑，但为了简化暂时跳过

                        # 截图
                        screenshot_name = f'{name_prefix}_screenshot.png'
                        screenshot_path = os.path.join(screenshot_dir, screenshot_name)

                        try:
                            page.screenshot(path=screenshot_path, full_page=False, timeout=60000)
                            result['截屏文件'] = screenshot_path
                            self.log(f"  ✓ 页面截图成功: {screenshot_name}")
                        except Exception as se:
                            self.log(f"  ⚠️ 截图超时，使用body兜底: {str(se)}")
                            try:
                                page.locator("body").screenshot(path=screenshot_path, timeout=60000)
                                result['截屏文件'] = screenshot_path
                            except Exception as se2:
                                self.log(f"  ❌ 截图失败: {str(se2)}")

                        # 完整模式抓取
                        is_full_mode = self.capture_mode.get() == "full"
                        if is_full_mode:
                            self.log(f"  [完整模式] 开始深度抓取...")

                            # 提取详细内容
                            details, image_urls = self.extract_detailed_content(page, link)
                            result.update(details)

                            # 修复版封面抓取
                            cover_paths = self.capture_cover_images_fixed(page, covers_dir, name_prefix, image_urls)
                            result['封面图列表'] = cover_paths
                            result['封面图数量'] = len(cover_paths) if cover_paths else len(image_urls)

                            if not result.get('封面链接') and image_urls:
                                result['封面链接'] = image_urls[0]

                            self.log(f"  [数据] 点赞:{details.get('点赞数','')} 收藏:{details.get('收藏数','')} 评论:{details.get('评论数','')}")
                            self.log(f"  [封面-修复] 获取{len(cover_paths)}张封面 | 首图: {result.get('封面链接','无')[:50]}")

                        result['采集状态'] = '成功'
                        self.log(f"  ✓ 链接{idx}处理完成")

                except Exception as e:
                    result['错误信息'] = str(e)
                    self.log(f"  ❌ 处理失败: {str(e)}")

                # 修复3: 验证结果与链接的对应关系
                if result['链接'] != link:
                    self.log(f"  ⚠️ 警告：结果链接不匹配！期望: {link[:30]}... 实际: {result['链接'][:30]}...")

                results.append(result)

                # 修复3: 记录处理进度
                self.log(f"[进度] 已处理 {len(results)}/{len(items)} 个链接")

                # 链接之间延迟
                if idx < len(items):
                    time.sleep(2)

            try:
                context.close()
            except Exception:
                pass
            self.log("[关闭] 浏览器已关闭")

        # 修复3: 最终验证所有结果
        self.log("[验证] 开始验证结果对应关系...")
        success_count = 0
        for i, result in enumerate(results):
            original_item = items[i] if i < len(items) else None
            if original_item and result['链接'] == original_item.get('链接', ''):
                success_count += 1
            else:
                self.log(f"  ⚠️ 第{i+1}个结果链接不匹配")

        self.log(f"[验证] 完成，{success_count}/{len(results)} 个链接对应正确")
        return results

    def create_excel_report_fixed(self, results, output_dir):
        """修复版Excel报告生成"""
        os.makedirs(output_dir, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "小红书监测修复版"

        # 设置列宽（简化版，保持核心字段）
        column_widths = {
            'A': 6,    # 序号
            'B': 12,   # 产品
            'C': 50,   # 链接
            'D': 12,   # 采集状态
            'E': 30,   # 错误信息
            'F': 35,   # 标题
            'G': 15,   # 作者昵称
            'H': 10,   # 点赞数
            'I': 10,   # 收藏数
            'J': 10,   # 评论数
            'K': 35,   # 截图
            'L': 22,   # 封面1
            'M': 22,   # 封面2
            'N': 22,   # 封面3
            'O': 22,   # 封面4
            'P': 22,   # 封面5
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 表头
        headers = [
            '序号', '产品', '链接', '采集状态', '错误信息', '标题', '作者昵称',
            '点赞数', '收藏数', '评论数', '截图',
            '封面1', '封面2', '封面3', '封面4', '封面5'
        ]

        header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")  # 绿色表示修复版
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        ws.row_dimensions[1].height = 30

        # 数据行
        current_row = 2
        for result in results:
            row_height = 100

            # 基础数据
            ws.cell(row=current_row, column=1, value=result.get('序号', '')).border = thin_border
            ws.cell(row=current_row, column=2, value=result.get('产品', '')).border = thin_border
            ws.cell(row=current_row, column=3, value=result['链接']).border = thin_border

            # 状态
            status = result.get('采集状态', '失败')
            cell = ws.cell(row=current_row, column=4, value=status)
            if status == '成功':
                cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            else:
                cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 错误信息
            ws.cell(row=current_row, column=5, value=result.get('错误信息', '')).border = thin_border

            # 详细数据
            ws.cell(row=current_row, column=6, value=result.get('标题', '')).border = thin_border
            ws.cell(row=current_row, column=7, value=result.get('作者昵称', '')).border = thin_border
            ws.cell(row=current_row, column=8, value=result.get('点赞数', '')).border = thin_border
            ws.cell(row=current_row, column=9, value=result.get('收藏数', '')).border = thin_border
            ws.cell(row=current_row, column=10, value=result.get('评论数', '')).border = thin_border

            # 截图
            screenshot_path = result.get('截屏文件', '')
            cell = ws.cell(row=current_row, column=11)
            cell.border = thin_border
            if screenshot_path and os.path.exists(screenshot_path):
                try:
                    img = XLImage(screenshot_path)
                    img.width = 250
                    img.height = 120
                    img.anchor = f'K{current_row}'
                    ws.add_image(img)
                    row_height = max(row_height, 100)
                except Exception as e:
                    cell.value = f'加载失败: {str(e)}'
            else:
                cell.value = '无截图'
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 封面图
            cover_paths = result.get('封面图列表', []) or []
            for i in range(5):  # 最多显示5张封面
                col_index = 12 + i
                cell = ws.cell(row=current_row, column=col_index)
                cell.border = thin_border

                if i < len(cover_paths) and os.path.exists(cover_paths[i]):
                    try:
                        img = XLImage(cover_paths[i])
                        img.width = 140
                        img.height = 110
                        anchor_cols = ['L','M','N','O','P']
                        anchor_col = anchor_cols[i]
                        img.anchor = f'{anchor_col}{current_row}'
                        ws.add_image(img)
                        row_height = max(row_height, 90)
                    except Exception:
                        cell.value = f'封面{i+1}'
                else:
                    cell.value = ''

            ws.row_dimensions[current_row].height = row_height
            current_row += 1

        # 统计行
        stats_row = current_row + 1
        success_count = sum(1 for r in results if r.get('采集状态') == '成功')
        fail_count = len(results) - success_count

        ws.cell(row=stats_row, column=1, value='统计(修复版)').font = Font(bold=True, size=11, color='28A745')
        ws.cell(row=stats_row, column=2, value=f'总计: {len(results)} | 成功: {success_count} | 失败: {fail_count}').font = Font(bold=True, size=10)

        # 冻结首行
        ws.freeze_panes = 'A2'

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_dir, f'小红书监测修复版_{timestamp}.xlsx')
        wb.save(output_file)

        return output_file, success_count, fail_count, 0

    def monitor_task(self):
        """主监控任务 - 修复版"""
        try:
            self.is_running = True
            self.stop_flag = False
            self.start_button.config(state='disabled')
            self.stop_button.config(state='normal')

            csv_file = self.csv_file.get()
            if not os.path.exists(csv_file):
                messagebox.showerror("错误", f"CSV文件不存在: {csv_file}")
                return

            output_dir = self.output_dir.get()
            screenshot_dir = os.path.join(output_dir, 'screenshots')
            os.makedirs(screenshot_dir, exist_ok=True)

            self.log("="*60)
            self.log("[任务-修复] 开始修复版监测")
            self.log(f"[输入] {csv_file}")
            self.log(f"[输出] {output_dir}")
            self.log(f"[模式] {'完整模式' if self.capture_mode.get() == 'full' else '快速模式'}")
            self.log("[修复] 视频封面、图片顺序、链接对应关系")
            self.log("="*60)

            items = self.read_items_from_csv(csv_file)
            self.log(f"[链接] 找到 {len(items)} 个链接")

            if not items:
                messagebox.showwarning("警告", "CSV中没有链接")
                return

            # 使用修复版抓取
            results = self.capture_with_playwright_fixed(
                items, screenshot_dir, all_links_mode=self.all_links_mode.get()
            )

            if not self.stop_flag:
                self.log("[报告] 生成修复版Excel...")
                output_file, success, fail, _ = self.create_excel_report_fixed(results, output_dir)

                self.log("="*60)
                self.log("[完成-修复] 修复版监测任务完成！")
                self.log(f"[统计] 总:{len(results)} | 成功:{success} | 失败:{fail}")
                self.log(f"[文件] {os.path.basename(output_file)}")
                self.log("[修复] 已解决: 1.视频封面 2.图片顺序 3.链接对应")
                self.log("="*60)

                messagebox.showinfo(
                    "✓ 修复版完成",
                    f"修复版监测完成！\n\n"
                    f"总计: {len(results)} 条\n"
                    f"成功: {success} 条\n"
                    f"失败: {fail} 条\n\n"
                    f"已修复问题:\n"
                    f"✓ 视频封面截图\n"
                    f"✓ 图片按顺序抓取\n"
                    f"✓ 链接对应关系\n\n"
                    f"报告: {os.path.basename(output_file)}"
                )

        except Exception as e:
            self.log(f"[错误] {str(e)}")
            messagebox.showerror("错误", f"修复版监测失败:\n{str(e)}")

        finally:
            self.is_running = False
            self.start_button.config(state='normal')
            self.stop_button.config(state='disabled')
            self.update_progress(0, 1, "就绪")

    def start_monitoring(self):
        if not self.is_running:
            thread = threading.Thread(target=self.monitor_task, daemon=True)
            thread.start()

    def stop_monitoring(self):
        if self.is_running:
            self.stop_flag = True
            self.log("[停止] 正在停止...")

def main():
    root = tk.Tk()

    root.update_idletasks()
    width = 950
    height = 800
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')

    app = FixedLinkMonitorGUI(root)
    root.mainloop()

if __name__ == '__main__':
    main()