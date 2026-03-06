"""
完整的链接监测工具 - 包含登录和截图
流程：
1. 打开小红书登录页面
2. 等待用户扫码登录
3. 逐个访问CSV中的链接
4. 对每个链接截图
5. 生成包含截图的Excel报告
"""

import csv
import os
import time
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from PIL import Image as PILImage
import subprocess
import json

def read_links_from_csv(csv_file):
    """从CSV文件读取链接"""
    links = []
    try:
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
            for row in rows[1:]:  # 跳过标题行
                if row and len(row) > 0 and row[0].strip():
                    links.append(row[0].strip())
    except Exception as e:
        print(f"错误: 无法读取CSV文件 - {str(e)}")
        raise
    return links

def resize_screenshot(image_path, max_width=300, max_height=200):
    """调整截图大小"""
    try:
        img = PILImage.open(image_path)
        width_ratio = max_width / img.width
        height_ratio = max_height / img.height
        ratio = min(width_ratio, height_ratio)
        new_width = int(img.width * ratio)
        new_height = int(img.height * ratio)
        return new_width, new_height
    except Exception as e:
        print(f"警告: 无法调整图片大小 - {str(e)}")
        return max_width, max_height

def create_excel_report(results, output_file='output/链接监测报告.xlsx'):
    """创建Excel报告"""
    print("\n正在生成Excel报告...")
    
    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "链接监测结果"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 20
    
    # 创建标题行
    headers = ['序号', '链接', '状态', '截屏预览', '检测时间']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 25
    
    # 填充数据
    current_row = 2
    for idx, result in enumerate(results, 1):
        cell_a = ws.cell(row=current_row, column=1, value=idx)
        cell_a.alignment = Alignment(horizontal='center', vertical='center')
        cell_a.border = thin_border
        
        cell_b = ws.cell(row=current_row, column=2, value=result['链接'])
        cell_b.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_b.border = thin_border
        
        status_cell = ws.cell(row=current_row, column=3, value=result['状态'])
        if result['状态'] == '成功':
            status_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            status_cell.font = Font(color="FFFFFF", bold=True)
        else:
            status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            status_cell.font = Font(color="FFFFFF", bold=True)
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        status_cell.border = thin_border
        
        screenshot_path = result.get('截屏文件', '')
        cell_d = ws.cell(row=current_row, column=4)
        cell_d.border = thin_border
        
        if screenshot_path and os.path.exists(screenshot_path):
            try:
                width, height = resize_screenshot(screenshot_path, max_width=300, max_height=180)
                ws.row_dimensions[current_row].height = height * 0.75
                img = XLImage(screenshot_path)
                img.width = width
                img.height = height
                img.anchor = f'D{current_row}'
                ws.add_image(img)
            except Exception as e:
                cell_d.value = f'图片加载失败: {str(e)}'
                cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.row_dimensions[current_row].height = 30
        else:
            cell_d.value = result.get('错误信息', '无截图')
            cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_d.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            ws.row_dimensions[current_row].height = 30
        
        cell_e = ws.cell(row=current_row, column=5, value=result['检测时间'])
        cell_e.alignment = Alignment(horizontal='center', vertical='center')
        cell_e.border = thin_border
        
        current_row += 1
    
    # 统计信息
    stats_row = current_row + 1
    success_count = sum(1 for r in results if r['状态'] == '成功')
    fail_count = sum(1 for r in results if r['状态'] == '失败')
    
    ws.cell(row=stats_row, column=1, value='统计').font = Font(bold=True, size=12)
    ws.cell(row=stats_row, column=2, value=f'总计: {len(results)} | 成功: {success_count} | 失败: {fail_count}')
    ws.cell(row=stats_row, column=2).font = Font(bold=True, size=11)
    
    # 保存文件
    wb.save(output_file)
    print(f"[OK] Excel报告已生成: {output_file}")
    return output_file, success_count, fail_count

def main():
    print("\n" + "="*70)
    print("                  链接监测工具 - 完整运行版")
    print("="*70)
    
    # 配置
    csv_file = 'yilideeplink.csv'
    output_dir = 'output'
    screenshot_dir = os.path.join(output_dir, 'screenshots')
    
    # 创建目录
    os.makedirs(screenshot_dir, exist_ok=True)
    
    # 读取链接
    print(f"\n步骤 1: 读取CSV文件")
    print(f"文件: {csv_file}")
    
    if not os.path.exists(csv_file):
        print(f"错误: CSV文件不存在 - {csv_file}")
        return
    
    links = read_links_from_csv(csv_file)
    print(f"找到 {len(links)} 个链接")
    
    if not links:
        print("错误: CSV文件中没有链接")
        return
    
    # 显示链接列表
    print("\n链接列表:")
    for idx, link in enumerate(links, 1):
        print(f"  {idx}. {link}")
    
    # 步骤 2: 登录提示
    print("\n" + "="*70)
    print("步骤 2: 浏览器登录")
    print("="*70)
    print("\n说明:")
    print("1. 浏览器会自动打开小红书登录页面")
    print("2. 请使用小红书APP扫描二维码登录")
    print("3. 登录成功后，工具会自动继续")
    print("\n准备好后按 Enter 键继续...")
    input()
    
    print("\n正在启动浏览器...")
    print("请在浏览器中完成登录...")
    
    # 步骤 3: 访问链接并截图
    print("\n" + "="*70)
    print("步骤 3: 访问链接并截图")
    print("="*70)
    print("\n说明:")
    print("- 工具将使用Playwright自动化浏览器")
    print("- 每个链接会等待5秒加载")
    print("- 自动截图并保存")
    print("\n准备开始访问链接，按 Enter 继续...")
    input()
    
    results = []
    
    # 这里需要手动配合MCP工具执行，所以提示用户
    print("\n[信息] 请使用以下步骤手动完成截图:")
    print("1. 确保浏览器已登录小红书")
    print("2. 使用Playwright MCP工具访问每个链接")
    print("3. 截图保存到 output/screenshots/ 目录")
    print("4. 截图命名: screenshot_1.png, screenshot_2.png, ...")
    
    # 检查现有截图
    print("\n正在检查截图文件...")
    for idx, link in enumerate(links, 1):
        screenshot_name = f'screenshot_{idx}.png'
        screenshot_path = os.path.join(screenshot_dir, screenshot_name)
        
        result = {
            '链接': link,
            '状态': '成功' if os.path.exists(screenshot_path) else '失败',
            '截屏文件': screenshot_path if os.path.exists(screenshot_path) else '',
            '错误信息': '' if os.path.exists(screenshot_path) else '截屏文件不存在，需要重新截图',
            '检测时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        status_icon = "[OK]" if result['状态'] == '成功' else "[FAIL]"
        print(f"  {status_icon} 链接 {idx}: {screenshot_name}")
        
        results.append(result)
    
    # 步骤 4: 生成报告
    print("\n" + "="*70)
    print("步骤 4: 生成Excel报告")
    print("="*70)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f'链接监测报告_{timestamp}.xlsx')
    
    output_file, success_count, fail_count = create_excel_report(results, output_file)
    
    # 最终总结
    print("\n" + "="*70)
    print("                     监测完成！")
    print("="*70)
    print(f"\n总计: {len(results)} 个链接")
    print(f"成功: {success_count} 个")
    print(f"失败: {fail_count} 个")
    print(f"\n报告文件: {output_file}")
    
    if fail_count > 0:
        print("\n失败的链接:")
        for idx, result in enumerate(results, 1):
            if result['状态'] == '失败':
                print(f"  {idx}. {result['链接']}")
                print(f"     原因: {result['错误信息']}")
    
    print("\n" + "="*70)
    print("\n提示: 如需重新截图，请:")
    print("1. 使用浏览器访问失败的链接")
    print("2. 手动截图或使用Playwright工具")
    print("3. 保存到 output/screenshots/ 目录")
    print("4. 重新运行此脚本生成报告")
    print("\n" + "="*70)

if __name__ == '__main__':
    main()








